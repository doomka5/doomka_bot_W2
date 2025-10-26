"""Telegram bot with basic user management backed by PostgreSQL."""

from __future__ import annotations

import asyncio
import logging
import os
import subprocess
from io import BytesIO
from pathlib import Path
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any, Awaitable, Callable, Dict, Optional, Tuple

import asyncpg
from aiogram import BaseMiddleware, Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    KeyboardButton,
    Message,
    TelegramObject,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    BufferedInputFile,
)
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO)

# === –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –æ–∫—Ä—É–∂–µ–Ω–∏—è ===
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN environment variable is not set")

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "botdb")
DB_USER = os.getenv("DB_USER", "botuser")
DB_PASS = os.getenv("DB_PASS", "botpass")

def _resolve_update_script_path() -> Path:
    env_path = os.getenv("UPDATE_SCRIPT_PATH")
    if env_path:
        return Path(env_path)

    default_path = Path("/share/3D/doomka_bot_W2/update.sh")
    if default_path.exists():
        return default_path

    return Path(__file__).resolve().parent.parent / "update.sh"


UPDATE_SCRIPT_PATH = _resolve_update_script_path()

db_pool: Optional[asyncpg.Pool] = None

WARSAW_TZ = ZoneInfo("Europe/Warsaw")


# === –ü—Ä–æ–≤–µ—Ä–∫–∞ –¥–æ—Å—Ç—É–ø–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π ===
async def user_has_access(tg_id: int) -> bool:
    if db_pool is None:
        logging.warning("Database pool is not initialised when checking access")
        return False
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT 1 FROM users WHERE tg_id = $1", tg_id)
    return row is not None


async def user_is_admin(tg_id: int) -> bool:
    if db_pool is None:
        logging.warning("Database pool is not initialised when checking admin role")
        return False
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT role FROM users WHERE tg_id = $1", tg_id)
    if not row:
        return False
    role = (row["role"] or "").lower()
    return "–∞–¥–º–∏–Ω–∏—Å—Ç" in role or "admin" in role


async def ensure_admin_access(message: Message, state: Optional[FSMContext] = None) -> bool:
    if not message.from_user:
        return False
    if await user_is_admin(message.from_user.id):
        return True
    if state is not None:
        await state.clear()
    await message.answer("üö´ –£ –≤–∞—Å –Ω–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –ø—Ä–∞–≤ –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –Ω–∞—Å—Ç—Ä–æ–π–∫–∞–º–∏.", reply_markup=MAIN_MENU_KB)
    return False


# === –ú–∏–¥–ª–≤–∞—Ä—å –¥–æ—Å—Ç—É–ø–∞ ===
class AccessControlMiddleware(BaseMiddleware):
    async def __call__(
        self,
        handler: Callable[[TelegramObject, Dict[str, Any]], Awaitable[Any]],
        event: TelegramObject,
        data: Dict[str, Any],
    ) -> Any:
        user_id: Optional[int] = None
        if isinstance(event, Message) and event.from_user:
            user_id = event.from_user.id
        if user_id is None:
            return await handler(event, data)
        if await user_has_access(user_id):
            return await handler(event, data)
        if isinstance(event, Message):
            await event.answer("üö´ –£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–º—É –±–æ—Ç—É. –û–±—Ä–∞—Ç–∏—Ç–µ—Å—å –∫ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä—É.")
        return None


# === –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö ===
async def init_database() -> None:
    global db_pool
    db_pool = await asyncpg.create_pool(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS, database=DB_NAME
    )

    async with db_pool.acquire() as conn:
        async with conn.transaction():
            # –¢–∞–±–ª–∏—Ü–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    tg_id BIGINT UNIQUE NOT NULL,
                    username TEXT NOT NULL,
                    position TEXT NOT NULL,
                    role TEXT NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            # –¢–∞–±–ª–∏—Ü–∞ —Å–∫–ª–∞–¥–∞ –ø–ª–∞—Å—Ç–∏–∫–æ–≤
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS warehouse_plastics (
                    id SERIAL PRIMARY KEY,
                    article TEXT NOT NULL,
                    material TEXT,
                    thickness NUMERIC(10, 2),
                    color TEXT,
                    length NUMERIC(10, 2),
                    width NUMERIC(10, 2),
                    warehouse TEXT,
                    comment TEXT,
                    employee_id BIGINT,
                    employee_name TEXT,
                    arrival_date DATE,
                    arrival_at TIMESTAMPTZ
                )
                """
            )
            await conn.execute(
                """
                ALTER TABLE warehouse_plastics
                ADD COLUMN IF NOT EXISTS arrival_at TIMESTAMPTZ
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS written_off_plastics (
                    id SERIAL PRIMARY KEY,
                    source_id INTEGER,
                    article TEXT NOT NULL,
                    material TEXT,
                    thickness NUMERIC(10, 2),
                    color TEXT,
                    length NUMERIC(10, 2),
                    width NUMERIC(10, 2),
                    warehouse TEXT,
                    comment TEXT,
                    employee_id BIGINT,
                    employee_name TEXT,
                    arrival_date DATE,
                    arrival_at TIMESTAMPTZ,
                    project TEXT,
                    written_off_by_id BIGINT,
                    written_off_by_name TEXT,
                    written_off_at TIMESTAMPTZ NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                ALTER TABLE written_off_plastics
                ADD COLUMN IF NOT EXISTS written_off_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                """
            )
            # –¢–∞–±–ª–∏—Ü–∞ —Ç–∏–ø–æ–≤ –ø–ª–∞—Å—Ç–∏–∫–æ–≤
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS plastic_material_types (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS plastic_material_thicknesses (
                    id SERIAL PRIMARY KEY,
                    material_id INTEGER NOT NULL REFERENCES plastic_material_types(id) ON DELETE CASCADE,
                    thickness NUMERIC(10, 2) NOT NULL,
                    UNIQUE(material_id, thickness)
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS plastic_material_colors (
                    id SERIAL PRIMARY KEY,
                    material_id INTEGER NOT NULL REFERENCES plastic_material_types(id) ON DELETE CASCADE,
                    color TEXT NOT NULL,
                    UNIQUE(material_id, color)
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS plastic_storage_locations (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS film_manufacturers (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_manufacturers (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_storage_locations (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_series (
                    id SERIAL PRIMARY KEY,
                    manufacturer_id INTEGER NOT NULL REFERENCES led_module_manufacturers(id) ON DELETE CASCADE,
                    name TEXT NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now()),
                    UNIQUE(manufacturer_id, name)
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_colors (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_power_options (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_voltage_options (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_module_lens_counts (
                    id SERIAL PRIMARY KEY,
                    value INTEGER UNIQUE NOT NULL CHECK (value > 0),
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS generated_led_modules (
                    id SERIAL PRIMARY KEY,
                    article TEXT UNIQUE NOT NULL,
                    manufacturer_id INTEGER NOT NULL REFERENCES led_module_manufacturers(id) ON DELETE RESTRICT,
                    series_id INTEGER NOT NULL REFERENCES led_module_series(id) ON DELETE RESTRICT,
                    color_id INTEGER NOT NULL REFERENCES led_module_colors(id) ON DELETE RESTRICT,
                    lens_count_id INTEGER NOT NULL REFERENCES led_module_lens_counts(id) ON DELETE RESTRICT,
                    power_option_id INTEGER NOT NULL REFERENCES led_module_power_options(id) ON DELETE RESTRICT,
                    voltage_option_id INTEGER NOT NULL REFERENCES led_module_voltage_options(id) ON DELETE RESTRICT,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS warehouse_led_modules (
                    id SERIAL PRIMARY KEY,
                    led_module_id INTEGER NOT NULL REFERENCES generated_led_modules(id) ON DELETE RESTRICT,
                    article TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    added_by_id BIGINT,
                    added_by_name TEXT,
                    added_at TIMESTAMPTZ NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                ALTER TABLE warehouse_led_modules
                DROP CONSTRAINT IF EXISTS warehouse_led_modules_quantity_check
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS written_off_led_modules (
                    id SERIAL PRIMARY KEY,
                    led_module_id INTEGER NOT NULL REFERENCES generated_led_modules(id) ON DELETE RESTRICT,
                    article TEXT NOT NULL,
                    quantity INTEGER NOT NULL CHECK (quantity > 0),
                    project TEXT,
                    written_off_by_id BIGINT,
                    written_off_by_name TEXT,
                    written_off_at TIMESTAMPTZ NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS led_strip_manufacturers (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS power_supply_manufacturers (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS film_series (
                    id SERIAL PRIMARY KEY,
                    manufacturer_id INTEGER NOT NULL REFERENCES film_manufacturers(id) ON DELETE CASCADE,
                    name TEXT NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now()),
                    UNIQUE(manufacturer_id, name)
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS film_storage_locations (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS order_types (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS clients (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    phone TEXT,
                    contact_person TEXT,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS client_addresses (
                    id SERIAL PRIMARY KEY,
                    client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                    address TEXT,
                    google_maps_link TEXT,
                    created_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                CREATE SEQUENCE IF NOT EXISTS order_number_seq START 1
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS orders (
                    id SERIAL PRIMARY KEY,
                    order_number INTEGER NOT NULL UNIQUE DEFAULT nextval('order_number_seq'),
                    client_id INTEGER NOT NULL REFERENCES clients(id),
                    client_name TEXT NOT NULL,
                    title TEXT NOT NULL,
                    order_type TEXT NOT NULL,
                    folder_path TEXT NOT NULL,
                    due_date DATE NOT NULL,
                    is_urgent BOOLEAN NOT NULL DEFAULT FALSE,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT timezone('utc', now()),
                    created_by_id BIGINT,
                    created_by_name TEXT
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS warehouse_films (
                    id SERIAL PRIMARY KEY,
                    article TEXT NOT NULL,
                    manufacturer TEXT,
                    series TEXT,
                    color_code TEXT,
                    color TEXT,
                    width NUMERIC(10, 2),
                    length NUMERIC(10, 2),
                    warehouse TEXT,
                    comment TEXT,
                    employee_id BIGINT,
                    employee_nick TEXT,
                    recorded_at TIMESTAMPTZ
                )
                """
            )
            await conn.execute(
                """
                CREATE TABLE IF NOT EXISTS written_off_films (
                    id SERIAL PRIMARY KEY,
                    source_id INTEGER,
                    article TEXT NOT NULL,
                    manufacturer TEXT,
                    series TEXT,
                    color_code TEXT,
                    color TEXT,
                    width NUMERIC(10, 2),
                    length NUMERIC(10, 2),
                    warehouse TEXT,
                    comment TEXT,
                    employee_id BIGINT,
                    employee_nick TEXT,
                    recorded_at TIMESTAMPTZ,
                    project TEXT,
                    written_off_by_id BIGINT,
                    written_off_by_name TEXT,
                    written_off_at TIMESTAMPTZ NOT NULL DEFAULT timezone('utc', now())
                )
                """
            )
            await conn.execute(
                """
                ALTER TABLE written_off_films
                ADD COLUMN IF NOT EXISTS written_off_at TIMESTAMPTZ DEFAULT timezone('utc', now())
                """
            )
            # –î–æ–±–∞–≤–ª—è–µ–º –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞
            await conn.execute(
                """
                INSERT INTO users (tg_id, username, position, role)
                VALUES ($1, $2, $3, $4)
                ON CONFLICT (tg_id) DO UPDATE
                SET username = EXCLUDED.username,
                    position = EXCLUDED.position,
                    role = EXCLUDED.role
                """,
                37352491,
                "DooMka",
                "–ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä",
                "–∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä —Å –ø–æ–ª–Ω—ã–º–∏ –ø—Ä–∞–≤–∞–º–∏ –∏ –¥–æ—Å—Ç—É–ø–æ–º",
            )


async def close_database() -> None:
    global db_pool
    if db_pool:
        await db_pool.close()
        db_pool = None


# === –°–æ–±—ã—Ç–∏—è –∑–∞–ø—É—Å–∫–∞ –∏ –æ—Å—Ç–∞–Ω–æ–≤–∫–∏ ===
async def on_startup(bot: Bot) -> None:
    await init_database()
    logging.info("‚úÖ –ë–æ—Ç –∑–∞–ø—É—â–µ–Ω –∏ –ø–æ–¥–∫–ª—é—á—ë–Ω –∫ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö.")
    print("‚úÖ –ë–æ—Ç –∑–∞–ø—É—â–µ–Ω –∏ –ø–æ–¥–∫–ª—é—á—ë–Ω –∫ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö.")


async def on_shutdown(bot: Bot) -> None:
    await close_database()


# === Dispatcher ===
dp = Dispatcher()
dp.startup.register(on_startup)
dp.shutdown.register(on_shutdown)
dp.message.outer_middleware(AccessControlMiddleware())


# === FSM ===
class AddUserStates(StatesGroup):
    waiting_for_tg_id = State()
    waiting_for_username = State()
    waiting_for_position = State()
    waiting_for_role = State()
    waiting_for_created_at = State()


class ManageOrderTypeStates(StatesGroup):
    waiting_for_new_type_name = State()


class ManagePlasticMaterialStates(StatesGroup):
    waiting_for_new_material_name = State()
    waiting_for_material_name_to_delete = State()
    waiting_for_material_name_to_add_thickness = State()
    waiting_for_thickness_value_to_add = State()
    waiting_for_material_name_to_delete_thickness = State()
    waiting_for_thickness_value_to_delete = State()
    waiting_for_material_name_to_add_color = State()
    waiting_for_color_value_to_add = State()
    waiting_for_material_name_to_delete_color = State()
    waiting_for_color_value_to_delete = State()
    waiting_for_new_storage_location_name = State()
    waiting_for_storage_location_to_delete = State()


class ManageFilmManufacturerStates(StatesGroup):
    waiting_for_new_manufacturer_name = State()
    waiting_for_manufacturer_name_to_delete = State()


class ManageFilmSeriesStates(StatesGroup):
    waiting_for_manufacturer_for_new_series = State()
    waiting_for_new_series_name = State()
    waiting_for_manufacturer_for_series_deletion = State()
    waiting_for_series_name_to_delete = State()


class ManageFilmStorageStates(StatesGroup):
    waiting_for_new_storage_location_name = State()
    waiting_for_storage_location_to_delete = State()


class ManageLedModuleManufacturerStates(StatesGroup):
    waiting_for_new_manufacturer_name = State()
    waiting_for_manufacturer_name_to_delete = State()


class ManageLedModuleSeriesStates(StatesGroup):
    waiting_for_manufacturer_for_new_series = State()
    waiting_for_new_series_name = State()
    waiting_for_manufacturer_for_series_deletion = State()
    waiting_for_series_name_to_delete = State()


class ManageLedModuleStorageStates(StatesGroup):
    waiting_for_new_storage_location_name = State()
    waiting_for_storage_location_to_delete = State()


class ManageLedModuleLensStates(StatesGroup):
    waiting_for_new_lens_count = State()
    waiting_for_lens_count_to_delete = State()


class ManageLedModuleColorStates(StatesGroup):
    waiting_for_new_color_name = State()
    waiting_for_color_name_to_delete = State()


class ManageLedModulePowerStates(StatesGroup):
    waiting_for_new_power_value = State()
    waiting_for_power_value_to_delete = State()


class ManageLedModuleVoltageStates(StatesGroup):
    waiting_for_new_voltage_value = State()
    waiting_for_voltage_value_to_delete = State()


class GenerateLedModuleStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_manufacturer = State()
    waiting_for_series = State()
    waiting_for_color = State()
    waiting_for_lens_count = State()
    waiting_for_power = State()
    waiting_for_voltage = State()


class ManageLedStripManufacturerStates(StatesGroup):
    waiting_for_new_manufacturer_name = State()
    waiting_for_manufacturer_name_to_delete = State()


class ManagePowerSupplyManufacturerStates(StatesGroup):
    waiting_for_new_manufacturer_name = State()
    waiting_for_manufacturer_name_to_delete = State()


class AddWarehouseFilmStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_manufacturer = State()
    waiting_for_series = State()
    waiting_for_color_code = State()
    waiting_for_color = State()
    waiting_for_width = State()
    waiting_for_length = State()
    waiting_for_storage = State()
    waiting_for_comment = State()


class CommentWarehouseFilmStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_comment = State()


class MoveWarehouseFilmStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_new_location = State()


class WriteOffWarehouseFilmStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_project = State()


class SearchWarehouseFilmStates(StatesGroup):
    choosing_mode = State()
    waiting_for_article = State()
    waiting_for_number = State()
    waiting_for_color = State()


class AddWarehousePlasticStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_material = State()
    waiting_for_thickness = State()
    waiting_for_color = State()
    waiting_for_length = State()
    waiting_for_width = State()
    waiting_for_storage = State()
    waiting_for_comment = State()


class AddWarehousePlasticBatchStates(StatesGroup):
    waiting_for_quantity = State()
    waiting_for_material = State()
    waiting_for_thickness = State()
    waiting_for_color = State()
    waiting_for_length = State()
    waiting_for_width = State()
    waiting_for_storage = State()
    waiting_for_comment = State()


class AddWarehouseLedModuleStates(StatesGroup):
    waiting_for_module = State()
    waiting_for_quantity = State()


class WriteOffWarehouseLedModuleStates(StatesGroup):
    waiting_for_module = State()
    waiting_for_quantity = State()
    waiting_for_project = State()


class SearchWarehousePlasticStates(StatesGroup):
    choosing_mode = State()
    waiting_for_article = State()
    waiting_for_material = State()
    waiting_for_thickness = State()
    waiting_for_color = State()
    waiting_for_min_length = State()
    waiting_for_min_width = State()


class AddClientStates(StatesGroup):
    waiting_for_name = State()
    waiting_for_phone = State()
    waiting_for_contact_person = State()
    waiting_for_address = State()
    waiting_for_map_link = State()


class SearchClientStates(StatesGroup):
    waiting_for_query = State()


class CreateOrderStates(StatesGroup):
    waiting_for_client_query = State()
    waiting_for_client_selection = State()
    waiting_for_order_name = State()
    waiting_for_order_type = State()
    waiting_for_folder_path = State()
    waiting_for_due_date = State()
    waiting_for_urgency = State()


class CommentWarehousePlasticStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_comment = State()


class MoveWarehousePlasticStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_new_location = State()


class WriteOffWarehousePlasticStates(StatesGroup):
    waiting_for_article = State()
    waiting_for_project = State()


# === –ö–ª–∞–≤–∏–∞—Ç—É—Ä—ã ===
MAIN_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="üè¢ –°–∫–ª–∞–¥"),
            KeyboardButton(text="‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏"),
        ],
        [
            KeyboardButton(text="–ö–ª–∏–µ–Ω—Ç—ã"),
            KeyboardButton(text="–ó–∞–∫–∞–∑—ã"),
        ],
    ],
    resize_keyboard=True,
)

ORDERS_NEW_ORDER_TEXT = "üÜï –ù–æ–≤—ã–π –∑–∞–∫–∞–∑"
ORDERS_IN_PROGRESS_TEXT = "üìã –ó–∞–∫–∞–∑—ã –≤ —Ä–∞–±–æ—Ç–µ"
ORDERS_SETTINGS_TEXT = "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –∑–∞–∫–∞–∑–æ–≤"
ORDERS_SETTINGS_ORDER_TYPE_TEXT = "üóÇÔ∏è –¢–∏–ø –∑–∞–∫–∞–∑–∞"
ORDERS_SETTINGS_BACK_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –∑–∞–∫–∞–∑–∞–º"
ORDER_TYPE_ADD_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ç–∏–ø –∑–∞–∫–∞–∑–∞"
ORDER_TYPE_DELETE_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å —Ç–∏–ø –∑–∞–∫–∞–∑–∞"
ORDER_TYPE_BACK_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞–º –∑–∞–∫–∞–∑–æ–≤"

ORDERS_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=ORDERS_NEW_ORDER_TEXT)],
        [KeyboardButton(text=ORDERS_IN_PROGRESS_TEXT)],
        [KeyboardButton(text=ORDERS_SETTINGS_TEXT)],
        [KeyboardButton(text="‚¨ÖÔ∏è –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")],
    ],
    resize_keyboard=True,
)

ORDERS_SETTINGS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=ORDERS_SETTINGS_ORDER_TYPE_TEXT)],
        [KeyboardButton(text=ORDERS_SETTINGS_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

ORDERS_ORDER_TYPE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=ORDER_TYPE_ADD_TEXT)],
        [KeyboardButton(text=ORDER_TYPE_DELETE_TEXT)],
        [KeyboardButton(text=ORDER_TYPE_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

CLIENTS_ADD_CLIENT_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –∫–ª–∏–µ–Ω—Ç–∞"
CLIENTS_SEARCH_CLIENT_TEXT = "üîç –ü–æ–∏—Å–∫ –∫–ª–∏–µ–Ω—Ç–∞"

CLIENTS_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=CLIENTS_ADD_CLIENT_TEXT)],
        [KeyboardButton(text=CLIENTS_SEARCH_CLIENT_TEXT)],
        [KeyboardButton(text="‚¨ÖÔ∏è –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")],
    ],
    resize_keyboard=True,
)

SETTINGS_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="üë• –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏")],
        [KeyboardButton(text="üîÑ –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")],
    ],
    resize_keyboard=True,
)

USERS_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è")],
        [KeyboardButton(text="üìã –ü–æ—Å–º–æ—Ç—Ä–µ—Ç—å –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_ELECTRICS_TEXT = "‚ö° –≠–ª–µ–∫—Ç—Ä–∏–∫–∞"
WAREHOUSE_ELECTRICS_LED_STRIPS_TEXT = "üí° Led –ª–µ–Ω—Ç–∞"
WAREHOUSE_ELECTRICS_LED_MODULES_TEXT = "üß© Led –º–æ–¥—É–ª–∏"
WAREHOUSE_ELECTRICS_POWER_SUPPLIES_TEXT = "üîå –ë–ª–æ–∫–∏ –ø–∏—Ç–∞–Ω–∏—è"
WAREHOUSE_LED_MODULES_ADD_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å Led –º–æ–¥—É–ª–∏"
WAREHOUSE_LED_MODULES_STOCK_TEXT = "üì¶ –û—Å—Ç–∞—Ç–æ–∫ Led –º–æ–¥—É–ª–µ–π –Ω–∞ —Å–∫–ª–∞–¥–µ"
WAREHOUSE_LED_MODULES_WRITE_OFF_TEXT = "‚ûñ –°–ø–∏—Å–∞—Ç—å Led –º–æ–¥—É–ª–∏"
WAREHOUSE_LED_MODULES_BACK_TO_ELECTRICS_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Ä–∞–∑–¥–µ–ª—É ¬´–≠–ª–µ–∫—Ç—Ä–∏–∫–∞¬ª"

WAREHOUSE_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="üß± –ü–ª–∞—Å—Ç–∏–∫–∏")],
        [KeyboardButton(text="üéûÔ∏è –ü–ª–µ–Ω–∫–∏")],
        [KeyboardButton(text=WAREHOUSE_ELECTRICS_TEXT)],
        [KeyboardButton(text="‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_ELECTRICS_TEXT = "‚ö° –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚öôÔ∏è"
WAREHOUSE_SETTINGS_ELECTRICS_LED_STRIPS_TEXT = "üí° Led –ª–µ–Ω—Ç–∞ ‚öôÔ∏è"
WAREHOUSE_SETTINGS_ELECTRICS_LED_MODULES_TEXT = "üß© Led –º–æ–¥—É–ª–∏ ‚öôÔ∏è"
LED_MODULES_BASE_MENU_TEXT = "Led –º–æ–¥—É–ª–∏ baza"
WAREHOUSE_SETTINGS_ELECTRICS_POWER_SUPPLIES_TEXT = "üîå –ë–ª–æ–∫–∏ –ø–∏—Ç–∞–Ω–∏—è ‚öôÔ∏è"
WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —ç–ª–µ–∫—Ç—Ä–∏–∫–µ"

WAREHOUSE_SETTINGS_MENU_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="üß± –ü–ª–∞—Å—Ç–∏–∫")],
        [KeyboardButton(text="üéûÔ∏è –ü–ª–µ–Ω–∫–∏ ‚öôÔ∏è")],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_ELECTRICS_TEXT)],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_PLASTIC_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="üì¶ –ú–∞—Ç–µ—Ä–∏–∞–ª")],
        [KeyboardButton(text="üìè –¢–æ–ª—â–∏–Ω–∞")],
        [KeyboardButton(text="üé® –¶–≤–µ—Ç")],
        [KeyboardButton(text="üè∑Ô∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_FILM_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="üè≠ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å")],
        [KeyboardButton(text="üé¨ –°–µ—Ä–∏—è")],
        [KeyboardButton(text="üè¨ –°–∫–ª–∞–¥")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_ELECTRICS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=WAREHOUSE_SETTINGS_ELECTRICS_LED_STRIPS_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_ELECTRICS_LED_MODULES_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_ELECTRICS_POWER_SUPPLIES_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT)],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

LED_MODULES_MANUFACTURERS_MENU_TEXT = "üè≠ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_SERIES_MENU_TEXT = "üé¨ –°–µ—Ä–∏—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_STORAGE_MENU_TEXT = "üè¨ –ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_POWER_MENU_TEXT = "‚ö° –ú–æ—â–Ω–æ—Å—Ç—å –º–æ–¥—É–ª–µ–π"
LED_MODULES_LENS_MENU_TEXT = "üî¢ –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑"
LED_MODULES_COLORS_MENU_TEXT = "üé® –¶–≤–µ—Ç –º–æ–¥—É–ª–µ–π"
LED_MODULES_VOLTAGE_MENU_TEXT = "üîå –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ –º–æ–¥—É–ª–µ–π"
LED_MODULES_BACK_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ Led –º–æ–¥—É–ª—è–º"
LED_MODULES_GENERATE_TEXT = "–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å Led –º–æ–¥—É–ª—å"
LED_MODULES_DELETE_TEXT = "–£–¥–∞–ª–∏—Ç—å Led –º–æ–¥—É–ª—å"
LED_MODULES_ADD_MANUFACTURER_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_MANUFACTURER_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_ADD_SERIES_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_SERIES_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_ADD_STORAGE_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_STORAGE_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è Led –º–æ–¥—É–ª–µ–π"
LED_MODULES_ADD_POWER_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–æ—â–Ω–æ—Å—Ç—å –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_POWER_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –º–æ—â–Ω–æ—Å—Ç—å –º–æ–¥—É–ª–µ–π"
LED_MODULES_ADD_VOLTAGE_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_VOLTAGE_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ –º–æ–¥—É–ª–µ–π"
LED_MODULES_ADD_LENS_COUNT_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑"
LED_MODULES_REMOVE_LENS_COUNT_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑"
LED_MODULES_ADD_COLOR_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ü–≤–µ—Ç –º–æ–¥—É–ª–µ–π"
LED_MODULES_REMOVE_COLOR_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å —Ü–≤–µ—Ç –º–æ–¥—É–ª–µ–π"
LED_STRIPS_ADD_MANUFACTURER_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –ª–µ–Ω—Ç—ã"
LED_STRIPS_REMOVE_MANUFACTURER_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –ª–µ–Ω—Ç—ã"
POWER_SUPPLIES_ADD_MANUFACTURER_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –±–ª–æ–∫–æ–≤ –ø–∏—Ç–∞–Ω–∏—è"
POWER_SUPPLIES_REMOVE_MANUFACTURER_TEXT = "‚ûñ –£–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –±–ª–æ–∫–æ–≤ –ø–∏—Ç–∞–Ω–∏—è"

WAREHOUSE_SETTINGS_LED_MODULES_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_MANUFACTURERS_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_SERIES_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_STORAGE_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_BASE_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_COLORS_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_POWER_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_VOLTAGE_MENU_TEXT)],
        [KeyboardButton(text=LED_MODULES_LENS_MENU_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_GENERATE_TEXT)],
        [KeyboardButton(text=LED_MODULES_DELETE_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_MANUFACTURERS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_MANUFACTURER_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_MANUFACTURER_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_SERIES_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_SERIES_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)


WAREHOUSE_SETTINGS_LED_MODULES_STORAGE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_STORAGE_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_STORAGE_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)


WAREHOUSE_SETTINGS_LED_MODULES_COLORS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_COLOR_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_COLOR_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_POWER_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_POWER_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_POWER_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_VOLTAGE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_VOLTAGE_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_VOLTAGE_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_MODULES_LENS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_MODULES_ADD_LENS_COUNT_TEXT)],
        [KeyboardButton(text=LED_MODULES_REMOVE_LENS_COUNT_TEXT)],
        [KeyboardButton(text=LED_MODULES_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_LED_STRIPS_MANUFACTURERS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=LED_STRIPS_ADD_MANUFACTURER_TEXT)],
        [KeyboardButton(text=LED_STRIPS_REMOVE_MANUFACTURER_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_POWER_SUPPLIES_MANUFACTURERS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=POWER_SUPPLIES_ADD_MANUFACTURER_TEXT)],
        [KeyboardButton(text=POWER_SUPPLIES_REMOVE_MANUFACTURER_TEXT)],
        [KeyboardButton(text=WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_PLASTIC_MATERIALS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–∞—Ç–µ—Ä–∏–∞–ª")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å –º–∞—Ç–µ—Ä–∏–∞–ª")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–∞—Å—Ç–∏–∫—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–∞—Å—Ç–∏–∫—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ü–≤–µ—Ç")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å —Ü–≤–µ—Ç")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–∞—Å—Ç–∏–∫—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_PLASTIC_STORAGE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–∞—Å—Ç–∏–∫—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_FILM_MANUFACTURERS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–µ–Ω–∫–∞–º")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_FILM_SERIES_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–µ–Ω–∫–∞–º")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_SETTINGS_FILM_STORAGE_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–ª–µ–Ω–∫–∏")],
        [KeyboardButton(text="‚ûñ –£–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–ª–µ–Ω–∫–∏")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–µ–Ω–∫–∞–º")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_FILMS_ADD_TEXT = "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø–ª–µ–Ω–∫—É"
WAREHOUSE_FILMS_WRITE_OFF_TEXT = "‚ûñ –°–ø–∏—Å–∞—Ç—å –ø–ª–µ–Ω–∫—É"
WAREHOUSE_FILMS_COMMENT_TEXT = "üí¨ –ö–æ–º–º–µ–Ω—Ç–∏—Ä–æ–≤–∞—Ç—å –ø–ª–µ–Ω–∫—É"
WAREHOUSE_FILMS_MOVE_TEXT = "üîÅ –ü–µ—Ä–µ–º–µ—Å—Ç–∏—Ç—å –ø–ª–µ–Ω–∫—É"
WAREHOUSE_FILMS_SEARCH_TEXT = "üîç –ù–∞–π—Ç–∏ –ø–ª–µ–Ω–∫—É"
WAREHOUSE_FILMS_EXPORT_TEXT = "üì§ –≠–∫—Å–ø–æ—Ä—Ç –ø–ª–µ–Ω–æ–∫"

WAREHOUSE_FILMS_SEARCH_BY_ARTICLE_TEXT = "–ü–æ –∞—Ä—Ç–∏–∫—É–ª—É"
WAREHOUSE_FILMS_SEARCH_BY_NUMBER_TEXT = "–ü–æ –Ω–æ–º–µ—Ä—É"
WAREHOUSE_FILMS_SEARCH_BY_COLOR_TEXT = "–ü–æ —Ü–≤–µ—Ç—É"
WAREHOUSE_FILMS_SEARCH_BACK_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–µ–Ω–∫–∞–º"
FILM_SEARCH_RESULTS_LIMIT = 15

WAREHOUSE_FILMS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text=WAREHOUSE_FILMS_ADD_TEXT),
            KeyboardButton(text=WAREHOUSE_FILMS_WRITE_OFF_TEXT),
        ],
        [
            KeyboardButton(text=WAREHOUSE_FILMS_COMMENT_TEXT),
            KeyboardButton(text=WAREHOUSE_FILMS_MOVE_TEXT),
        ],
        [
            KeyboardButton(text=WAREHOUSE_FILMS_SEARCH_TEXT),
            KeyboardButton(text=WAREHOUSE_FILMS_EXPORT_TEXT),
        ],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_FILMS_SEARCH_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=WAREHOUSE_FILMS_SEARCH_BY_ARTICLE_TEXT)],
        [KeyboardButton(text=WAREHOUSE_FILMS_SEARCH_BY_NUMBER_TEXT)],
        [KeyboardButton(text=WAREHOUSE_FILMS_SEARCH_BY_COLOR_TEXT)],
        [KeyboardButton(text=WAREHOUSE_FILMS_SEARCH_BACK_TEXT)],
    ],
    resize_keyboard=True,
)

WAREHOUSE_PLASTICS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å"), KeyboardButton(text="++–¥–æ–±–∞–≤–∏—Ç—å –ø–∞—á–∫—É")],
        [KeyboardButton(text="‚ûñ –°–ø–∏—Å–∞—Ç—å"), KeyboardButton(text="üí¨ –ö–æ–º–º–µ–Ω—Ç–∏—Ä–æ–≤–∞—Ç—å")],
        [KeyboardButton(text="üîÅ –ü–µ—Ä–µ–º–µ—Å—Ç–∏—Ç—å"), KeyboardButton(text="üîç –ù–∞–π—Ç–∏")],
        [KeyboardButton(text="üì§ –≠–∫—Å–ø–æ—Ä—Ç")],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_ELECTRICS_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=WAREHOUSE_ELECTRICS_LED_STRIPS_TEXT)],
        [KeyboardButton(text=WAREHOUSE_ELECTRICS_LED_MODULES_TEXT)],
        [KeyboardButton(text=WAREHOUSE_ELECTRICS_POWER_SUPPLIES_TEXT)],
        [KeyboardButton(text="‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")],
    ],
    resize_keyboard=True,
)

WAREHOUSE_LED_MODULES_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=WAREHOUSE_LED_MODULES_ADD_TEXT)],
        [KeyboardButton(text=WAREHOUSE_LED_MODULES_STOCK_TEXT)],
        [KeyboardButton(text=WAREHOUSE_LED_MODULES_WRITE_OFF_TEXT)],
        [KeyboardButton(text=WAREHOUSE_LED_MODULES_BACK_TO_ELECTRICS_TEXT)],
    ],
    resize_keyboard=True,
)

SEARCH_BY_ARTICLE_TEXT = "üî¢ –ü–æ–∏—Å–∫ –ø–æ –∞—Ä—Ç–∏–∫—É–ª—É"
ADVANCED_SEARCH_TEXT = "üß≠ –†–∞—Å—à–∏—Ä–µ–Ω–Ω—ã–π –ø–æ–∏—Å–∫"
BACK_TO_PLASTICS_MENU_TEXT = "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –º–µ–Ω—é –ø–ª–∞—Å—Ç–∏–∫–∞"

WAREHOUSE_PLASTICS_SEARCH_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text=SEARCH_BY_ARTICLE_TEXT)],
        [KeyboardButton(text=ADVANCED_SEARCH_TEXT)],
        [KeyboardButton(text=BACK_TO_PLASTICS_MENU_TEXT)],
    ],
    resize_keyboard=True,
)

ADVANCED_SEARCH_SKIP_MATERIAL_TEXT = "‚û°Ô∏è –î–∞–ª–µ–µ"
ADVANCED_SEARCH_ALL_THICKNESSES_TEXT = "üìè –í—Å–µ —Ç–æ–ª—â–∏–Ω—ã"
ADVANCED_SEARCH_ALL_COLORS_TEXT = "üé® –í—Å–µ —Ü–≤–µ—Ç–∞"

CANCEL_TEXT = "‚ùå –û—Ç–º–µ–Ω–∞"
SKIP_TEXT = "–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å"

def build_article_input_keyboard(
    suggested_article: Optional[str] = None,
) -> ReplyKeyboardMarkup:
    keyboard: list[list[KeyboardButton]] = []
    if suggested_article:
        keyboard.append([KeyboardButton(text=suggested_article)])
    keyboard.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)


CANCEL_KB = build_article_input_keyboard()

SKIP_OR_CANCEL_KB = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text=SKIP_TEXT)], [KeyboardButton(text=CANCEL_TEXT)]],
    resize_keyboard=True,
)

ORDER_URGENCY_YES_TEXT = "–î–∞"
ORDER_URGENCY_NO_TEXT = "–ù–µ—Ç"

ORDER_URGENCY_KB = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text=ORDER_URGENCY_YES_TEXT),
            KeyboardButton(text=ORDER_URGENCY_NO_TEXT),
        ],
        [KeyboardButton(text=CANCEL_TEXT)],
    ],
    resize_keyboard=True,
)


async def _process_cancel_if_requested(message: Message, state: FSMContext) -> bool:
    if (message.text or "").strip() != CANCEL_TEXT:
        return False
    await handle_cancel(message, state)
    return True


async def _cancel_add_user_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=USERS_MENU_KB
    )


async def _cancel_add_client_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=CLIENTS_MENU_KB
    )


async def _cancel_search_client_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –ü–æ–∏—Å–∫ –∫–ª–∏–µ–Ω—Ç–∞ –æ—Ç–º–µ–Ω—ë–Ω.", reply_markup=CLIENTS_MENU_KB
    )


async def _cancel_create_order_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –°–æ–∑–¥–∞–Ω–∏–µ –∑–∞–∫–∞–∑–∞ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=ORDERS_MENU_KB
    )


async def _cancel_add_plastic_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –ø–ª–∞—Å—Ç–∏–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_PLASTICS_KB
    )


async def _cancel_add_plastic_batch_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –ø–∞—á–∫–∏ –ø–ª–∞—Å—Ç–∏–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_PLASTICS_KB
    )


async def _cancel_add_led_module_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ Led –º–æ–¥—É–ª—è –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_LED_MODULES_KB
    )


async def _cancel_write_off_led_module_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –°–ø–∏—Å–∞–Ω–∏–µ Led –º–æ–¥—É–ª–µ–π –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_LED_MODULES_KB
    )


async def _cancel_search_plastic_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚ùå –ü–æ–∏—Å–∫ –æ—Ç–º–µ–Ω—ë–Ω.", reply_markup=WAREHOUSE_PLASTICS_KB)


async def _cancel_search_film_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚ùå –ü–æ–∏—Å–∫ –æ—Ç–º–µ–Ω—ë–Ω.", reply_markup=WAREHOUSE_FILMS_KB)


async def _cancel_comment_plastic_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –ò–∑–º–µ–Ω–µ–Ω–∏–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏—è –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_PLASTICS_KB
    )


async def _cancel_move_plastic_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚ùå –ü–µ—Ä–µ–º–µ—â–µ–Ω–∏–µ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_PLASTICS_KB)


async def _cancel_write_off_plastic_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚ùå –°–ø–∏—Å–∞–Ω–∏–µ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_PLASTICS_KB)


async def _cancel_add_film_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –ø–ª–µ–Ω–∫–∏ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_FILMS_KB
    )


async def _cancel_comment_film_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –ò–∑–º–µ–Ω–µ–Ω–∏–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏—è –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_FILMS_KB
    )


async def _cancel_move_film_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –ü–µ—Ä–µ–º–µ—â–µ–Ω–∏–µ –ø–ª–µ–Ω–∫–∏ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_FILMS_KB
    )


async def _cancel_write_off_film_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚ùå –°–ø–∏—Å–∞–Ω–∏–µ –ø–ª–µ–Ω–∫–∏ –æ—Ç–º–µ–Ω–µ–Ω–æ.", reply_markup=WAREHOUSE_FILMS_KB)


async def _cancel_generate_led_module_flow(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ùå –ì–µ–Ω–µ—Ä–∞—Ü–∏—è Led –º–æ–¥—É–ª—è –æ—Ç–º–µ–Ω–µ–Ω–∞.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
    )


# === –†–∞–±–æ—Ç–∞ —Å –ë–î ===
async def upsert_user_in_db(
    tg_id: int,
    username: str,
    position: str,
    role: str,
    created_at: Optional[datetime] = None,
) -> None:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        await conn.execute(
            """
            INSERT INTO users (tg_id, username, position, role, created_at)
            VALUES ($1, $2, $3, $4, COALESCE($5, timezone('utc', now())))
            ON CONFLICT (tg_id) DO UPDATE
            SET username = EXCLUDED.username,
                position = EXCLUDED.position,
                role = EXCLUDED.role,
                created_at = CASE
                    WHEN $5 IS NULL THEN users.created_at
                    ELSE EXCLUDED.created_at
                END
            """,
            tg_id,
            username,
            position,
            role,
            created_at,
        )


async def fetch_all_users_from_db() -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT tg_id, username, position, role, created_at
            FROM users
            ORDER BY created_at DESC NULLS LAST, id DESC
            """
        )
    return [dict(row) for row in rows]


async def create_client_in_db(
    name: str,
    phone: Optional[str] = None,
    contact_person: Optional[str] = None,
) -> int:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO clients (name, phone, contact_person)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            name,
            phone,
            contact_person,
        )
    return int(row["id"])


async def add_client_address_in_db(
    client_id: int,
    address: Optional[str] = None,
    google_maps_link: Optional[str] = None,
) -> int:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO client_addresses (client_id, address, google_maps_link)
            VALUES ($1, $2, $3)
            RETURNING id
            """,
            client_id,
            address,
            google_maps_link,
        )
    return int(row["id"])


async def search_clients_by_name(
    query: str,
    limit: int = 10,
) -> Tuple[list[Dict[str, Any]], bool]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        client_rows = await conn.fetch(
            """
            SELECT id, name, phone, contact_person
            FROM clients
            WHERE name ILIKE $1
            ORDER BY LOWER(name), id
            LIMIT $2
            """,
            f"%{query}%",
            limit + 1,
        )
        has_more = len(client_rows) > limit
        trimmed_rows = list(client_rows[:limit])
        if not trimmed_rows:
            return [], False
        client_ids = [row["id"] for row in trimmed_rows]
        address_rows = await conn.fetch(
            """
            SELECT client_id, address, google_maps_link
            FROM client_addresses
            WHERE client_id = ANY($1::int[])
            ORDER BY client_id, created_at DESC NULLS LAST, id DESC
            """,
            client_ids,
        )
    addresses_map: Dict[int, list[Dict[str, Optional[str]]]] = {}
    for row in address_rows:
        address_value = row.get("address")
        map_link_value = row.get("google_maps_link")
        if address_value is None and map_link_value is None:
            continue
        addresses_map.setdefault(row["client_id"], []).append(
            {
                "address": address_value,
                "google_maps_link": map_link_value,
            }
        )
    results: list[Dict[str, Any]] = []
    for row in trimmed_rows:
        results.append(
            {
                "id": row["id"],
                "name": row["name"],
                "phone": row.get("phone"),
                "contact_person": row.get("contact_person"),
                "addresses": addresses_map.get(row["id"], []),
            }
        )
    return results, has_more


async def fetch_order_types() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM order_types ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_next_order_number() -> int:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT COALESCE(MAX(order_number), 0) + 1 AS next_number FROM orders"
        )
    return int(row["next_number"] or 1)


async def fetch_all_orders() -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                order_number,
                client_name,
                title,
                order_type,
                folder_path,
                due_date,
                is_urgent,
                created_at,
                created_by_name
            FROM orders
            ORDER BY due_date ASC, order_number ASC
            """
        )
    return [dict(row) for row in rows]


async def create_order_in_db(
    client_id: int,
    client_name: str,
    title: str,
    order_type: str,
    folder_path: str,
    due_date: date,
    is_urgent: bool,
    created_by_id: Optional[int],
    created_by_name: Optional[str],
) -> Dict[str, Any]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO orders (
                client_id,
                client_name,
                title,
                order_type,
                folder_path,
                due_date,
                is_urgent,
                created_by_id,
                created_by_name
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
            RETURNING id, order_number, client_id, client_name, title, order_type,
                      folder_path, due_date, is_urgent, created_at, created_by_id,
                      created_by_name
            """,
            client_id,
            client_name,
            title,
            order_type,
            folder_path,
            due_date,
            is_urgent,
            created_by_id,
            created_by_name,
        )
    if row is None:
        raise RuntimeError("Failed to insert order")
    return dict(row)


async def fetch_plastic_material_types() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM plastic_material_types ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_plastic_storage_locations() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM plastic_storage_locations ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_film_manufacturers() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM film_manufacturers ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_led_module_manufacturers() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_module_manufacturers ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_led_module_storage_locations() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_module_storage_locations ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_led_strip_manufacturers() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_strip_manufacturers ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_power_supply_manufacturers() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM power_supply_manufacturers ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def get_led_module_manufacturer_by_name(
    name: str,
) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, name
            FROM led_module_manufacturers
            WHERE LOWER(name) = LOWER($1)
            """,
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "name": row["name"]}


async def fetch_led_module_manufacturers_with_series() -> list[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturers_rows = await conn.fetch(
            "SELECT id, name FROM led_module_manufacturers ORDER BY LOWER(name)"
        )
        series_rows = await conn.fetch(
            """
            SELECT manufacturer_id, name
            FROM led_module_series
            ORDER BY manufacturer_id, LOWER(name)
            """
        )
    series_map: dict[int, list[str]] = {}
    for row in series_rows:
        series_map.setdefault(row["manufacturer_id"], []).append(row["name"])
    result: list[dict[str, Any]] = []
    for row in manufacturers_rows:
        result.append(
            {
                "id": row["id"],
                "name": row["name"],
                "series": series_map.get(row["id"], []),
            }
        )
    return result


async def fetch_led_module_colors() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_module_colors ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_led_module_power_options() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_module_power_options ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_led_module_voltage_options() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM led_module_voltage_options ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def fetch_generated_led_modules_with_details() -> list[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                glm.article,
                manufacturer.name AS manufacturer,
                series.name AS series,
                color.name AS color,
                lens.value AS lens_count,
                power.name AS power,
                voltage.name AS voltage
            FROM generated_led_modules AS glm
            JOIN led_module_manufacturers AS manufacturer ON manufacturer.id = glm.manufacturer_id
            JOIN led_module_series AS series ON series.id = glm.series_id
            JOIN led_module_colors AS color ON color.id = glm.color_id
            JOIN led_module_lens_counts AS lens ON lens.id = glm.lens_count_id
            JOIN led_module_power_options AS power ON power.id = glm.power_option_id
            JOIN led_module_voltage_options AS voltage ON voltage.id = glm.voltage_option_id
            ORDER BY glm.created_at DESC NULLS LAST, glm.id DESC
            """
        )
    return [dict(row) for row in rows]


async def fetch_led_module_stock_summary() -> list[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                glm.article,
                manufacturer.name AS manufacturer,
                series.name AS series,
                color.name AS color,
                lens.value AS lens_count,
                power.name AS power,
                voltage.name AS voltage,
                COALESCE(SUM(wlm.quantity), 0) AS total_quantity
            FROM generated_led_modules AS glm
            JOIN led_module_manufacturers AS manufacturer ON manufacturer.id = glm.manufacturer_id
            JOIN led_module_series AS series ON series.id = glm.series_id
            JOIN led_module_colors AS color ON color.id = glm.color_id
            JOIN led_module_lens_counts AS lens ON lens.id = glm.lens_count_id
            JOIN led_module_power_options AS power ON power.id = glm.power_option_id
            JOIN led_module_voltage_options AS voltage ON voltage.id = glm.voltage_option_id
            LEFT JOIN warehouse_led_modules AS wlm ON wlm.led_module_id = glm.id
            GROUP BY
                glm.id,
                glm.article,
                manufacturer.name,
                series.name,
                color.name,
                lens.value,
                power.name,
                voltage.name
            HAVING COALESCE(SUM(wlm.quantity), 0) > 0
            ORDER BY total_quantity DESC, LOWER(glm.article)
            """,
        )
    return [dict(row) for row in rows]


async def get_generated_led_module_details(module_id: int) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT
                glm.id,
                glm.article,
                manufacturer.name AS manufacturer,
                series.name AS series,
                color.name AS color,
                lens.value AS lens_count,
                power.name AS power,
                voltage.name AS voltage
            FROM generated_led_modules AS glm
            JOIN led_module_manufacturers AS manufacturer ON manufacturer.id = glm.manufacturer_id
            JOIN led_module_series AS series ON series.id = glm.series_id
            JOIN led_module_colors AS color ON color.id = glm.color_id
            JOIN led_module_lens_counts AS lens ON lens.id = glm.lens_count_id
            JOIN led_module_power_options AS power ON power.id = glm.power_option_id
            JOIN led_module_voltage_options AS voltage ON voltage.id = glm.voltage_option_id
            WHERE glm.id = $1
            """,
            module_id,
        )
    if row is None:
        return None
    return dict(row)


async def fetch_led_module_lens_counts() -> list[int]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT value FROM led_module_lens_counts ORDER BY value"
        )
    return [row["value"] for row in rows]


async def fetch_led_module_series_by_manufacturer(
    manufacturer_name: str,
) -> list[str]:
    manufacturer = await get_led_module_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        return []
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT name
            FROM led_module_series
            WHERE manufacturer_id = $1
            ORDER BY LOWER(name)
            """,
            manufacturer["id"],
        )
    return [row["name"] for row in rows]


async def get_led_module_series_by_name(
    manufacturer_id: int, name: str
) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, manufacturer_id, name
            FROM led_module_series
            WHERE manufacturer_id = $1 AND LOWER(name) = LOWER($2)
            """,
            manufacturer_id,
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "manufacturer_id": row["manufacturer_id"], "name": row["name"]}


async def get_led_module_color_by_name(name: str) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name FROM led_module_colors WHERE LOWER(name) = LOWER($1)",
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "name": row["name"]}


async def get_led_module_power_option_by_name(name: str) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name FROM led_module_power_options WHERE LOWER(name) = LOWER($1)",
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "name": row["name"]}


async def get_led_module_voltage_option_by_name(name: str) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name FROM led_module_voltage_options WHERE LOWER(name) = LOWER($1)",
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "name": row["name"]}


async def get_led_module_lens_count_by_value(value: int) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, value FROM led_module_lens_counts WHERE value = $1",
            value,
        )
    if row is None:
        return None
    return {"id": row["id"], "value": row["value"]}


async def get_generated_led_module_by_article(article: str) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, article, manufacturer_id, series_id, color_id,
                   lens_count_id, power_option_id, voltage_option_id, created_at
            FROM generated_led_modules
            WHERE LOWER(article) = LOWER($1)
            """,
            article,
        )
    if row is None:
        return None
    return dict(row)


async def insert_generated_led_module(
    *,
    article: str,
    manufacturer_id: int,
    series_id: int,
    color_id: int,
    lens_count_id: int,
    power_option_id: int,
    voltage_option_id: int,
) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO generated_led_modules (
                article,
                manufacturer_id,
                series_id,
                color_id,
                lens_count_id,
                power_option_id,
                voltage_option_id
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7)
            ON CONFLICT (article) DO NOTHING
            RETURNING id, article, manufacturer_id, series_id, color_id,
                      lens_count_id, power_option_id, voltage_option_id, created_at
            """,
            article,
            manufacturer_id,
            series_id,
            color_id,
            lens_count_id,
            power_option_id,
            voltage_option_id,
        )
    if row is None:
        return None
    return dict(row)


async def insert_warehouse_led_module_record(
    *,
    led_module_id: int,
    article: str,
    quantity: int,
    added_by_id: Optional[int],
    added_by_name: Optional[str],
) -> Dict[str, Any]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    added_at = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO warehouse_led_modules (
                led_module_id,
                article,
                quantity,
                added_by_id,
                added_by_name,
                added_at
            )
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING id, led_module_id, article, quantity, added_by_id, added_by_name, added_at
            """,
            led_module_id,
            article,
            quantity,
            added_by_id,
            added_by_name,
            added_at,
        )
    if row is None:
        return {}
    return dict(row)


async def get_led_module_stock_quantity(led_module_id: int) -> int:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        value = await conn.fetchval(
            "SELECT COALESCE(SUM(quantity), 0) FROM warehouse_led_modules WHERE led_module_id = $1",
            led_module_id,
        )
    return int(value or 0)


async def write_off_led_module_stock(
    *,
    led_module_id: int,
    article: str,
    quantity: int,
    project: str,
    written_off_by_id: Optional[int],
    written_off_by_name: Optional[str],
) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    if quantity <= 0:
        raise ValueError("Quantity for write-off must be positive")
    now_warsaw = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        async with conn.transaction():
            available = await conn.fetchval(
                "SELECT COALESCE(SUM(quantity), 0) FROM warehouse_led_modules WHERE led_module_id = $1",
                led_module_id,
            )
            if available is None or int(available) < quantity:
                return None
            ledger_row = await conn.fetchrow(
                """
                INSERT INTO warehouse_led_modules (
                    led_module_id,
                    article,
                    quantity,
                    added_by_id,
                    added_by_name,
                    added_at
                )
                VALUES ($1, $2, $3, $4, $5, $6)
                RETURNING id
                """,
                led_module_id,
                article,
                -quantity,
                written_off_by_id,
                written_off_by_name,
                now_warsaw,
            )
            if ledger_row is None:
                return None
            written_off_row = await conn.fetchrow(
                """
                INSERT INTO written_off_led_modules (
                    led_module_id,
                    article,
                    quantity,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                )
                VALUES ($1, $2, $3, $4, $5, $6, $7)
                RETURNING
                    id,
                    led_module_id,
                    article,
                    quantity,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                """,
                led_module_id,
                article,
                quantity,
                project,
                written_off_by_id,
                written_off_by_name,
                now_warsaw,
            )
    if written_off_row is None:
        return None
    return dict(written_off_row)


async def fetch_film_storage_locations() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT name FROM film_storage_locations ORDER BY LOWER(name)"
        )
    return [row["name"] for row in rows]


async def get_film_manufacturer_by_name(
    name: str,
) -> Optional[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, name FROM film_manufacturers WHERE LOWER(name) = LOWER($1)",
            name,
        )
    if row is None:
        return None
    return {"id": row["id"], "name": row["name"]}


async def fetch_film_manufacturers_with_series() -> list[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturers_rows = await conn.fetch(
            "SELECT id, name FROM film_manufacturers ORDER BY LOWER(name)"
        )
        series_rows = await conn.fetch(
            """
            SELECT manufacturer_id, name
            FROM film_series
            ORDER BY manufacturer_id, LOWER(name)
            """
        )
    series_map: dict[int, list[str]] = {}
    for row in series_rows:
        series_map.setdefault(row["manufacturer_id"], []).append(row["name"])
    result: list[dict[str, Any]] = []
    for row in manufacturers_rows:
        result.append(
            {
                "id": row["id"],
                "name": row["name"],
                "series": series_map.get(row["id"], []),
            }
        )
    return result


async def fetch_film_series_by_manufacturer(manufacturer_name: str) -> list[str]:
    manufacturer = await get_film_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        return []
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT name
            FROM film_series
            WHERE manufacturer_id = $1
            ORDER BY LOWER(name)
            """,
            manufacturer["id"],
        )
    return [row["name"] for row in rows]


async def fetch_max_plastic_article() -> Optional[int]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        value = await conn.fetchval(
            """
            SELECT MAX(article::BIGINT)
            FROM warehouse_plastics
            WHERE article ~ '^[0-9]+$'
            """
        )
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


async def fetch_max_film_article() -> Optional[int]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        value = await conn.fetchval(
            """
            SELECT MAX(article::BIGINT)
            FROM warehouse_films
            WHERE article ~ '^[0-9]+$'
            """
        )
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


async def insert_order_type(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO order_types (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def insert_plastic_material_type(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO plastic_material_types (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_plastic_material_type(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM plastic_material_types WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_plastic_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        existing_id = await conn.fetchval(
            "SELECT id FROM plastic_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
        if existing_id:
            return False
        row = await conn.fetchrow(
            """
            INSERT INTO plastic_storage_locations (name)
            VALUES ($1)
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_plastic_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM plastic_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_film_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO film_manufacturers (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_film_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM film_manufacturers WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_led_module_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_manufacturers (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def insert_led_module_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        existing_id = await conn.fetchval(
            "SELECT id FROM led_module_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
        if existing_id:
            return False
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_storage_locations (name)
            VALUES ($1)
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_led_module_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_manufacturers WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def delete_led_module_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_led_module_color(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        existing = await conn.fetchrow(
            "SELECT 1 FROM led_module_colors WHERE LOWER(name) = LOWER($1)",
            name,
        )
        if existing:
            return False
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_colors (name)
            VALUES ($1)
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_led_module_color(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_colors WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_led_module_lens_count(value: int) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_lens_counts (value)
            VALUES ($1)
            ON CONFLICT (value) DO NOTHING
            RETURNING id
            """,
            value,
        )
    return row is not None


async def insert_led_module_series(
    manufacturer_name: str, series_name: str
) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturer_row = await conn.fetchrow(
            """
            SELECT id, name
            FROM led_module_manufacturers
            WHERE LOWER(name) = LOWER($1)
            """,
            manufacturer_name,
        )
        if manufacturer_row is None:
            return "manufacturer_not_found"
        manufacturer_id = manufacturer_row["id"]
        existing_id = await conn.fetchval(
            """
            SELECT id
            FROM led_module_series
            WHERE manufacturer_id = $1 AND LOWER(name) = LOWER($2)
            """,
            manufacturer_id,
            series_name,
        )
        if existing_id:
            return "already_exists"
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_series (manufacturer_id, name)
            VALUES ($1, $2)
            RETURNING id
            """,
            manufacturer_id,
            series_name,
        )
    return "inserted" if row else "error"


async def delete_led_module_series(
    manufacturer_name: str, series_name: str
) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturer_row = await conn.fetchrow(
            """
            SELECT id
            FROM led_module_manufacturers
            WHERE LOWER(name) = LOWER($1)
            """,
            manufacturer_name,
        )
        if manufacturer_row is None:
            return "manufacturer_not_found"
        result = await conn.execute(
            """
            DELETE FROM led_module_series
            WHERE manufacturer_id = $1 AND LOWER(name) = LOWER($2)
            """,
            manufacturer_row["id"],
            series_name,
        )
    return "deleted" if result.endswith(" 1") else "not_found"


async def delete_led_module_lens_count(value: int) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_lens_counts WHERE value = $1",
            value,
        )
    return result.endswith(" 1")


async def insert_led_module_power_option(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_power_options (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_led_module_power_option(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_power_options WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_led_module_voltage_option(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO led_module_voltage_options (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_led_module_voltage_option(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_module_voltage_options WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_led_strip_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO led_strip_manufacturers (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_led_strip_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM led_strip_manufacturers WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_power_supply_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO power_supply_manufacturers (name)
            VALUES ($1)
            ON CONFLICT (name) DO NOTHING
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_power_supply_manufacturer(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM power_supply_manufacturers WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def insert_film_series(
    manufacturer_name: str, series_name: str
) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturer_row = await conn.fetchrow(
            "SELECT id, name FROM film_manufacturers WHERE LOWER(name) = LOWER($1)",
            manufacturer_name,
        )
        if manufacturer_row is None:
            return "manufacturer_not_found"
        manufacturer_id = manufacturer_row["id"]
        existing_id = await conn.fetchval(
            """
            SELECT id FROM film_series
            WHERE manufacturer_id = $1 AND LOWER(name) = LOWER($2)
            """,
            manufacturer_id,
            series_name,
        )
        if existing_id:
            return "already_exists"
        row = await conn.fetchrow(
            """
            INSERT INTO film_series (manufacturer_id, name)
            VALUES ($1, $2)
            RETURNING id
            """,
            manufacturer_id,
            series_name,
        )
    return "inserted" if row else "error"


async def delete_film_series(
    manufacturer_name: str, series_name: str
) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        manufacturer_row = await conn.fetchrow(
            "SELECT id FROM film_manufacturers WHERE LOWER(name) = LOWER($1)",
            manufacturer_name,
        )
        if manufacturer_row is None:
            return "manufacturer_not_found"
        result = await conn.execute(
            """
            DELETE FROM film_series
            WHERE manufacturer_id = $1 AND LOWER(name) = LOWER($2)
            """,
            manufacturer_row["id"],
            series_name,
        )
    return "deleted" if result.endswith(" 1") else "not_found"


async def insert_film_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        existing_id = await conn.fetchval(
            "SELECT id FROM film_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
        if existing_id:
            return False
        row = await conn.fetchrow(
            """
            INSERT INTO film_storage_locations (name)
            VALUES ($1)
            RETURNING id
            """,
            name,
        )
    return row is not None


async def delete_film_storage_location(name: str) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM film_storage_locations WHERE LOWER(name) = LOWER($1)",
            name,
        )
    return result.endswith(" 1")


async def fetch_materials_with_thicknesses() -> list[dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT p.name,
                   COALESCE(
                       (
                           SELECT ARRAY_AGG(t.thickness ORDER BY t.thickness)
                           FROM plastic_material_thicknesses t
                           WHERE t.material_id = p.id
                       ),
                       ARRAY[]::NUMERIC[]
                   ) AS thicknesses,
                   COALESCE(
                       (
                           SELECT ARRAY_AGG(c.color ORDER BY LOWER(c.color))
                           FROM plastic_material_colors c
                           WHERE c.material_id = p.id
                       ),
                       ARRAY[]::TEXT[]
                   ) AS colors
            FROM plastic_material_types p
            ORDER BY LOWER(p.name)
            """
        )
    return [dict(row) for row in rows]


async def fetch_material_thicknesses(material_name: str) -> list[Decimal]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT t.thickness
            FROM plastic_material_thicknesses t
            JOIN plastic_material_types p ON p.id = t.material_id
            WHERE LOWER(p.name) = LOWER($1)
            ORDER BY t.thickness
            """,
            material_name,
        )
    return [row["thickness"] for row in rows]


async def insert_material_thickness(material_name: str, thickness: Decimal) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        material_id = await conn.fetchval(
            "SELECT id FROM plastic_material_types WHERE LOWER(name) = LOWER($1)",
            material_name,
        )
        if material_id is None:
            return "material_not_found"
        row = await conn.fetchrow(
            """
            INSERT INTO plastic_material_thicknesses (material_id, thickness)
            VALUES ($1, $2)
            ON CONFLICT (material_id, thickness) DO NOTHING
            RETURNING id
            """,
            material_id,
            thickness,
        )
        if row:
            return "added"
        return "exists"


async def delete_material_thickness(material_name: str, thickness: Decimal) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        material_id = await conn.fetchval(
            "SELECT id FROM plastic_material_types WHERE LOWER(name) = LOWER($1)",
            material_name,
        )
        if material_id is None:
            return "material_not_found"
        result = await conn.execute(
            """
            DELETE FROM plastic_material_thicknesses
            WHERE material_id = $1 AND thickness = $2
            """,
            material_id,
            thickness,
        )
    if result.endswith(" 1"):
        return "deleted"
    return "not_found"


async def fetch_material_colors(material_name: str) -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT c.color
            FROM plastic_material_colors c
            JOIN plastic_material_types p ON p.id = c.material_id
            WHERE LOWER(p.name) = LOWER($1)
            ORDER BY LOWER(c.color)
            """,
            material_name,
        )
    return [row["color"] for row in rows]


async def fetch_all_material_thicknesses() -> list[Decimal]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT DISTINCT thickness
            FROM plastic_material_thicknesses
            ORDER BY thickness
            """
        )
    return [row["thickness"] for row in rows]


async def fetch_all_material_colors() -> list[str]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT DISTINCT color
            FROM plastic_material_colors
            ORDER BY LOWER(color)
            """
        )
    return [row["color"] for row in rows]


async def insert_material_color(material_name: str, color: str) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        material_id = await conn.fetchval(
            "SELECT id FROM plastic_material_types WHERE LOWER(name) = LOWER($1)",
            material_name,
        )
        if material_id is None:
            return "material_not_found"
        exists = await conn.fetchval(
            """
            SELECT 1
            FROM plastic_material_colors
            WHERE material_id = $1 AND LOWER(color) = LOWER($2)
            """,
            material_id,
            color,
        )
        if exists:
            return "exists"
        await conn.execute(
            """
            INSERT INTO plastic_material_colors (material_id, color)
            VALUES ($1, $2)
            """,
            material_id,
            color,
        )
    return "added"


async def delete_material_color(material_name: str, color: str) -> str:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        material_id = await conn.fetchval(
            "SELECT id FROM plastic_material_types WHERE LOWER(name) = LOWER($1)",
            material_name,
        )
        if material_id is None:
            return "material_not_found"
        result = await conn.execute(
            """
            DELETE FROM plastic_material_colors
            WHERE material_id = $1 AND LOWER(color) = LOWER($2)
            """,
            material_id,
            color,
        )
    if result.endswith(" 1"):
        return "deleted"
    return "not_found"


async def insert_warehouse_plastic_record(
    article: str,
    material: str,
    thickness: Decimal,
    color: str,
    length_mm: Decimal,
    width_mm: Decimal,
    warehouse: str,
    comment: Optional[str],
    employee_id: Optional[int],
    employee_name: Optional[str],
) -> Dict[str, Any]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    now_warsaw = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO warehouse_plastics (
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_id,
                employee_name,
                arrival_date,
                arrival_at
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
            RETURNING
                id,
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_id,
                employee_name,
                arrival_date,
                arrival_at
            """,
            article,
            material,
            thickness,
            color,
            length_mm,
            width_mm,
            warehouse,
            comment,
            employee_id,
            employee_name,
            now_warsaw.date(),
            now_warsaw,
        )
    if row is None:
        return {}
    return dict(row)


async def insert_warehouse_film_record(
    article: str,
    manufacturer: str,
    series: str,
    color_code: str,
    color: str,
    width_mm: Decimal,
    length_mm: Decimal,
    warehouse: str,
    comment: Optional[str],
    employee_id: Optional[int],
    employee_nick: Optional[str],
) -> Dict[str, Any]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    recorded_at = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO warehouse_films (
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            )
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
            RETURNING
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            """,
            article,
            manufacturer,
            series,
            color_code,
            color,
            width_mm,
            length_mm,
            warehouse,
            comment,
            employee_id,
            employee_nick,
            recorded_at,
        )
    if row is None:
        return {}
    return dict(row)


async def search_warehouse_plastic_records(query: str, limit: int = 5) -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    pattern = f"%{query}%"
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_name,
                arrival_at
            FROM warehouse_plastics
            WHERE article ILIKE $1
               OR material ILIKE $1
               OR color ILIKE $1
               OR warehouse ILIKE $1
               OR comment ILIKE $1
            ORDER BY arrival_at DESC NULLS LAST, id DESC
            LIMIT $2
            """,
            pattern,
            limit,
        )
    return [dict(row) for row in rows]


async def fetch_all_warehouse_plastics() -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_name,
                arrival_date,
                arrival_at
            FROM warehouse_plastics
            ORDER BY arrival_at DESC NULLS LAST, id DESC
            """
        )
    return [dict(row) for row in rows]


async def fetch_all_warehouse_films() -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            FROM warehouse_films
            ORDER BY recorded_at DESC NULLS LAST, id DESC
            """
        )
    return [dict(row) for row in rows]


async def fetch_warehouse_plastic_by_article(article: str) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT
                id,
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_name,
                arrival_at
            FROM warehouse_plastics
            WHERE article = $1
            ORDER BY arrival_at DESC NULLS LAST, id DESC
            LIMIT 1
            """,
            article,
        )
    if row is None:
        return None
    return dict(row)


async def fetch_warehouse_film_by_article(article: str) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            FROM warehouse_films
            WHERE article = $1
            ORDER BY recorded_at DESC NULLS LAST, id DESC
            LIMIT 1
            """,
            article,
        )
    if row is None:
        return None
    return dict(row)


async def write_off_warehouse_film(
    record_id: int,
    project: str,
    written_off_by_id: Optional[int],
    written_off_by_name: Optional[str],
) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    now_warsaw = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        async with conn.transaction():
            original_row = await conn.fetchrow(
                """
                DELETE FROM warehouse_films
                WHERE id = $1
                RETURNING
                    id,
                    article,
                    manufacturer,
                    series,
                    color_code,
                    color,
                    width,
                    length,
                    warehouse,
                    comment,
                    employee_id,
                    employee_nick,
                    recorded_at
                """,
                record_id,
            )
            if original_row is None:
                return None
            inserted_row = await conn.fetchrow(
                """
                INSERT INTO written_off_films (
                    source_id,
                    article,
                    manufacturer,
                    series,
                    color_code,
                    color,
                    width,
                    length,
                    warehouse,
                    comment,
                    employee_id,
                    employee_nick,
                    recorded_at,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                )
                VALUES (
                    $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17
                )
                RETURNING
                    id,
                    source_id,
                    article,
                    manufacturer,
                    series,
                    color_code,
                    color,
                    width,
                    length,
                    warehouse,
                    comment,
                    employee_id,
                    employee_nick,
                    recorded_at,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                """,
                original_row["id"],
                original_row["article"],
                original_row["manufacturer"],
                original_row["series"],
                original_row["color_code"],
                original_row["color"],
                original_row["width"],
                original_row["length"],
                original_row["warehouse"],
                original_row["comment"],
                original_row["employee_id"],
                original_row["employee_nick"],
                original_row["recorded_at"],
                project,
                written_off_by_id,
                written_off_by_name,
                now_warsaw,
            )
    if inserted_row is None:
        return None
    return dict(inserted_row)


async def fetch_warehouse_film_by_id(record_id: int) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            FROM warehouse_films
            WHERE id = $1
            """,
            record_id,
        )
    if row is None:
        return None
    return dict(row)


async def search_warehouse_films_by_color_code(
    color_code: str, limit: int = FILM_SEARCH_RESULTS_LIMIT
) -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            FROM warehouse_films
            WHERE color_code ILIKE '%' || $1 || '%'
            ORDER BY recorded_at DESC NULLS LAST, id DESC
            LIMIT $2
            """,
            color_code,
            limit,
        )
    return [dict(row) for row in rows]


async def search_warehouse_films_by_color(
    color_query: str, limit: int = FILM_SEARCH_RESULTS_LIMIT
) -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            FROM warehouse_films
            WHERE color ILIKE '%' || $1 || '%'
            ORDER BY recorded_at DESC NULLS LAST, id DESC
            LIMIT $2
            """,
            color_query,
            limit,
        )
    return [dict(row) for row in rows]


async def search_warehouse_plastics_advanced(
    material: Optional[str] = None,
    thickness: Optional[Decimal] = None,
    color: Optional[str] = None,
    min_length: Optional[Decimal] = None,
    min_width: Optional[Decimal] = None,
) -> list[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    conditions: list[str] = []
    params: list[Any] = []
    param_index = 1
    if material:
        conditions.append(f"LOWER(material) = LOWER(${param_index})")
        params.append(material)
        param_index += 1
    if thickness is not None:
        conditions.append(f"thickness = ${param_index}")
        params.append(thickness)
        param_index += 1
    if color:
        conditions.append(f"LOWER(color) = LOWER(${param_index})")
        params.append(color)
        param_index += 1
    if min_length is not None:
        conditions.append(f"length >= ${param_index}")
        params.append(min_length)
        param_index += 1
    if min_width is not None:
        conditions.append(f"width >= ${param_index}")
        params.append(min_width)
        param_index += 1
    where_clause = ""
    if conditions:
        where_clause = " WHERE " + " AND ".join(conditions)
    query = (
        """
        SELECT
            id,
            article,
            material,
            thickness,
            color,
            length,
            width,
            warehouse,
            comment,
            employee_name,
            arrival_at
        FROM warehouse_plastics
        """
        + where_clause
        + " ORDER BY length DESC NULLS LAST, width DESC NULLS LAST, arrival_at DESC NULLS LAST, id DESC"
    )
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(query, *params)
    return [dict(row) for row in rows]


async def update_warehouse_plastic_comment(
    record_id: int, comment: Optional[str]
) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            """
            UPDATE warehouse_plastics
            SET comment = $2
            WHERE id = $1
            """,
            record_id,
            comment,
    )
    return result.endswith(" 1")


async def update_warehouse_film_comment(
    record_id: int, comment: Optional[str]
) -> bool:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        result = await conn.execute(
            """
            UPDATE warehouse_films
            SET comment = $2
            WHERE id = $1
            """,
            record_id,
            comment,
        )
    return result.endswith(" 1")


async def update_warehouse_film_location(
    record_id: int,
    new_location: str,
    employee_id: Optional[int],
    employee_nick: Optional[str],
) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE warehouse_films
            SET warehouse = $2,
                employee_id = COALESCE($3, employee_id),
                employee_nick = COALESCE($4, employee_nick)
            WHERE id = $1
            RETURNING
                id,
                article,
                manufacturer,
                series,
                color_code,
                color,
                width,
                length,
                warehouse,
                comment,
                employee_id,
                employee_nick,
                recorded_at
            """,
            record_id,
            new_location,
            employee_id,
            employee_nick,
        )
    if row is None:
        return None
    return dict(row)


async def update_warehouse_plastic_location(
    record_id: int,
    new_location: str,
    employee_id: Optional[int],
    employee_name: Optional[str],
) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE warehouse_plastics
            SET warehouse = $2,
                employee_id = COALESCE($3, employee_id),
                employee_name = COALESCE($4, employee_name)
            WHERE id = $1
            RETURNING
                id,
                article,
                material,
                thickness,
                color,
                length,
                width,
                warehouse,
                comment,
                employee_name,
                arrival_at
            """,
            record_id,
            new_location,
            employee_id,
            employee_name,
        )
    if row is None:
        return None
    return dict(row)


async def write_off_warehouse_plastic(
    record_id: int,
    project: str,
    written_off_by_id: Optional[int],
    written_off_by_name: Optional[str],
) -> Optional[Dict[str, Any]]:
    if db_pool is None:
        raise RuntimeError("Database pool is not initialised")
    now_warsaw = datetime.now(WARSAW_TZ)
    async with db_pool.acquire() as conn:
        async with conn.transaction():
            original_row = await conn.fetchrow(
                """
                DELETE FROM warehouse_plastics
                WHERE id = $1
                RETURNING
                    id,
                    article,
                    material,
                    thickness,
                    color,
                    length,
                    width,
                    warehouse,
                    comment,
                    employee_id,
                    employee_name,
                    arrival_date,
                    arrival_at
                """,
                record_id,
            )
            if original_row is None:
                return None
            inserted_row = await conn.fetchrow(
                """
                INSERT INTO written_off_plastics (
                    source_id,
                    article,
                    material,
                    thickness,
                    color,
                    length,
                    width,
                    warehouse,
                    comment,
                    employee_id,
                    employee_name,
                    arrival_date,
                    arrival_at,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                )
                VALUES (
                    $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17
                )
                RETURNING
                    id,
                    source_id,
                    article,
                    material,
                    thickness,
                    color,
                    length,
                    width,
                    warehouse,
                    comment,
                    employee_id,
                    employee_name,
                    arrival_date,
                    arrival_at,
                    project,
                    written_off_by_id,
                    written_off_by_name,
                    written_off_at
                """,
                original_row["id"],
                original_row["article"],
                original_row["material"],
                original_row["thickness"],
                original_row["color"],
                original_row["length"],
                original_row["width"],
                original_row["warehouse"],
                original_row["comment"],
                original_row["employee_id"],
                original_row["employee_name"],
                original_row["arrival_date"],
                original_row["arrival_at"],
                project,
                written_off_by_id,
                written_off_by_name,
                now_warsaw,
            )
    if inserted_row is None:
        return None
    return dict(inserted_row)


def format_materials_list(materials: list[str]) -> str:
    if not materials:
        return "‚Äî"
    return "\n".join(f"‚Ä¢ {item}" for item in materials)


def format_order_types_list(order_types: list[str]) -> str:
    if not order_types:
        return "‚Äî"
    return "\n".join(f"‚Ä¢ {name}" for name in order_types)


def _format_datetime(value: Optional[datetime]) -> str:
    if value is None:
        return "‚Äî"
    try:
        localised = value.astimezone(WARSAW_TZ)
    except Exception:
        localised = value
    return localised.strftime("%Y-%m-%d %H:%M")


def _format_date(value: Optional[date]) -> str:
    if value is None:
        return "‚Äî"
    return value.strftime("%d.%m.%Y")


def _parse_due_date_input(value: str) -> Optional[date]:
    cleaned = value.strip()
    if not cleaned:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def _decimal_to_excel_number(value: Optional[Decimal]) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError, InvalidOperation):
        try:
            return float(Decimal(str(value)))
        except (InvalidOperation, ValueError, TypeError):
            return None


def _format_date_for_excel(value: Optional[date], fallback: Optional[datetime] = None) -> str:
    if value is not None:
        return value.strftime("%Y-%m-%d")
    if fallback is None:
        return ""
    try:
        localised = fallback.astimezone(WARSAW_TZ)
    except Exception:
        localised = fallback
    return localised.strftime("%Y-%m-%d")


def _format_datetime_for_excel(value: Optional[datetime]) -> str:
    if value is None:
        return ""
    try:
        localised = value.astimezone(WARSAW_TZ)
    except Exception:
        localised = value
    return localised.strftime("%Y-%m-%d %H:%M")


def parse_user_created_at_input(text: str) -> Optional[datetime]:
    cleaned = text.strip()
    if not cleaned:
        return None

    datetime_formats = [
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y %H:%M",
    ]
    date_formats = ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"]

    for fmt in datetime_formats:
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return parsed.replace(tzinfo=WARSAW_TZ)
        except ValueError:
            continue

    for fmt in date_formats:
        try:
            parsed_date = datetime.strptime(cleaned, fmt).date()
            combined = datetime.combine(parsed_date, time.min, tzinfo=WARSAW_TZ)
            return combined
        except ValueError:
            continue

    return None


def format_user_record_for_message(record: Dict[str, Any], index: int) -> str:
    tg_id = record.get("tg_id") or "‚Äî"
    username = record.get("username") or "‚Äî"
    position = record.get("position") or "‚Äî"
    role = record.get("role") or "‚Äî"
    created_at = record.get("created_at")
    created_text = _format_datetime(created_at)
    return (
        f"{index}. üë§ {username}\n"
        f"   ‚Ä¢ TG ID: {tg_id}\n"
        f"   ‚Ä¢ –î–æ–ª–∂–Ω–æ—Å—Ç—å: {position}\n"
        f"   ‚Ä¢ –†–æ–ª—å: {role}\n"
        f"   ‚Ä¢ –î–æ–±–∞–≤–ª–µ–Ω: {created_text}"
    )


def split_text_into_messages(text: str, limit: int = 4000) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts = text.split("\n\n")
    chunks: list[str] = []
    current = ""
    for part in parts:
        candidate = part if not current else f"{current}\n\n{part}"
        if len(candidate) <= limit:
            current = candidate
            continue
        if current:
            chunks.append(current)
        if len(part) > limit:
            for start in range(0, len(part), limit):
                chunks.append(part[start : start + limit])
            current = ""
        else:
            current = part
    if current:
        chunks.append(current)
    return chunks


def format_thickness_value(thickness: Decimal) -> str:
    as_str = format(thickness, "f").rstrip("0").rstrip(".")
    if not as_str:
        as_str = "0"
    return f"{as_str} –º–º"


def format_dimension_value(value: Optional[Decimal]) -> str:
    if value is None:
        return "‚Äî"
    as_str = format(value, "f").rstrip("0").rstrip(".")
    if not as_str:
        as_str = "0"
    return f"{as_str} –º–º"


def format_thicknesses_list(thicknesses: list[Decimal]) -> str:
    if not thicknesses:
        return "‚Äî"
    return ", ".join(format_thickness_value(value) for value in thicknesses)


def format_colors_list(colors: list[str]) -> str:
    if not colors:
        return "‚Äî"
    return ", ".join(colors)


def format_series_list(series: list[str]) -> str:
    if not series:
        return "‚Äî"
    return ", ".join(series)


def format_storage_locations_list(locations: list[str]) -> str:
    if not locations:
        return "‚Äî"
    return "\n".join(f"‚Ä¢ {item}" for item in locations)


def build_plastics_export_file(records: list[Dict[str, Any]]) -> BufferedInputFile:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plastics"

    headers = [
        "–ê—Ä—Ç–∏–∫—É–ª",
        "–ú–∞—Ç–µ—Ä–∏–∞–ª",
        "–¢–æ–ª—â–∏–Ω–∞ (–º–º)",
        "–¶–≤–µ—Ç",
        "–î–ª–∏–Ω–∞ (–º–º)",
        "–®–∏—Ä–∏–Ω–∞ (–º–º)",
        "–ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è",
        "–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π",
        "–û—Ç–≤–µ—Ç—Å—Ç–≤–µ–Ω–Ω—ã–π",
        "–î–∞—Ç–∞ –ø—Ä–∏–±—ã—Ç–∏—è",
        "–î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è –ø—Ä–∏–±—ã—Ç–∏—è",
    ]
    sheet.append(headers)

    for record in records:
        arrival_at: Optional[datetime] = record.get("arrival_at")
        arrival_date: Optional[date] = record.get("arrival_date")
        row = [
            record.get("article"),
            record.get("material"),
            _decimal_to_excel_number(record.get("thickness")),
            record.get("color"),
            _decimal_to_excel_number(record.get("length")),
            _decimal_to_excel_number(record.get("width")),
            record.get("warehouse"),
            record.get("comment"),
            record.get("employee_name"),
            _format_date_for_excel(arrival_date, arrival_at),
            _format_datetime_for_excel(arrival_at),
        ]
        sheet.append(row)

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        adjusted_width = min(max(12, max_length + 2), 40)
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    timestamp = datetime.now(WARSAW_TZ).strftime("%Y%m%d_%H%M%S")
    filename = f"plastics_export_{timestamp}.xlsx"
    return BufferedInputFile(buffer.getvalue(), filename=filename)


def build_films_export_file(records: list[Dict[str, Any]]) -> BufferedInputFile:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Films"

    headers = [
        "–ê—Ä—Ç–∏–∫—É–ª",
        "–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å",
        "–°–µ—Ä–∏—è",
        "–ö–æ–¥ —Ü–≤–µ—Ç–∞",
        "–¶–≤–µ—Ç",
        "–®–∏—Ä–∏–Ω–∞ (–º–º)",
        "–î–ª–∏–Ω–∞ (–º–º)",
        "–ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è",
        "–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π",
        "–ù–∏–∫ —Å–æ—Ç—Ä—É–¥–Ω–∏–∫–∞",
        "ID —Å–æ—Ç—Ä—É–¥–Ω–∏–∫–∞",
        "–î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è –∑–∞–ø–∏—Å–∏",
    ]
    sheet.append(headers)

    for record in records:
        row = [
            record.get("article"),
            record.get("manufacturer"),
            record.get("series"),
            record.get("color_code"),
            record.get("color"),
            _decimal_to_excel_number(record.get("width")),
            _decimal_to_excel_number(record.get("length")),
            record.get("warehouse"),
            record.get("comment"),
            record.get("employee_nick"),
            record.get("employee_id"),
            _format_datetime_for_excel(record.get("recorded_at")),
        ]
        sheet.append(row)

    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        adjusted_width = min(max(12, max_length + 2), 40)
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = adjusted_width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    timestamp = datetime.now(WARSAW_TZ).strftime("%Y%m%d_%H%M%S")
    filename = f"films_export_{timestamp}.xlsx"
    return BufferedInputFile(buffer.getvalue(), filename=filename)


def format_plastic_record_for_message(record: Dict[str, Any]) -> str:
    thickness = record.get("thickness")
    arrival_at = record.get("arrival_at")
    if arrival_at:
        try:
            arrival_local = arrival_at.astimezone(WARSAW_TZ)
        except Exception:
            arrival_local = arrival_at
        arrival_text = arrival_local.strftime("%Y-%m-%d %H:%M")
    else:
        arrival_text = "‚Äî"
    return (
        f"–ê—Ä—Ç–∏–∫—É–ª: {record.get('article') or '‚Äî'}\n"
        f"–ú–∞—Ç–µ—Ä–∏–∞–ª: {record.get('material') or '‚Äî'}\n"
        f"–¢–æ–ª—â–∏–Ω–∞: {format_thickness_value(thickness) if thickness is not None else '‚Äî'}\n"
        f"–¶–≤–µ—Ç: {record.get('color') or '‚Äî'}\n"
        f"–î–ª–∏–Ω–∞: {format_dimension_value(record.get('length'))}\n"
        f"–®–∏—Ä–∏–Ω–∞: {format_dimension_value(record.get('width'))}\n"
        f"–°–∫–ª–∞–¥: {record.get('warehouse') or '‚Äî'}\n"
        f"–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {record.get('comment') or '‚Äî'}\n"
        f"–î–æ–±–∞–≤–∏–ª: {record.get('employee_name') or '‚Äî'}\n"
        f"–î–æ–±–∞–≤–ª–µ–Ω–æ: {arrival_text}"
    )


def format_film_record_for_message(record: Dict[str, Any]) -> str:
    recorded_at = record.get("recorded_at")
    if recorded_at:
        try:
            recorded_local = recorded_at.astimezone(WARSAW_TZ)
        except Exception:
            recorded_local = recorded_at
        recorded_text = recorded_local.strftime("%Y-%m-%d %H:%M")
    else:
        recorded_text = "‚Äî"
    return (
        f"–ê—Ä—Ç–∏–∫—É–ª: {record.get('article') or '‚Äî'}\n"
        f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {record.get('manufacturer') or '‚Äî'}\n"
        f"–°–µ—Ä–∏—è: {record.get('series') or '‚Äî'}\n"
        f"–ö–æ–¥ —Ü–≤–µ—Ç–∞: {record.get('color_code') or '‚Äî'}\n"
        f"–¶–≤–µ—Ç: {record.get('color') or '‚Äî'}\n"
        f"–®–∏—Ä–∏–Ω–∞: {format_dimension_value(record.get('width'))}\n"
        f"–î–ª–∏–Ω–∞: {format_dimension_value(record.get('length'))}\n"
        f"–°–∫–ª–∞–¥: {record.get('warehouse') or '‚Äî'}\n"
        f"–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {record.get('comment') or '‚Äî'}\n"
        f"–ù–∏–∫: {record.get('employee_nick') or '‚Äî'}\n"
        f"ID: {record.get('employee_id') or '‚Äî'}\n"
        f"–î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è: {recorded_text}"
    )


def format_film_records_list_for_message(records: list[Dict[str, Any]]) -> str:
    parts: list[str] = []
    for index, record in enumerate(records, start=1):
        formatted = format_film_record_for_message(record)
        parts.append(f"{index}.\n{formatted}")
    return "\n\n".join(parts)


def format_written_off_film_record(record: Dict[str, Any]) -> str:
    base_info = format_film_record_for_message(record)
    project = record.get("project") or "‚Äî"
    written_off_at = record.get("written_off_at")
    if written_off_at:
        try:
            written_off_local = written_off_at.astimezone(WARSAW_TZ)
        except Exception:
            written_off_local = written_off_at
        written_off_text = written_off_local.strftime("%Y-%m-%d %H:%M")
    else:
        written_off_text = "‚Äî"
    written_off_by_name = record.get("written_off_by_name") or "‚Äî"
    written_off_by_id = record.get("written_off_by_id")
    written_off_by_id_text = "‚Äî" if written_off_by_id is None else str(written_off_by_id)
    return (
        f"{base_info}\n"
        f"–ü—Ä–æ–µ–∫—Ç: {project}\n"
        f"–°–ø–∏—Å–∞–ª: {written_off_by_name}\n"
        f"ID —Å–ø–∏—Å–∞–≤—à–µ–≥–æ: {written_off_by_id_text}\n"
        f"–°–ø–∏—Å–∞–Ω–æ: {written_off_text}"
    )


def format_written_off_plastic_record(record: Dict[str, Any]) -> str:
    base_info = format_plastic_record_for_message(record)
    project = record.get("project") or "‚Äî"
    written_off_at = record.get("written_off_at")
    if written_off_at:
        try:
            written_off_local = written_off_at.astimezone(WARSAW_TZ)
        except Exception:
            written_off_local = written_off_at
        written_off_text = written_off_local.strftime("%Y-%m-%d %H:%M")
    else:
        written_off_text = "‚Äî"
    written_off_by_name = record.get("written_off_by_name") or "‚Äî"
    written_off_by_id = record.get("written_off_by_id")
    if written_off_by_id is None:
        written_off_by_id_text = "‚Äî"
    else:
        written_off_by_id_text = str(written_off_by_id)
    return (
        f"{base_info}\n"
        f"–ü—Ä–æ–µ–∫—Ç: {project}\n"
        f"–°–ø–∏—Å–∞–ª: {written_off_by_name}\n"
        f"ID —Å–ø–∏—Å–∞–≤—à–µ–≥–æ: {written_off_by_id_text}\n"
        f"–°–ø–∏—Å–∞–Ω–æ: {written_off_text}"
    )


def parse_thickness_input(raw_text: str) -> Optional[Decimal]:
    if raw_text is None:
        return None
    cleaned = raw_text.strip().lower()
    for suffix in ("–º–º", "mm"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    cleaned = cleaned.replace(" ", "").replace(",", ".")
    if not cleaned:
        return None
    try:
        value = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if value <= 0:
        return None
    return value.quantize(Decimal("0.01"))


def parse_dimension_filter_value(raw_text: str) -> Optional[Decimal]:
    if raw_text is None:
        return None
    cleaned = raw_text.strip().lower()
    for suffix in ("–º–º", "mm"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    cleaned = cleaned.replace(" ", "").replace(",", ".")
    if not cleaned:
        return None
    try:
        value = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if value < 0:
        return None
    return value.quantize(Decimal("0.01"))


def parse_positive_decimal(raw_text: str) -> Optional[Decimal]:
    if raw_text is None:
        return None
    cleaned = raw_text.strip().lower()
    for suffix in ("–º–º", "mm"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    cleaned = cleaned.replace(" ", "").replace(",", ".")
    if not cleaned:
        return None
    try:
        value = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None
    if value <= 0:
        return None
    return value.quantize(Decimal("0.01"))


def parse_positive_integer(raw_text: str) -> Optional[int]:
    if raw_text is None:
        return None
    cleaned = (raw_text or "").strip()
    if not cleaned.isdigit():
        return None
    value = int(cleaned)
    if value <= 0:
        return None
    return value


def build_materials_keyboard(materials: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for name in materials:
        rows.append([KeyboardButton(text=name)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_manufacturers_keyboard(manufacturers: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for name in manufacturers:
        rows.append([KeyboardButton(text=name)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_series_keyboard(series: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for name in series:
        rows.append([KeyboardButton(text=name)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_power_values_keyboard(values: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in values:
        rows.append([KeyboardButton(text=value)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_voltage_values_keyboard(values: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in values:
        rows.append([KeyboardButton(text=value)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_lens_counts_keyboard(counts: list[int]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in counts:
        rows.append([KeyboardButton(text=str(value))])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_led_module_articles_keyboard(articles: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for article in articles:
        rows.append([KeyboardButton(text=article)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_thickness_keyboard(thicknesses: list[Decimal]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in thicknesses:
        rows.append([KeyboardButton(text=format_thickness_value(value))])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_colors_keyboard(colors: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in colors:
        rows.append([KeyboardButton(text=value)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_advanced_materials_keyboard(materials: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for name in materials:
        rows.append([KeyboardButton(text=name)])
    rows.append([KeyboardButton(text=ADVANCED_SEARCH_SKIP_MATERIAL_TEXT)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_advanced_thickness_keyboard(thicknesses: list[Decimal]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in thicknesses:
        rows.append([KeyboardButton(text=format_thickness_value(value))])
    rows.append([KeyboardButton(text=ADVANCED_SEARCH_ALL_THICKNESSES_TEXT)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_advanced_colors_keyboard(colors: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for value in colors:
        rows.append([KeyboardButton(text=value)])
    rows.append([KeyboardButton(text=ADVANCED_SEARCH_ALL_COLORS_TEXT)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def build_storage_locations_keyboard(locations: list[str]) -> ReplyKeyboardMarkup:
    rows: list[list[KeyboardButton]] = []
    for location in locations:
        rows.append([KeyboardButton(text=location)])
    rows.append([KeyboardButton(text=CANCEL_TEXT)])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


# === –°–µ—Ä–≤–∏—Å–Ω—ã–µ —Ñ—É–Ω–∫—Ü–∏–∏ ===
async def send_plastic_settings_overview(message: Message) -> None:
    materials = await fetch_materials_with_thicknesses()
    storage_locations = await fetch_plastic_storage_locations()
    if materials:
        lines = []
        for material in materials:
            name = material["name"]
            thicknesses = material.get("thicknesses") or []
            formatted_thicknesses = format_thicknesses_list(thicknesses)
            colors = material.get("colors") or []
            formatted_colors = format_colors_list(colors)
            lines.append(
                "\n".join(
                    [
                        f"‚Ä¢ {name}",
                        f"   –¢–æ–ª—â–∏–Ω—ã: {formatted_thicknesses}",
                        f"   –¶–≤–µ—Ç–∞: {formatted_colors}",
                    ]
                )
            )
        materials_list = "\n".join(lines)
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–∞—Å—Ç–∏–∫.\n\n"
            "–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã, —Ç–æ–ª—â–∏–Ω—ã –∏ —Ü–≤–µ—Ç–∞:\n"
            f"{materials_list}"
        )
    else:
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–∞—Å—Ç–∏–∫.\n\n"
            "–ú–∞—Ç–µ—Ä–∏–∞–ª—ã –µ—â—ë –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã. –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ."
        )
    storage_text = format_storage_locations_list(storage_locations)
    text = f"{text}\n\n–ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n{storage_text}"
    await message.answer(text, reply_markup=WAREHOUSE_SETTINGS_PLASTIC_KB)


async def send_storage_locations_overview(message: Message) -> None:
    locations = await fetch_plastic_storage_locations()
    formatted = format_storage_locations_list(locations)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–∞—Å—Ç–∏–∫ ‚Üí –ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ.",
        reply_markup=WAREHOUSE_SETTINGS_PLASTIC_STORAGE_KB,
    )


async def send_film_settings_overview(message: Message) -> None:
    manufacturers = await fetch_film_manufacturers_with_series()
    if manufacturers:
        lines = []
        for manufacturer in manufacturers:
            name = manufacturer["name"]
            series = manufacturer.get("series") or []
            formatted_series = format_series_list(series)
            lines.append(
                "\n".join(
                    [
                        f"‚Ä¢ {name}",
                        f"   –°–µ—Ä–∏–∏: {formatted_series}",
                    ]
                )
            )
        formatted = "\n".join(lines)
        intro = "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏ –∏ —Å–µ—Ä–∏–∏:"
    else:
        formatted = (
            "–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏ –µ—â—ë –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π,"
            " –∞ –∑–∞—Ç–µ–º —É–∫–∞–∂–∏—Ç–µ –¥–ª—è –Ω–∏—Ö —Å–µ—Ä–∏–∏."
        )
        intro = "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç."
    storage_locations = await fetch_film_storage_locations()
    storage_text = format_storage_locations_list(storage_locations)
    text = (
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–µ–Ω–∫–∏.\n\n"
        f"{intro}\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ ¬´üè≠ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å¬ª –∏ ¬´üé¨ –°–µ—Ä–∏—è¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å —Å–ø–∏—Å–∫–∞–º–∏."\
        "\n\n"
        "–ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n"
        f"{storage_text}\n\n"
        "–ö–Ω–æ–ø–∫–∞ ¬´üè¨ –°–∫–ª–∞–¥¬ª –ø–æ–º–æ–∂–µ—Ç –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è."
    )
    await message.answer(text, reply_markup=WAREHOUSE_SETTINGS_FILM_KB)


async def send_film_manufacturers_menu(message: Message) -> None:
    manufacturers = await fetch_film_manufacturers()
    formatted = format_materials_list(manufacturers)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–µ–Ω–∫–∏ ‚Üí –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
        reply_markup=WAREHOUSE_SETTINGS_FILM_MANUFACTURERS_KB,
    )


async def send_film_storage_overview(message: Message) -> None:
    locations = await fetch_film_storage_locations()
    formatted = format_storage_locations_list(locations)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–µ–Ω–∫–∏ ‚Üí –°–∫–ª–∞–¥.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ.",
        reply_markup=WAREHOUSE_SETTINGS_FILM_STORAGE_KB,
    )


async def send_electrics_settings_overview(message: Message) -> None:
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞.\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é, –∫–æ—Ç–æ—Ä—É—é —Ö–æ—Ç–∏—Ç–µ –Ω–∞—Å—Ç—Ä–æ–∏—Ç—å.",
        reply_markup=WAREHOUSE_SETTINGS_ELECTRICS_KB,
    )

async def send_led_modules_settings_overview(message: Message) -> None:
    manufacturers = await fetch_led_module_manufacturers_with_series()
    storage_locations = await fetch_led_module_storage_locations()
    lens_counts = await fetch_led_module_lens_counts()
    colors = await fetch_led_module_colors()
    power_options = await fetch_led_module_power_options()
    voltage_options = await fetch_led_module_voltage_options()
    formatted_lens_counts = format_materials_list([str(value) for value in lens_counts])
    formatted_colors = format_materials_list(colors)
    formatted_power = format_materials_list(power_options)
    formatted_voltage = format_materials_list(voltage_options)
    formatted_storage = format_storage_locations_list(storage_locations)
    if manufacturers:
        lines: list[str] = []
        for manufacturer in manufacturers:
            name = manufacturer["name"]
            formatted_series = format_series_list(manufacturer.get("series") or [])
            lines.append(
                "\n".join(
                    [
                        f"‚Ä¢ {name}",
                        f"   –°–µ—Ä–∏–∏: {formatted_series}",
                    ]
                )
            )
        formatted = "\n".join(lines)
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏.\n\n"
            "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏ –∏ —Å–µ—Ä–∏–∏:\n"
            f"{formatted}\n\n"
            "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ ¬´üè≠ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å Led –º–æ–¥—É–ª–µ–π¬ª –∏ ¬´üé¨ –°–µ—Ä–∏—è Led –º–æ–¥—É–ª–µ–π¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å —Å–ø–∏—Å–∫–∞–º–∏."
        )
    else:
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏.\n\n"
            "–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏ –µ—â—ë –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, —á—Ç–æ–±—ã –Ω–∞—á–∞—Ç—å."\
            " –ó–∞—Ç–µ–º –º–æ–∂–Ω–æ –±—É–¥–µ—Ç —É–∫–∞–∑–∞—Ç—å —Å–µ—Ä–∏–∏."
        )
    text += (
        "\n\n–ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n"
        f"{formatted_storage}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´üè¨ –ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è Led –º–æ–¥—É–ª–µ–π¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å —Å–ø–∏—Å–∫–æ–º."
    )
    text += (
        "\n\n–î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ª–∏–Ω–∑:\n"
        f"{formatted_lens_counts}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´üî¢ –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å –æ–±—â–∏–º —Å–ø–∏—Å–∫–æ–º –∑–Ω–∞—á–µ–Ω–∏–π."
    )
    text += (
        "\n\n–î–æ—Å—Ç—É–ø–Ω—ã–µ —Ü–≤–µ—Ç–∞ –º–æ–¥—É–ª–µ–π:\n"
        f"{formatted_colors}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´üé® –¶–≤–µ—Ç –º–æ–¥—É–ª–µ–π¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å –æ–±—â–∏–º —Å–ø–∏—Å–∫–æ–º —Ü–≤–µ—Ç–æ–≤."
    )
    text += (
        "\n\n–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–æ—â–Ω–æ—Å—Ç–∏ –º–æ–¥—É–ª–µ–π:\n"
        f"{formatted_power}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´‚ö° –ú–æ—â–Ω–æ—Å—Ç—å –º–æ–¥—É–ª–µ–π¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å –æ–±—â–∏–º —Å–ø–∏—Å–∫–æ–º –∑–Ω–∞—á–µ–Ω–∏–π."
    )
    text += (
        "\n\n–î–æ—Å—Ç—É–ø–Ω—ã–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è –º–æ–¥—É–ª–µ–π:\n"
        f"{formatted_voltage}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´üîå –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ –º–æ–¥—É–ª–µ–π¬ª, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å –æ–±—â–∏–º —Å–ø–∏—Å–∫–æ–º –∑–Ω–∞—á–µ–Ω–∏–π."
    )
    await message.answer(text, reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB)


async def send_led_module_manufacturers_menu(message: Message) -> None:
    manufacturers = await fetch_led_module_manufacturers()
    formatted = format_materials_list(manufacturers)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_MANUFACTURERS_KB,
    )


async def send_led_module_storage_overview(message: Message) -> None:
    locations = await fetch_led_module_storage_locations()
    formatted = format_storage_locations_list(locations)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –ú–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_STORAGE_KB,
    )


async def send_led_module_base_menu(message: Message) -> None:
    modules = await fetch_generated_led_modules_with_details()
    if modules:
        lines = []
        for module in modules:
            article = module.get("article", "‚Äî")
            manufacturer = module.get("manufacturer", "‚Äî")
            series = module.get("series", "‚Äî")
            color = module.get("color", "‚Äî")
            lens_count = module.get("lens_count")
            power = module.get("power", "‚Äî")
            voltage = module.get("voltage", "‚Äî")
            lens_text = "‚Äî" if lens_count is None else str(lens_count)
            lines.append(
                " | ".join(
                    [
                        f"–ê—Ä—Ç–∏–∫—É–ª: {article}",
                        f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {manufacturer}",
                        f"–°–µ—Ä–∏—è: {series}",
                        f"–¶–≤–µ—Ç: {color}",
                        f"–õ–∏–Ω–∑: {lens_text}",
                        f"–ú–æ—â–Ω–æ—Å—Ç—å: {power}",
                        f"–ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ: {voltage}",
                    ]
                )
            )
        generated_text = "üìã –£–∂–µ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ Led –º–æ–¥—É–ª–∏:\n" + "\n".join(lines)
    else:
        generated_text = (
            "‚ÑπÔ∏è –ü–æ–∫–∞ –Ω–µ—Ç —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω–Ω—ã—Ö Led –º–æ–¥—É–ª–µ–π. –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å Led –º–æ–¥—É–ª—å¬ª."
        )
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí Led –º–æ–¥—É–ª–∏ baza.\n\n"
        f"{generated_text}\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å –±–∞–∑–æ–π Led –º–æ–¥—É–ª–µ–π.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
    )


async def send_led_module_colors_menu(message: Message) -> None:
    colors = await fetch_led_module_colors()
    formatted = format_materials_list(colors)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –¶–≤–µ—Ç –º–æ–¥—É–ª–µ–π.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ —Ü–≤–µ—Ç–∞:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å —Ü–≤–µ—Ç.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_COLORS_KB,
    )


async def send_led_module_power_menu(message: Message) -> None:
    power_options = await fetch_led_module_power_options()
    formatted = format_materials_list(power_options)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –ú–æ—â–Ω–æ—Å—Ç—å –º–æ–¥—É–ª–µ–π.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –º–æ—â–Ω–æ—Å—Ç–∏:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –∑–Ω–∞—á–µ–Ω–∏–µ.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_POWER_KB,
    )


async def send_led_module_voltage_menu(message: Message) -> None:
    voltage_options = await fetch_led_module_voltage_options()
    formatted = format_materials_list(voltage_options)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ –º–æ–¥—É–ª–µ–π.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –∑–Ω–∞—á–µ–Ω–∏–µ.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_VOLTAGE_KB,
    )


async def send_led_module_lens_menu(message: Message) -> None:
    lens_counts = await fetch_led_module_lens_counts()
    formatted = format_materials_list([str(value) for value in lens_counts])
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ª–∏–Ω–∑:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –∑–Ω–∞—á–µ–Ω–∏–µ.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_LENS_KB,
    )


async def send_led_module_series_menu(message: Message) -> None:
    manufacturers = await fetch_led_module_manufacturers_with_series()
    if manufacturers:
        lines: list[str] = []
        for manufacturer in manufacturers:
            name = manufacturer["name"]
            formatted_series = format_series_list(manufacturer.get("series") or [])
            lines.append(
                "\n".join(
                    [
                        f"‚Ä¢ {name}",
                        f"   –°–µ—Ä–∏–∏: {formatted_series}",
                    ]
                )
            )
        formatted = "\n".join(lines)
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –°–µ—Ä–∏—è.\n\n"
            "–î–æ—Å—Ç—É–ø–Ω—ã–µ —Å–µ—Ä–∏–∏ –ø–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è–º:\n"
            f"{formatted}\n\n"
            "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é."
        )
    else:
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí –°–µ—Ä–∏—è.\n\n"
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π, —á—Ç–æ–±—ã —Å–æ–∑–¥–∞–≤–∞—Ç—å —Å–µ—Ä–∏–∏."
        )
    await message.answer(text, reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB)


async def send_led_strips_settings_overview(message: Message) -> None:
    manufacturers = await fetch_led_strip_manufacturers()
    formatted = format_materials_list(manufacturers)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –ª–µ–Ω—Ç–∞.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
        reply_markup=WAREHOUSE_SETTINGS_LED_STRIPS_MANUFACTURERS_KB,
    )


async def send_power_supplies_settings_overview(message: Message) -> None:
    manufacturers = await fetch_power_supply_manufacturers()
    formatted = format_materials_list(manufacturers)
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí –ë–ª–æ–∫–∏ –ø–∏—Ç–∞–Ω–∏—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏:\n"
        f"{formatted}\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
        reply_markup=WAREHOUSE_SETTINGS_POWER_SUPPLIES_MANUFACTURERS_KB,
    )


# === –ö–æ–º–∞–Ω–¥—ã ===
@dp.message(CommandStart())
async def handle_start(message: Message) -> None:
    await message.answer("üëã –ü—Ä–∏–≤–µ—Ç! –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=MAIN_MENU_KB)


@dp.message(Command("settings"))
@dp.message(F.text == "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏")
async def handle_settings(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await message.answer("‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=SETTINGS_MENU_KB)


@dp.message(F.text == "üîÑ –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∑–∏—Ç—å")
async def handle_restart(message: Message) -> None:
    if not await ensure_admin_access(message):
        return

    if not UPDATE_SCRIPT_PATH.exists():
        await message.answer(
            "‚ö†Ô∏è –§–∞–π–ª update.sh –Ω–µ –Ω–∞–π–¥–µ–Ω –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ.", reply_markup=SETTINGS_MENU_KB
        )
        return

    await message.answer(
        "‚ôªÔ∏è –ü–µ—Ä–µ–∑–∞–ø—É—Å–∫ —Å–∏—Å—Ç–µ–º—ã –Ω–∞—á–∞—Ç... –ü–æ–¥–æ–∂–¥–∏ –Ω–µ–º–Ω–æ–≥–æ ‚è≥",
        reply_markup=SETTINGS_MENU_KB,
    )

    try:
        subprocess.Popen(
            ["bash", str(UPDATE_SCRIPT_PATH)],
            cwd=str(UPDATE_SCRIPT_PATH.parent),
        )
    except Exception as exc:
        await message.answer(
            f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è:\n`{exc}`",
            reply_markup=SETTINGS_MENU_KB,
        )
        return

    await message.answer(
        "‚úÖ –°–∫—Ä–∏–ø—Ç –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –∑–∞–ø—É—â–µ–Ω!\n–Ø –ø—Ä–∏—à–ª—é —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ, –∫–æ–≥–¥–∞ –ø—Ä–æ—Ü–µ—Å—Å –∑–∞–≤–µ—Ä—à–∏—Ç—Å—è.",
        reply_markup=SETTINGS_MENU_KB,
    )


@dp.message(F.text == "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞")
async def handle_warehouse_settings(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await message.answer("‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=WAREHOUSE_SETTINGS_MENU_KB)


@dp.message(F.text == "üë• –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏")
async def handle_users_menu(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await message.answer("üë• –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=USERS_MENU_KB)


@dp.message(F.text == "üìã –ü–æ—Å–º–æ—Ç—Ä–µ—Ç—å –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π")
async def handle_list_all_users(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    users = await fetch_all_users_from_db()
    if not users:
        await message.answer(
            "‚ÑπÔ∏è –í –±–∞–∑–µ –ø–æ–∫–∞ –Ω–µ—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π.", reply_markup=USERS_MENU_KB
        )
        return
    formatted_records = [
        format_user_record_for_message(record, index)
        for index, record in enumerate(users, start=1)
    ]
    header = f"üìã –í—Å–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {len(users)}"
    full_text = f"{header}\n\n" + "\n\n".join(formatted_records)
    chunks = split_text_into_messages(full_text)
    for idx, chunk in enumerate(chunks):
        if idx == 0:
            await message.answer(chunk, reply_markup=USERS_MENU_KB)
        else:
            await message.answer(chunk)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è")
async def handle_add_user_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await state.set_state(AddUserStates.waiting_for_tg_id)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ Telegram ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (—Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddUserStates.waiting_for_tg_id)
async def process_add_user_tg_id(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_user_flow(message, state)
        return
    if not text.isdigit():
        await message.answer(
            "‚ö†Ô∏è Telegram ID –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(tg_id=int(text))
    await state.set_state(AddUserStates.waiting_for_username)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∏–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–∫–∞–∫ –±—É–¥–µ—Ç –æ—Ç–æ–±—Ä–∞–∂–∞—Ç—å—Å—è –≤ —Å–ø–∏—Å–∫–µ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddUserStates.waiting_for_username)
async def process_add_user_username(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_user_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ò–º—è –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –í–≤–µ–¥–∏—Ç–µ –∏–º—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(username=text)
    await state.set_state(AddUserStates.waiting_for_position)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –¥–æ–ª–∂–Ω–æ—Å—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddUserStates.waiting_for_position)
async def process_add_user_position(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_user_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –î–æ–ª–∂–Ω–æ—Å—Ç—å –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç–æ–π. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(position=text)
    await state.set_state(AddUserStates.waiting_for_role)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ —Ä–æ–ª—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä –∏–ª–∏ —Å–æ—Ç—Ä—É–¥–Ω–∏–∫).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddUserStates.waiting_for_role)
async def process_add_user_role(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_user_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –†–æ–ª—å –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç–æ–π. –í–≤–µ–¥–∏—Ç–µ –æ–ø–∏—Å–∞–Ω–∏–µ —Ä–æ–ª–∏.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(role=text)
    await state.set_state(AddUserStates.waiting_for_created_at)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –¥–∞—Ç—É –∏ –≤—Ä–µ–º—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, 2024-01-31 –∏–ª–∏"
        " 31.01.2024 09:30).\n–ï—Å–ª–∏ —Ö–æ—Ç–∏—Ç–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Ç–µ–∫—É—â–µ–µ –≤—Ä–µ–º—è, –Ω–∞–∂–º–∏—Ç–µ"
        " ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddUserStates.waiting_for_created_at)
async def process_add_user_created_at(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_user_flow(message, state)
        return

    custom_created_at: Optional[datetime]
    if text == SKIP_TEXT:
        custom_created_at = None
    else:
        parsed = parse_user_created_at_input(text)
        if parsed is None:
            await message.answer(
                "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å –¥–∞—Ç—É. –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ —Ñ–æ—Ä–º–∞—Ç –ì–ì–ì–ì-–ú–ú-–î–î –∏–ª–∏"
                " –ì–ì–ì–ì-–ú–ú-–î–î –ß–ß:–ú–ú. –ú–æ–∂–Ω–æ —Ç–∞–∫–∂–µ –≤–≤–µ—Å—Ç–∏ 31.01.2024 –∏–ª–∏ 31.01.2024"
                " 09:30.",
                reply_markup=SKIP_OR_CANCEL_KB,
            )
            return
        custom_created_at = parsed

    data = await state.get_data()
    tg_id = data.get("tg_id")
    username = data.get("username")
    position = data.get("position")
    role = data.get("role")

    if tg_id is None or username is None or position is None or role is None:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –≤–≤–µ–¥—ë–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=USERS_MENU_KB,
        )
        return

    try:
        await upsert_user_in_db(
            tg_id=int(tg_id),
            username=str(username),
            position=str(position),
            role=str(role),
            created_at=custom_created_at,
        )
    except Exception as exc:
        logging.exception("Failed to add or update user")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.\n"
            f"–¢–µ—Ö–Ω–∏—á–µ—Å–∫–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è: {exc}",
            reply_markup=USERS_MENU_KB,
        )
        return

    await state.clear()
    created_info = (
        _format_datetime(custom_created_at)
        if custom_created_at is not None
        else "—Ç–µ–∫—É—â–µ–µ –≤—Ä–µ–º—è (–ø–æ —É–º–æ–ª—á–∞–Ω–∏—é)"
    )
    await message.answer(
        "‚úÖ –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Å–æ—Ö—Ä–∞–Ω—ë–Ω.\n"
        f"üë§ –ò–º—è: {username}\n"
        f"üÜî TG ID: {tg_id}\n"
        f"üè¢ –î–æ–ª–∂–Ω–æ—Å—Ç—å: {position}\n"
        f"üîê –†–æ–ª—å: {role}\n"
        f"üóì –î–∞—Ç–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è: {created_info}",
        reply_markup=USERS_MENU_KB,
    )


@dp.message(F.text == "‚¨ÖÔ∏è –ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é")
async def handle_back_to_main(message: Message) -> None:
    await message.answer("–ì–ª–∞–≤–Ω–æ–µ –º–µ–Ω—é.", reply_markup=MAIN_MENU_KB)


@dp.message(F.text == "–ö–ª–∏–µ–Ω—Ç—ã")
async def handle_clients_section(message: Message) -> None:
    await message.answer(
        "üë• –†–∞–∑–¥–µ–ª ¬´–ö–ª–∏–µ–Ω—Ç—ã¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=CLIENTS_MENU_KB
    )


@dp.message(F.text == CLIENTS_ADD_CLIENT_TEXT)
async def handle_clients_add(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(AddClientStates.waiting_for_name)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞:",
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == CLIENTS_SEARCH_CLIENT_TEXT)
async def handle_clients_search(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(SearchClientStates.waiting_for_query)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞ –∏–ª–∏ –µ–≥–æ —á–∞—Å—Ç—å:",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddClientStates.waiting_for_name)
async def process_add_client_name(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_client_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(name=text)
    await state.set_state(AddClientStates.waiting_for_phone)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ —Ç–µ–ª–µ—Ñ–æ–Ω –∫–ª–∏–µ–Ω—Ç–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª:",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddClientStates.waiting_for_phone)
async def process_add_client_phone(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_client_flow(message, state)
        return
    phone: Optional[str]
    if text == SKIP_TEXT or not text:
        phone = None
    else:
        phone = text
    await state.update_data(phone=phone)
    await state.set_state(AddClientStates.waiting_for_contact_person)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∫–æ–Ω—Ç–∞–∫—Ç–Ω–æ–µ –ª–∏—Ü–æ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª:",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddClientStates.waiting_for_contact_person)
async def process_add_client_contact_person(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_client_flow(message, state)
        return
    contact_person: Optional[str]
    if text == SKIP_TEXT or not text:
        contact_person = None
    else:
        contact_person = text
    await state.update_data(contact_person=contact_person)
    await state.set_state(AddClientStates.waiting_for_address)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∞–¥—Ä–µ—Å –∫–ª–∏–µ–Ω—Ç–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª:",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddClientStates.waiting_for_address)
async def process_add_client_address(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_client_flow(message, state)
        return
    address: Optional[str]
    if text == SKIP_TEXT or not text:
        address = None
    else:
        address = text
    await state.update_data(address=address)
    await state.set_state(AddClientStates.waiting_for_map_link)
    await message.answer(
        "–î–æ–±–∞–≤—å—Ç–µ —Å—Å—ã–ª–∫—É –Ω–∞ Google –ö–∞—Ä—Ç—ã –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª:",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddClientStates.waiting_for_map_link)
async def process_add_client_map_link(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_client_flow(message, state)
        return
    map_link: Optional[str]
    if text == SKIP_TEXT or not text:
        map_link = None
    else:
        map_link = text
    data = await state.get_data()
    name = data.get("name")
    phone = data.get("phone")
    contact_person = data.get("contact_person")
    address = data.get("address")
    if not name:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∫–ª–∏–µ–Ω—Ç–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=CLIENTS_MENU_KB,
        )
        return
    try:
        client_id = await create_client_in_db(
            name=name,
            phone=phone,
            contact_person=contact_person,
        )
        if address is not None or map_link is not None:
            await add_client_address_in_db(
                client_id=client_id,
                address=address,
                google_maps_link=map_link,
            )
    except Exception as exc:
        logging.exception("Failed to add client")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –∫–ª–∏–µ–Ω—Ç–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.\n"
            f"–¢–µ—Ö–Ω–∏—á–µ—Å–∫–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è: {exc}",
            reply_markup=CLIENTS_MENU_KB,
        )
        return

    await state.clear()
    phone_text = phone or "‚Äî"
    contact_text = contact_person or "‚Äî"
    address_text = address or "‚Äî"
    map_text = map_link or "‚Äî"
    await message.answer(
        "‚úÖ –ö–ª–∏–µ–Ω—Ç –¥–æ–±–∞–≤–ª–µ–Ω.\n"
        f"üè¢ –ù–∞–∑–≤–∞–Ω–∏–µ: {name}\n"
        f"üìû –¢–µ–ª–µ—Ñ–æ–Ω: {phone_text}\n"
        f"üë§ –ö–æ–Ω—Ç–∞–∫—Ç–Ω–æ–µ –ª–∏—Ü–æ: {contact_text}\n"
        f"üìç –ê–¥—Ä–µ—Å: {address_text}\n"
        f"üîó Google –ö–∞—Ä—Ç—ã: {map_text}",
        reply_markup=CLIENTS_MENU_KB,
    )


@dp.message(SearchClientStates.waiting_for_query)
async def process_clients_search_query(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_client_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    try:
        matches, has_more = await search_clients_by_name(text)
    except Exception as exc:
        logging.exception("Failed to search clients")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –≤—ã–ø–æ–ª–Ω–∏—Ç—å –ø–æ–∏—Å–∫ –∫–ª–∏–µ–Ω—Ç–æ–≤. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.\n"
            f"–¢–µ—Ö–Ω–∏—á–µ—Å–∫–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è: {exc}",
            reply_markup=CLIENTS_MENU_KB,
        )
        return

    await state.clear()
    if not matches:
        await message.answer(
            "‚ÑπÔ∏è –ö–ª–∏–µ–Ω—Ç—ã —Å —Ç–∞–∫–∏–º –Ω–∞–∑–≤–∞–Ω–∏–µ–º –Ω–µ –Ω–∞–π–¥–µ–Ω—ã.",
            reply_markup=CLIENTS_MENU_KB,
        )
        return

    cards: list[str] = []
    for client in matches:
        card_lines = [
            f"üè¢ –ù–∞–∑–≤–∞–Ω–∏–µ: {client['name']}",
            f"üìû –¢–µ–ª–µ—Ñ–æ–Ω: {client.get('phone') or '‚Äî'}",
            f"üë§ –ö–æ–Ω—Ç–∞–∫—Ç–Ω–æ–µ –ª–∏—Ü–æ: {client.get('contact_person') or '‚Äî'}",
        ]
        addresses = client.get("addresses") or []
        if addresses:
            card_lines.append("üìç –ê–¥—Ä–µ—Å–∞:")
            for entry in addresses:
                address_value = entry.get("address") or "‚Äî"
                map_link_value = entry.get("google_maps_link")
                card_lines.append(f"  ‚Ä¢ {address_value}")
                if map_link_value:
                    card_lines.append(f"    üîó {map_link_value}")
        else:
            card_lines.append("üìç –ê–¥—Ä–µ—Å: ‚Äî")
            card_lines.append("üîó Google –ö–∞—Ä—Ç—ã: ‚Äî")
        cards.append("\n".join(card_lines))

    response_text = "üîé –†–µ–∑—É–ª—å—Ç–∞—Ç—ã –ø–æ–∏—Å–∫–∞:\n\n" + "\n\n".join(cards)
    if has_more:
        response_text += (
            "\n\n‚ÑπÔ∏è –ü–æ–∫–∞–∑–∞–Ω—ã –ø–µ—Ä–≤—ã–µ 10 —Å–æ–≤–ø–∞–¥–µ–Ω–∏–π. –£—Ç–æ—á–Ω–∏—Ç–µ –∑–∞–ø—Ä–æ—Å –¥–ª—è –±–æ–ª–µ–µ —Ç–æ—á–Ω–æ–≥–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞."
        )
    await message.answer(response_text, reply_markup=CLIENTS_MENU_KB)


@dp.message(F.text == "–ó–∞–∫–∞–∑—ã")
async def handle_orders_section(message: Message) -> None:
    await message.answer(
        "üßæ –†–∞–∑–¥–µ–ª ¬´–ó–∞–∫–∞–∑—ã¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=ORDERS_MENU_KB
    )


def _format_client_search_results_for_order(
    clients: list[Dict[str, Any]],
    has_more: bool,
) -> str:
    result_lines = ["üîé –ù–∞–π–¥–µ–Ω–Ω—ã–µ –∑–∞–∫–∞–∑—á–∏–∫–∏:", ""]
    formatted_entries: list[str] = []
    for index, client in enumerate(clients, start=1):
        entry_lines = [f"{index}. {client['name']}"]
        contact_person = (client.get("contact_person") or "").strip()
        phone = (client.get("phone") or "").strip()
        if contact_person:
            entry_lines.append(f"   üë§ –ö–æ–Ω—Ç–∞–∫—Ç: {contact_person}")
        if phone:
            entry_lines.append(f"   üìû –¢–µ–ª–µ—Ñ–æ–Ω: {phone}")
        addresses = client.get("addresses") or []
        if addresses:
            first_address = addresses[0]
            address_value = (first_address.get("address") or "").strip()
            if address_value:
                entry_lines.append(f"   üìç –ê–¥—Ä–µ—Å: {address_value}")
        formatted_entries.append("\n".join(entry_lines))
    result_lines.append("\n\n".join(formatted_entries))
    if has_more:
        result_lines.extend(
            [
                "",
                "‚ÑπÔ∏è –ü–æ–∫–∞–∑–∞–Ω—ã –ø–µ—Ä–≤—ã–µ —Å–æ–≤–ø–∞–¥–µ–Ω–∏—è. –£—Ç–æ—á–Ω–∏—Ç–µ –∑–∞–ø—Ä–æ—Å –¥–ª—è –±–æ–ª–µ–µ —Ç–æ—á–Ω–æ–≥–æ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞.",
            ]
        )
    result_lines.extend(
        [
            "",
            "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–º–µ—Ä –ø–æ–¥—Ö–æ–¥—è—â–µ–≥–æ –∑–∞–∫–∞–∑—á–∏–∫–∞ –∏–ª–∏ –≤–≤–µ–¥–∏—Ç–µ –Ω–æ–≤—ã–π –∑–∞–ø—Ä–æ—Å –¥–ª—è –ø–æ–∏—Å–∫–∞.",
        ]
    )
    return "\n".join(result_lines)


def _build_order_summary_message(
    order_row: Dict[str, Any],
    client_details: Dict[str, Any],
) -> str:
    due_date_raw = order_row.get("due_date")
    if isinstance(due_date_raw, datetime):
        due_date_value: Optional[date] = due_date_raw.date()
    else:
        due_date_value = due_date_raw
    due_date_text = _format_date(due_date_value)
    deadline_line = "‚è≥ –î–µ–¥–ª–∞–π–Ω: ‚Äî"
    if isinstance(due_date_value, date):
        today = datetime.now(WARSAW_TZ).date()
        days_diff = (due_date_value - today).days
        if days_diff > 0:
            deadline_line = f"‚è≥ –î–µ–¥–ª–∞–π–Ω: –æ—Å—Ç–∞–ª–æ—Å—å {days_diff} –¥–Ω."
        elif days_diff == 0:
            deadline_line = "‚è≥ –î–µ–¥–ª–∞–π–Ω: —Å–µ–≥–æ–¥–Ω—è"
        else:
            deadline_line = f"‚è≥ –î–µ–¥–ª–∞–π–Ω: –ø—Ä–æ—Å—Ä–æ—á–∫–∞ {abs(days_diff)} –¥–Ω."
    created_text = _format_datetime(order_row.get("created_at"))
    is_urgent = bool(order_row.get("is_urgent"))
    creator_name = order_row.get("created_by_name")
    if not creator_name:
        creator_id = order_row.get("created_by_id")
        creator_name = f"ID: {creator_id}" if creator_id else "‚Äî"
    client_lines = [f"üè¢ –ó–∞–∫–∞–∑—á–∏–∫: {client_details['name']}"]
    contact_person = (client_details.get("contact_person") or "").strip()
    phone = (client_details.get("phone") or "").strip()
    if contact_person:
        client_lines.append(f"üë§ –ö–æ–Ω—Ç–∞–∫—Ç–Ω–æ–µ –ª–∏—Ü–æ: {contact_person}")
    if phone:
        client_lines.append(f"üìû –¢–µ–ª–µ—Ñ–æ–Ω: {phone}")
    summary_lines = [
        f"‚úÖ –ó–∞–∫–∞–∑ ‚Ññ{order_row['order_number']} —Å–æ–∑–¥–∞–Ω.",
        "",
        "\n".join(client_lines),
        "",
        f"üìù –ù–∞–∑–≤–∞–Ω–∏–µ: {order_row['title']}",
        f"üóÇÔ∏è –¢–∏–ø –∑–∞–∫–∞–∑–∞: {order_row['order_type']}",
        f"üìÇ –ü–∞–ø–∫–∞: {order_row['folder_path']}",
        f"üìÖ –°—Ä–æ–∫ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è: {due_date_text}",
        deadline_line,
        f"üî• –°—Ä–æ—á–Ω—ã–π: {'–î–∞' if is_urgent else '–ù–µ—Ç'}",
        "",
        f"üïí –°–æ–∑–¥–∞–Ω–æ: {created_text}",
        f"‚úçÔ∏è –°–æ–∑–¥–∞–ª: {creator_name}",
    ]
    return "\n".join(summary_lines)


def format_orders_overview(order_rows: list[Dict[str, Any]]) -> str:
    if not order_rows:
        return "‚ÑπÔ∏è –í —Ç–∞–±–ª–∏—Ü–µ –∑–∞–∫–∞–∑–æ–≤ –ø–æ–∫–∞ –Ω–µ—Ç –∑–∞–ø–∏—Å–µ–π."

    overview_lines = ["üìã –ó–∞–∫–∞–∑—ã –≤ —Ä–∞–±–æ—Ç–µ:", ""]
    formatted_entries: list[str] = []

    for index, row in enumerate(order_rows, start=1):
        due_date_text = _format_date(row.get("due_date"))
        created_at_text = _format_datetime(row.get("created_at"))
        responsible = (row.get("created_by_name") or "").strip() or "‚Äî"
        urgency_text = "–î–∞" if row.get("is_urgent") else "–ù–µ—Ç"

        entry_lines = [
            f"{index}. ‚Ññ{row.get('order_number')} ‚Äî {row.get('title')}",
            f"   üè¢ –ó–∞–∫–∞–∑—á–∏–∫: {row.get('client_name')}",
            f"   üóÇÔ∏è –¢–∏–ø: {row.get('order_type')}",
            f"   üìÅ –ü–∞–ø–∫–∞: {row.get('folder_path')}",
            f"   üìÖ –°—Ä–æ–∫: {due_date_text}",
            f"   ‚ö° –°—Ä–æ—á–Ω—ã–π: {urgency_text}",
            f"   üïí –°–æ–∑–¥–∞–Ω: {created_at_text}",
            f"   üë§ –û—Ç–≤–µ—Ç—Å—Ç–≤–µ–Ω–Ω—ã–π: {responsible}",
        ]
        formatted_entries.append("\n".join(entry_lines))

    overview_lines.append("\n\n".join(formatted_entries))
    return "\n".join(overview_lines)


async def _send_client_search_prompt(
    message: Message,
    state: FSMContext,
    query: str,
) -> None:
    matches, has_more = await search_clients_by_name(query)
    if not matches:
        await message.answer(
            "‚ÑπÔ∏è –ó–∞–∫–∞–∑—á–∏–∫–∏ —Å —Ç–∞–∫–∏–º –Ω–∞–∑–≤–∞–Ω–∏–µ–º –Ω–µ –Ω–∞–π–¥–µ–Ω—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∑–∞–ø—Ä–æ—Å.",
            reply_markup=CANCEL_KB,
        )
        await state.set_state(CreateOrderStates.waiting_for_client_query)
        return
    await state.update_data(client_search_results=matches)
    await state.set_state(CreateOrderStates.waiting_for_client_selection)
    await message.answer(
        _format_client_search_results_for_order(matches, has_more),
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == ORDERS_IN_PROGRESS_TEXT)
async def handle_orders_in_progress(message: Message) -> None:
    try:
        orders = await fetch_all_orders()
    except Exception:
        logging.exception("Failed to fetch orders overview")
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –∑–∞–∫–∞–∑–æ–≤. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=ORDERS_MENU_KB,
        )
        return

    overview_text = format_orders_overview(orders)
    await message.answer(overview_text, reply_markup=ORDERS_MENU_KB)


@dp.message(F.text == ORDERS_NEW_ORDER_TEXT)
async def handle_orders_new_order(message: Message, state: FSMContext) -> None:
    await state.clear()
    order_types = await fetch_order_types()
    if not order_types:
        await message.answer(
            "‚ÑπÔ∏è –°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ —Ö–æ—Ç—è –±—ã –æ–¥–∏–Ω —Ç–∏–ø –∑–∞–∫–∞–∑–∞ –≤ —Ä–∞–∑–¥–µ–ª–µ ¬´‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –∑–∞–∫–∞–∑–æ–≤¬ª.",
            reply_markup=ORDERS_MENU_KB,
        )
        return
    next_number = await fetch_next_order_number()
    await state.update_data(order_number_hint=next_number)
    await state.set_state(CreateOrderStates.waiting_for_client_query)
    await message.answer(
        (
            f"üÜï –°–æ–∑–¥–∞–Ω–∏–µ –∑–∞–∫–∞–∑–∞ ‚Ññ{next_number}.\n"
            "–í–≤–µ–¥–∏—Ç–µ –Ω–µ –º–µ–Ω–µ–µ –¥–≤—É—Ö —Å–∏–º–≤–æ–ª–æ–≤ –Ω–∞–∑–≤–∞–Ω–∏—è –∑–∞–∫–∞–∑—á–∏–∫–∞, —á—Ç–æ–±—ã –Ω–∞–π—Ç–∏ –µ–≥–æ –≤ –±–∞–∑–µ."
        ),
        reply_markup=CANCEL_KB,
    )


@dp.message(CreateOrderStates.waiting_for_client_query)
async def process_order_client_query(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    if len(text) < 2:
        await message.answer(
            "‚ö†Ô∏è –í–≤–µ–¥–∏—Ç–µ –º–∏–Ω–∏–º—É–º –¥–≤–∞ —Å–∏–º–≤–æ–ª–∞ –¥–ª—è –ø–æ–∏—Å–∫–∞ –∑–∞–∫–∞–∑—á–∏–∫–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await _send_client_search_prompt(message, state, text)


@dp.message(CreateOrderStates.waiting_for_client_selection)
async def process_order_client_selection(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    data = await state.get_data()
    candidates = data.get("client_search_results") or []
    if text.isdigit():
        selected_index = int(text)
        if 1 <= selected_index <= len(candidates):
            selected_client = candidates[selected_index - 1]
            await state.update_data(
                selected_client={
                    "id": selected_client["id"],
                    "name": selected_client["name"],
                    "phone": selected_client.get("phone"),
                    "contact_person": selected_client.get("contact_person"),
                },
                client_search_results=None,
            )
            await state.set_state(CreateOrderStates.waiting_for_order_name)
            await message.answer(
                (
                    f"‚úÖ –ó–∞–∫–∞–∑—á–∏–∫ ¬´{selected_client['name']}¬ª –≤—ã–±—Ä–∞–Ω.\n"
                    "–¢–µ–ø–µ—Ä—å –≤–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –∑–∞–∫–∞–∑–∞."
                ),
                reply_markup=CANCEL_KB,
            )
            return
    if len(text) >= 2:
        await _send_client_search_prompt(message, state, text)
        return
    await message.answer(
        "‚ÑπÔ∏è –£–∫–∞–∂–∏—Ç–µ –Ω–æ–º–µ—Ä –∏–∑ —Å–ø–∏—Å–∫–∞ –∏–ª–∏ –≤–≤–µ–¥–∏—Ç–µ –Ω–æ–≤—ã–π –∑–∞–ø—Ä–æ—Å (–º–∏–Ω–∏–º—É–º –¥–≤–∞ —Å–∏–º–≤–æ–ª–∞).",
        reply_markup=CANCEL_KB,
    )


@dp.message(CreateOrderStates.waiting_for_order_name)
async def process_order_name(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –∑–∞–∫–∞–∑–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(order_name=text)
    order_types = await fetch_order_types()
    if not order_types:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø–∏—Å–æ–∫ —Ç–∏–ø–æ–≤ –∑–∞–∫–∞–∑–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Ç–∏–ø—ã –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=ORDERS_MENU_KB,
        )
        return
    await state.update_data(order_type_options=order_types)
    options_lines = [f"{index}. {name}" for index, name in enumerate(order_types, start=1)]
    prompt_lines = [
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –∑–∞–∫–∞–∑–∞ –∏–∑ —Å–ø–∏—Å–∫–∞:",
        "",
        "\n".join(options_lines),
        "",
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–º–µ—Ä –∏–ª–∏ –Ω–∞–∑–≤–∞–Ω–∏–µ —Ç–∏–ø–∞ –∑–∞–∫–∞–∑–∞.",
    ]
    await state.set_state(CreateOrderStates.waiting_for_order_type)
    await message.answer("\n".join(prompt_lines), reply_markup=CANCEL_KB)


@dp.message(CreateOrderStates.waiting_for_order_type)
async def process_order_type(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    data = await state.get_data()
    order_types = data.get("order_type_options") or []
    if not order_types:
        order_types = await fetch_order_types()
        if not order_types:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –°–ø–∏—Å–æ–∫ —Ç–∏–ø–æ–≤ –∑–∞–∫–∞–∑–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Ç–∏–ø—ã –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
                reply_markup=ORDERS_MENU_KB,
            )
            return
        await state.update_data(order_type_options=order_types)
    selected_type: Optional[str] = None
    if text.isdigit():
        index = int(text)
        if 1 <= index <= len(order_types):
            selected_type = order_types[index - 1]
    else:
        lowered = text.lower()
        for option in order_types:
            if option.lower() == lowered:
                selected_type = option
                break
    if not selected_type:
        options_lines = "\n".join(
            f"{index}. {name}" for index, name in enumerate(order_types, start=1)
        )
        await message.answer(
            (
                "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ç–∏–ø –∑–∞–∫–∞–∑–∞.\n"
                "–í—ã–±–µ—Ä–∏—Ç–µ –Ω–æ–º–µ—Ä –∏–∑ —Å–ø–∏—Å–∫–∞ –∏–ª–∏ —É–∫–∞–∂–∏—Ç–µ —Ç–æ—á–Ω–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ.\n\n"
                f"–î–æ—Å—Ç—É–ø–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã:\n{options_lines}"
            ),
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(order_type=selected_type)
    await state.set_state(CreateOrderStates.waiting_for_folder_path)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –ø—É—Ç—å –∫ –ø–∞–ø–∫–µ –∑–∞–∫–∞–∑–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä, —Å–µ—Ç–µ–≤–æ–π –ø—É—Ç—å –∏–ª–∏ –∫–∞—Ç–∞–ª–æ–≥ –Ω–∞ –¥–∏—Å–∫–µ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(CreateOrderStates.waiting_for_folder_path)
async def process_order_folder_path(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ü—É—Ç—å –∫ –ø–∞–ø–∫–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –ø—É—Ç—å.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(order_folder_path=text)
    await state.set_state(CreateOrderStates.waiting_for_due_date)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –¥–∞—Ç—É, –∫–æ–≥–¥–∞ –∑–∞–∫–∞–∑ –¥–æ–ª–∂–µ–Ω –±—ã—Ç—å –≤—ã–ø–æ–ª–Ω–µ–Ω (—Ñ–æ—Ä–º–∞—Ç –î–î.–ú–ú.–ì–ì–ì–ì).",
        reply_markup=CANCEL_KB,
    )


@dp.message(CreateOrderStates.waiting_for_due_date)
async def process_order_due_date(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    due_date = _parse_due_date_input(text)
    if not due_date:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å –¥–∞—Ç—É. –£–∫–∞–∂–∏—Ç–µ –¥–∞—Ç—É –≤ —Ñ–æ—Ä–º–∞—Ç–µ –î–î.–ú–ú.–ì–ì–ì–ì.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(order_due_date=due_date)
    await state.set_state(CreateOrderStates.waiting_for_urgency)
    await message.answer(
        "–≠—Ç–æ —Å—Ä–æ—á–Ω—ã–π –∑–∞–∫–∞–∑? –û—Ç–≤–µ—Ç—å—Ç–µ ¬´–î–∞¬ª –∏–ª–∏ ¬´–ù–µ—Ç¬ª.",
        reply_markup=ORDER_URGENCY_KB,
    )


@dp.message(CreateOrderStates.waiting_for_urgency)
async def process_order_urgency(message: Message, state: FSMContext) -> None:
    text_raw = message.text or ""
    text = text_raw.strip().lower()
    if text_raw.strip() == CANCEL_TEXT:
        await _cancel_create_order_flow(message, state)
        return
    if text in {ORDER_URGENCY_YES_TEXT.lower(), "yes", "y"}:
        is_urgent = True
    elif text in {ORDER_URGENCY_NO_TEXT.lower(), "no", "n"}:
        is_urgent = False
    else:
        await message.answer(
            "‚ö†Ô∏è –û—Ç–≤–µ—Ç –Ω–µ —Ä–∞—Å–ø–æ–∑–Ω–∞–Ω. –£–∫–∞–∂–∏—Ç–µ ¬´–î–∞¬ª –∏–ª–∏ ¬´–ù–µ—Ç¬ª.",
            reply_markup=ORDER_URGENCY_KB,
        )
        return
    data = await state.get_data()
    selected_client = data.get("selected_client")
    if not selected_client:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –∑–∞–∫–∞–∑—á–∏–∫–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–æ–∑–¥–∞—Ç—å –∑–∞–∫–∞–∑ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=ORDERS_MENU_KB,
        )
        return
    order_name = data.get("order_name")
    order_type = data.get("order_type")
    folder_path = data.get("order_folder_path")
    due_date = data.get("order_due_date")
    if not all([order_name, order_type, folder_path, due_date]):
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –î–∞–Ω–Ω—ã–µ –∑–∞–∫–∞–∑–∞ –Ω–µ–ø–æ–ª–Ω—ã–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–æ–∑–¥–∞—Ç—å –∑–∞–∫–∞–∑ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=ORDERS_MENU_KB,
        )
        return
    creator_id = message.from_user.id if message.from_user else None
    creator_name = message.from_user.full_name if message.from_user else None
    try:
        order_row = await create_order_in_db(
            client_id=int(selected_client["id"]),
            client_name=selected_client["name"],
            title=order_name,
            order_type=order_type,
            folder_path=folder_path,
            due_date=due_date,
            is_urgent=is_urgent,
            created_by_id=creator_id,
            created_by_name=creator_name,
        )
    except Exception:
        logging.exception("Failed to create order")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –∑–∞–∫–∞–∑. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=ORDERS_MENU_KB,
        )
        return
    await state.clear()
    summary = _build_order_summary_message(order_row, selected_client)
    await message.answer(summary, reply_markup=ORDERS_MENU_KB)


@dp.message(F.text == ORDERS_SETTINGS_TEXT)
async def handle_orders_settings(message: Message) -> None:
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –∑–∞–∫–∞–∑–æ–≤. –í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é:",
        reply_markup=ORDERS_SETTINGS_KB,
    )


async def send_order_type_settings_overview(message: Message) -> None:
    order_types = await fetch_order_types()
    lines = ["üìÅ –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –∑–∞–∫–∞–∑–æ–≤ ‚Üí –¢–∏–ø –∑–∞–∫–∞–∑–∞.", ""]
    if order_types:
        lines.extend(
            [
                "–°–µ–π—á–∞—Å –¥–æ—Å—Ç—É–ø–Ω—ã —Ç–∏–ø—ã –∑–∞–∫–∞–∑–æ–≤:",
                format_order_types_list(order_types),
                "",
                "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã —É–ø—Ä–∞–≤–ª—è—Ç—å —Ç–∏–ø–∞–º–∏.",
            ]
        )
    else:
        lines.extend(
            [
                "–¢–∏–ø—ã –∑–∞–∫–∞–∑–æ–≤ –µ—â—ë –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã.",
                "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É ¬´‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ç–∏–ø –∑–∞–∫–∞–∑–∞¬ª, —á—Ç–æ–±—ã —Å–æ–∑–¥–∞—Ç—å –ø–µ—Ä–≤—ã–π —Ç–∏–ø.",
            ]
        )
    await message.answer("\n".join(lines), reply_markup=ORDERS_ORDER_TYPE_KB)


@dp.message(F.text == ORDERS_SETTINGS_ORDER_TYPE_TEXT)
async def handle_orders_settings_order_type(message: Message) -> None:
    await send_order_type_settings_overview(message)


@dp.message(F.text == ORDER_TYPE_ADD_TEXT)
async def handle_order_type_add(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManageOrderTypeStates.waiting_for_new_type_name)
    order_types = await fetch_order_types()
    prompt_lines = ["–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ —Ç–∏–ø–∞ –∑–∞–∫–∞–∑–∞."]
    if order_types:
        prompt_lines.extend(
            [
                "",
                "–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:",
                format_order_types_list(order_types),
            ]
        )
    else:
        prompt_lines.extend(
            [
                "",
                "–ü–æ–∫–∞ –µ—â—ë –Ω–µ –¥–æ–±–∞–≤–ª–µ–Ω–æ –Ω–∏ –æ–¥–Ω–æ–≥–æ —Ç–∏–ø–∞ –∑–∞–∫–∞–∑–∞.",
            ]
        )
    await message.answer("\n".join(prompt_lines), reply_markup=CANCEL_KB)


@dp.message(F.text == ORDER_TYPE_DELETE_TEXT)
async def handle_order_type_delete(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await message.answer(
        "‚ûñ –£–¥–∞–ª–µ–Ω–∏–µ —Ç–∏–ø–æ–≤ –∑–∞–∫–∞–∑–æ–≤ –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–µ.",
        reply_markup=ORDERS_ORDER_TYPE_KB,
    )


@dp.message(ManageOrderTypeStates.waiting_for_new_type_name)
async def process_new_order_type(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_order_type(name):
        await message.answer(f"‚úÖ –¢–∏–ø –∑–∞–∫–∞–∑–∞ ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –¢–∏–ø –∑–∞–∫–∞–∑–∞ ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_order_type_settings_overview(message)


@dp.message(F.text == ORDER_TYPE_BACK_TEXT)
async def handle_back_to_orders_settings(message: Message) -> None:
    await message.answer(
        "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ –∑–∞–∫–∞–∑–æ–≤. –í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é:",
        reply_markup=ORDERS_SETTINGS_KB,
    )


@dp.message(F.text == ORDERS_SETTINGS_BACK_TEXT)
async def handle_back_to_orders(message: Message) -> None:
    await message.answer(
        "üßæ –†–∞–∑–¥–µ–ª ¬´–ó–∞–∫–∞–∑—ã¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=ORDERS_MENU_KB,
    )


@dp.message(F.text == "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏")
async def handle_back_to_settings(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await handle_settings(message)


@dp.message(F.text == "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ —Å–∫–ª–∞–¥—É")
async def handle_back_to_warehouse(message: Message, state: FSMContext) -> None:
    await state.clear()
    await handle_warehouse_menu(message)


# === –°–∫–ª–∞–¥ ===
@dp.message(F.text == "üè¢ –°–∫–ª–∞–¥")
async def handle_warehouse_menu(message: Message) -> None:
    await message.answer("üè¢ –°–∫–ª–∞–¥. –í—ã–±–µ—Ä–∏—Ç–µ —Ä–∞–∑–¥–µ–ª:", reply_markup=WAREHOUSE_MENU_KB)


@dp.message(F.text == "üß± –ü–ª–∞—Å—Ç–∏–∫–∏")
async def handle_warehouse_plastics(message: Message) -> None:
    await message.answer("üì¶ –†–∞–∑–¥–µ–ª ¬´–ü–ª–∞—Å—Ç–∏–∫–∏¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:", reply_markup=WAREHOUSE_PLASTICS_KB)


@dp.message(F.text == "üéûÔ∏è –ü–ª–µ–Ω–∫–∏")
async def handle_warehouse_films(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "üéûÔ∏è –†–∞–∑–¥–µ–ª ¬´–ü–ª–µ–Ω–∫–∏¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(F.text == WAREHOUSE_ELECTRICS_TEXT)
async def handle_warehouse_electrics(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "‚ö° –†–∞–∑–¥–µ–ª ¬´–≠–ª–µ–∫—Ç—Ä–∏–∫–∞¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é:",
        reply_markup=WAREHOUSE_ELECTRICS_KB,
    )


@dp.message(F.text == WAREHOUSE_ELECTRICS_LED_STRIPS_TEXT)
async def handle_warehouse_electrics_led_strips(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    await message.answer(
        "üí° –†–∞–∑–¥–µ–ª ¬´Led –ª–µ–Ω—Ç–∞¬ª. –§—É–Ω–∫—Ü–∏–æ–Ω–∞–ª –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–µ.",
        reply_markup=WAREHOUSE_ELECTRICS_KB,
    )


@dp.message(F.text == WAREHOUSE_ELECTRICS_LED_MODULES_TEXT)
async def handle_warehouse_electrics_led_modules(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    await message.answer(
        "üß© –†–∞–∑–¥–µ–ª ¬´Led –º–æ–¥—É–ª–∏¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=WAREHOUSE_LED_MODULES_KB,
    )


@dp.message(F.text == WAREHOUSE_LED_MODULES_STOCK_TEXT)
async def handle_led_module_stock(message: Message, state: FSMContext) -> None:
    await state.clear()
    stock = await fetch_led_module_stock_summary()
    if not stock:
        await message.answer(
            "‚ÑπÔ∏è –ù–∞ —Å–∫–ª–∞–¥–µ –ø–æ–∫–∞ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π. –î–æ–±–∞–≤—å—Ç–µ –ø–æ–∑–∏—Ü–∏–∏ —á–µ—Ä–µ–∑ –∫–Ω–æ–ø–∫—É "
            f"¬´{WAREHOUSE_LED_MODULES_ADD_TEXT}¬ª.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    lines = []
    for item in stock:
        details = (
            f"{item['manufacturer']} / {item['series']} / {item['color']}, "
            f"{item['lens_count']} –ª–∏–Ω–∑, {item['power']} / {item['voltage']}"
        )
        lines.append(
            f"‚Ä¢ {item['article']} ‚Äî {item['total_quantity']} —à—Ç. ({details})"
        )
    await message.answer(
        "üì¶ –û—Å—Ç–∞—Ç–æ–∫ Led –º–æ–¥—É–ª–µ–π –Ω–∞ —Å–∫–ª–∞–¥–µ:\n\n" + "\n".join(lines),
        reply_markup=WAREHOUSE_LED_MODULES_KB,
    )


@dp.message(F.text == WAREHOUSE_LED_MODULES_ADD_TEXT)
async def handle_add_warehouse_led_modules(message: Message, state: FSMContext) -> None:
    await state.clear()
    modules = await fetch_generated_led_modules_with_details()
    if not modules:
        await message.answer(
            "‚ÑπÔ∏è –í –±–∞–∑–µ –ø–æ–∫–∞ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö —á–µ—Ä–µ–∑ ¬´‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí "
            "–≠–ª–µ–∫—Ç—Ä–∏–∫–∞ ‚Üí Led –º–æ–¥—É–ª–∏ ‚Üí Led –º–æ–¥—É–ª–∏ baza¬ª.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    overview_lines = []
    for module in modules:
        line = (
            f"‚Ä¢ {module['article']} ‚Äî {module['manufacturer']} / {module['series']} / {module['color']}, "
            f"{module['lens_count']} –ª–∏–Ω–∑, {module['power']} / {module['voltage']}"
        )
        overview_lines.append(line)
    await state.set_state(AddWarehouseLedModuleStates.waiting_for_module)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ Led –º–æ–¥—É–ª—å, –∫–æ—Ç–æ—Ä—ã–π –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å –Ω–∞ —Å–∫–ª–∞–¥.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø–æ–∑–∏—Ü–∏–∏:\n"
        + "\n".join(overview_lines),
        reply_markup=build_led_module_articles_keyboard([module["article"] for module in modules]),
    )


@dp.message(AddWarehouseLedModuleStates.waiting_for_module)
async def process_add_led_module_selection(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_led_module_flow(message, state)
        return
    if not text:
        modules = await fetch_generated_led_modules_with_details()
        if not modules:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –í –±–∞–∑–µ –ø–æ–∫–∞ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö —á–µ—Ä–µ–∑ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏.",
                reply_markup=WAREHOUSE_LED_MODULES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –£–∫–∞–∂–∏—Ç–µ –∞—Ä—Ç–∏–∫—É–ª Led –º–æ–¥—É–ª—è, –∏—Å–ø–æ–ª—å–∑—É—è –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ.",
            reply_markup=build_led_module_articles_keyboard([module["article"] for module in modules]),
        )
        return
    module = await get_generated_led_module_by_article(text)
    if module is None:
        modules = await fetch_generated_led_modules_with_details()
        if not modules:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –í –±–∞–∑–µ –ø–æ–∫–∞ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö —á–µ—Ä–µ–∑ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏.",
                reply_markup=WAREHOUSE_LED_MODULES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è Led –º–æ–¥—É–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_led_module_articles_keyboard([module["article"] for module in modules]),
        )
        return
    await state.update_data(
        selected_led_module_id=module["id"],
        selected_led_module_article=module["article"],
    )
    await state.set_state(AddWarehouseLedModuleStates.waiting_for_quantity)
    await message.answer(
        f"–£–∫–∞–∂–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–ª—è –∞—Ä—Ç–∏–∫—É–ª–∞ {module['article']} (–ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω–æ–µ —á–∏—Å–ª–æ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehouseLedModuleStates.waiting_for_quantity)
async def process_add_led_module_quantity(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_led_module_flow(message, state)
        return
    quantity = parse_positive_integer(text)
    if quantity is None:
        await message.answer(
            "‚ö†Ô∏è –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–æ–ª–∂–Ω–æ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    data = await state.get_data()
    module_id = data.get("selected_led_module_id")
    article = data.get("selected_led_module_article")
    if module_id is None or article is None:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–π Led –º–æ–¥—É–ª—å. –ù–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    added_by_id = message.from_user.id if message.from_user else None
    added_by_name = message.from_user.full_name if message.from_user else None
    try:
        record = await insert_warehouse_led_module_record(
            led_module_id=module_id,
            article=article,
            quantity=quantity,
            added_by_id=added_by_id,
            added_by_name=added_by_name,
        )
    except Exception:
        logging.exception("Failed to insert warehouse led module record")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –∑–∞–ø–∏—Å—å. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    if not record:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –∑–∞–ø–∏—Å—å. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    details = await get_generated_led_module_details(module_id)
    await state.clear()
    details_lines: list[str] = []
    if details:
        details_lines.extend(
            [
                f"–ê—Ä—Ç–∏–∫—É–ª: {details['article']}",
                f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {details['manufacturer']}",
                f"–°–µ—Ä–∏—è: {details['series']}",
                f"–¶–≤–µ—Ç: {details['color']}",
                f"–õ–∏–Ω–∑: {details['lens_count']}",
                f"–ú–æ—â–Ω–æ—Å—Ç—å: {details['power']}",
                f"–ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ: {details['voltage']}",
            ]
        )
    else:
        details_lines.append(f"–ê—Ä—Ç–∏–∫—É–ª: {article}")
    quantity_value = record.get("quantity", quantity)
    summary_employee = record.get("added_by_name") or added_by_name or "‚Äî"
    added_at_text = _format_datetime(record.get("added_at"))
    details_lines.append(f"–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ: {quantity_value} —à—Ç")
    details_lines.append(f"–î–æ–±–∞–≤–∏–ª: {summary_employee}")
    details_lines.append(f"–î–æ–±–∞–≤–ª–µ–Ω–æ: {added_at_text}")
    await message.answer(
        "‚úÖ Led –º–æ–¥—É–ª—å –¥–æ–±–∞–≤–ª–µ–Ω –Ω–∞ —Å–∫–ª–∞–¥.\n\n" + "\n".join(details_lines),
        reply_markup=WAREHOUSE_LED_MODULES_KB,
    )


@dp.message(F.text == WAREHOUSE_LED_MODULES_WRITE_OFF_TEXT)
async def handle_write_off_warehouse_led_modules(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    stock = await fetch_led_module_stock_summary()
    available_modules = [
        item for item in stock if int(item.get("total_quantity") or 0) > 0
    ]
    if not available_modules:
        await message.answer(
            "‚ÑπÔ∏è –ù–∞ —Å–∫–ª–∞–¥–µ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    overview_lines: list[str] = []
    for item in available_modules:
        details = (
            f"{item['manufacturer']} / {item['series']} / {item['color']}, "
            f"{item['lens_count']} –ª–∏–Ω–∑, {item['power']} / {item['voltage']}"
        )
        overview_lines.append(
            f"‚Ä¢ {item['article']} ‚Äî {details}. –û—Å—Ç–∞—Ç–æ–∫: {item['total_quantity']} —à—Ç"
        )
    await state.set_state(WriteOffWarehouseLedModuleStates.waiting_for_module)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ Led –º–æ–¥—É–ª—å, –∫–æ—Ç–æ—Ä—ã–π –Ω—É–∂–Ω–æ —Å–ø–∏—Å–∞—Ç—å —Å–æ —Å–∫–ª–∞–¥–∞.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –ø–æ–∑–∏—Ü–∏–∏:\n" + "\n".join(overview_lines),
        reply_markup=build_led_module_articles_keyboard(
            [item["article"] for item in available_modules]
        ),
    )


@dp.message(WriteOffWarehouseLedModuleStates.waiting_for_module)
async def process_write_off_led_module_selection(
    message: Message, state: FSMContext
) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_write_off_led_module_flow(message, state)
        return
    if not text:
        stock = await fetch_led_module_stock_summary()
        available_modules = [
            item for item in stock if int(item.get("total_quantity") or 0) > 0
        ]
        if not available_modules:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –ù–∞ —Å–∫–ª–∞–¥–µ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è.",
                reply_markup=WAREHOUSE_LED_MODULES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –í—ã–±–µ—Ä–∏—Ç–µ –∞—Ä—Ç–∏–∫—É–ª Led –º–æ–¥—É–ª—è –∏–∑ —Å–ø–∏—Å–∫–∞ –Ω–∏–∂–µ.",
            reply_markup=build_led_module_articles_keyboard(
                [item["article"] for item in available_modules]
            ),
        )
        return
    module = await get_generated_led_module_by_article(text)
    if module is None:
        stock = await fetch_led_module_stock_summary()
        available_modules = [
            item for item in stock if int(item.get("total_quantity") or 0) > 0
        ]
        if not available_modules:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –ù–∞ —Å–∫–ª–∞–¥–µ –Ω–µ—Ç Led –º–æ–¥—É–ª–µ–π –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è.",
                reply_markup=WAREHOUSE_LED_MODULES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è Led –º–æ–¥—É–ª—å —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_led_module_articles_keyboard(
                [item["article"] for item in available_modules]
            ),
        )
        return
    available_quantity = await get_led_module_stock_quantity(module["id"])
    if available_quantity <= 0:
        await message.answer(
            "‚ÑπÔ∏è –£–∫–∞–∑–∞–Ω–Ω—ã–π Led –º–æ–¥—É–ª—å –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –Ω–∞ —Å–∫–ª–∞–¥–µ. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        await state.clear()
        return
    await state.update_data(
        selected_led_module_id=module["id"],
        selected_led_module_article=module["article"],
        available_quantity=available_quantity,
    )
    await state.set_state(WriteOffWarehouseLedModuleStates.waiting_for_quantity)
    await message.answer(
        f"–£–∫–∞–∂–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è (–¥–æ—Å—Ç—É–ø–Ω–æ {available_quantity} —à—Ç).",
        reply_markup=CANCEL_KB,
    )


@dp.message(WriteOffWarehouseLedModuleStates.waiting_for_quantity)
async def process_write_off_led_module_quantity(
    message: Message, state: FSMContext
) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_write_off_led_module_flow(message, state)
        return
    quantity = parse_positive_integer(text)
    if quantity is None:
        await message.answer(
            "‚ö†Ô∏è –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–æ–ª–∂–Ω–æ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    data = await state.get_data()
    available_quantity = int(data.get("available_quantity") or 0)
    if available_quantity <= 0:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –£–∫–∞–∑–∞–Ω–Ω—ã–π Led –º–æ–¥—É–ª—å –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –Ω–∞ —Å–∫–ª–∞–¥–µ.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    if quantity > available_quantity:
        await message.answer(
            f"‚ö†Ô∏è –î–ª—è —Å–ø–∏—Å–∞–Ω–∏—è –¥–æ—Å—Ç—É–ø–Ω–æ —Ç–æ–ª—å–∫–æ {available_quantity} —à—Ç. –£–∫–∞–∂–∏—Ç–µ –¥—Ä—É–≥–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(write_off_quantity=quantity)
    await state.set_state(WriteOffWarehouseLedModuleStates.waiting_for_project)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –∑–∞–∫–∞–∑, –¥–ª—è –∫–æ—Ç–æ—Ä–æ–≥–æ —Å–ø–∏—Å—ã–≤–∞—é—Ç—Å—è Led –º–æ–¥—É–ª–∏.",
        reply_markup=CANCEL_KB,
    )


@dp.message(WriteOffWarehouseLedModuleStates.waiting_for_project)
async def process_write_off_led_module_project(
    message: Message, state: FSMContext
) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_write_off_led_module_flow(message, state)
        return
    project = text
    if not project:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –∑–∞–∫–∞–∑–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –∑–∞–∫–∞–∑.",
            reply_markup=CANCEL_KB,
        )
        return
    data = await state.get_data()
    module_id = data.get("selected_led_module_id")
    article = data.get("selected_led_module_article")
    quantity = data.get("write_off_quantity")
    if module_id is None or article is None or quantity is None:
        await _cancel_write_off_led_module_flow(message, state)
        return
    written_off_by_id = message.from_user.id if message.from_user else None
    written_off_by_name = message.from_user.full_name if message.from_user else None
    try:
        result = await write_off_led_module_stock(
            led_module_id=int(module_id),
            article=str(article),
            quantity=int(quantity),
            project=project,
            written_off_by_id=written_off_by_id,
            written_off_by_name=written_off_by_name,
        )
    except Exception:
        logging.exception("Failed to write off led modules")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–ø–∏—Å–∞—Ç—å Led –º–æ–¥—É–ª–∏. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_LED_MODULES_KB,
        )
        return
    if result is None:
        current_available = await get_led_module_stock_quantity(int(module_id))
        if current_available <= 0:
            await state.clear()
            await message.answer(
                "‚ÑπÔ∏è –£–∫–∞–∑–∞–Ω–Ω—ã–π Led –º–æ–¥—É–ª—å –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –Ω–∞ —Å–∫–ª–∞–¥–µ.",
                reply_markup=WAREHOUSE_LED_MODULES_KB,
            )
            return
        await state.update_data(available_quantity=current_available)
        await state.set_state(WriteOffWarehouseLedModuleStates.waiting_for_quantity)
        await message.answer(
            f"‚ö†Ô∏è –î–ª—è —Å–ø–∏—Å–∞–Ω–∏—è –¥–æ—Å—Ç—É–ø–Ω–æ —Ç–æ–ª—å–∫–æ {current_available} —à—Ç. –£–∫–∞–∂–∏—Ç–µ –¥—Ä—É–≥–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ.",
            reply_markup=CANCEL_KB,
        )
        return
    details = await get_generated_led_module_details(int(module_id))
    remaining_quantity = await get_led_module_stock_quantity(int(module_id))
    await state.clear()
    summary_lines: list[str] = []
    if details:
        summary_lines.extend(
            [
                f"–ê—Ä—Ç–∏–∫—É–ª: {details['article']}",
                f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {details['manufacturer']}",
                f"–°–µ—Ä–∏—è: {details['series']}",
                f"–¶–≤–µ—Ç: {details['color']}",
                f"–õ–∏–Ω–∑: {details['lens_count']}",
                f"–ú–æ—â–Ω–æ—Å—Ç—å: {details['power']}",
                f"–ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ: {details['voltage']}",
            ]
        )
    else:
        summary_lines.append(f"–ê—Ä—Ç–∏–∫—É–ª: {article}")
    summary_lines.append(f"–°–ø–∏—Å–∞–Ω–æ: {quantity} —à—Ç")
    summary_lines.append(f"–û—Å—Ç–∞—Ç–æ–∫ –Ω–∞ —Å–∫–ª–∞–¥–µ: {remaining_quantity} —à—Ç")
    summary_lines.append(f"–ó–∞–∫–∞–∑: {project}")
    summary_lines.append(
        f"–°–ø–∏—Å–∞–ª: {result.get('written_off_by_name') or written_off_by_name or '‚Äî'}"
    )
    summary_lines.append(
        f"–î–∞—Ç–∞ —Å–ø–∏—Å–∞–Ω–∏—è: {_format_datetime(result.get('written_off_at'))}"
    )
    await message.answer(
        "‚úÖ Led –º–æ–¥—É–ª–∏ —Å–ø–∏—Å–∞–Ω—ã —Å–æ —Å–∫–ª–∞–¥–∞.\n\n" + "\n".join(summary_lines),
        reply_markup=WAREHOUSE_LED_MODULES_KB,
    )


@dp.message(F.text == WAREHOUSE_LED_MODULES_BACK_TO_ELECTRICS_TEXT)
async def handle_back_to_electrics_from_led_modules(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    await message.answer(
        "‚ö° –†–∞–∑–¥–µ–ª ¬´–≠–ª–µ–∫—Ç—Ä–∏–∫–∞¬ª. –í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é:",
        reply_markup=WAREHOUSE_ELECTRICS_KB,
    )


@dp.message(F.text == WAREHOUSE_ELECTRICS_POWER_SUPPLIES_TEXT)
async def handle_warehouse_electrics_power_supplies(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    await message.answer(
        "üîå –†–∞–∑–¥–µ–ª ¬´–ë–ª–æ–∫–∏ –ø–∏—Ç–∞–Ω–∏—è¬ª. –§—É–Ω–∫—Ü–∏–æ–Ω–∞–ª –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–µ.",
        reply_markup=WAREHOUSE_ELECTRICS_KB,
    )


async def _reply_films_feature_in_development(message: Message, feature: str) -> None:
    await message.answer(
        f"‚öôÔ∏è –§—É–Ω–∫—Ü–∏—è ¬´{feature}¬ª –¥–ª—è —Ä–∞–∑–¥–µ–ª–∞ ¬´–ü–ª–µ–Ω–∫–∏¬ª –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–µ.",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(F.text == WAREHOUSE_FILMS_ADD_TEXT)
async def handle_add_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    manufacturers = await fetch_film_manufacturers()
    if not manufacturers:
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –¥–∞–Ω–Ω—ã–µ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.set_state(AddWarehouseFilmStates.waiting_for_article)
    suggested_article: Optional[str] = None
    last_article = await fetch_max_film_article()
    if last_article is not None:
        suggested_article = str(last_article + 1)
    await state.update_data(article_suggestion=suggested_article)
    prompt_lines = ["–í–≤–µ–¥–∏—Ç–µ –∞—Ä—Ç–∏–∫—É–ª –ø–ª–µ–Ω–∫–∏ (—Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã)."]
    if last_article is not None and suggested_article is not None:
        prompt_lines.append("")
        prompt_lines.append(
            "–ü–æ—Å–ª–µ–¥–Ω–∏–π –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–π –∞—Ä—Ç–∏–∫—É–ª: "
            f"{last_article}. –ù–∞–∂–º–∏—Ç–µ –∫–Ω–æ–ø–∫—É –Ω–∏–∂–µ, —á—Ç–æ–±—ã –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Å–ª–µ–¥—É—é—â–∏–π –Ω–æ–º–µ—Ä."
        )
    await message.answer(
        "\n".join(prompt_lines),
        reply_markup=build_article_input_keyboard(suggested_article),
    )


@dp.message(F.text == WAREHOUSE_FILMS_WRITE_OFF_TEXT)
async def handle_write_off_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(WriteOffWarehouseFilmStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã —Å–ø–∏—Å–∞—Ç—å –ø–ª–µ–Ω–∫—É —Å–æ —Å–∫–ª–∞–¥–∞.",
        reply_markup=CANCEL_KB,
    )


@dp.message(WriteOffWarehouseFilmStates.waiting_for_article)
async def process_write_off_film_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_write_off_film_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_film_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∞ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(film_id=record["id"], article=record.get("article"))
    formatted = format_film_record_for_message(record)
    await state.set_state(WriteOffWarehouseFilmStates.waiting_for_project)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{formatted}\n\n"
        "–£–∫–∞–∂–∏—Ç–µ –ø—Ä–æ–µ–∫—Ç, –Ω–∞ –∫–æ—Ç–æ—Ä—ã–π –≤—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è —Å–ø–∏—Å–∞–Ω–∏–µ.",
        reply_markup=CANCEL_KB,
    )


@dp.message(WriteOffWarehouseFilmStates.waiting_for_project)
async def process_write_off_film_project(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_write_off_film_flow(message, state)
        return
    project = (message.text or "").strip()
    if not project:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–µ–∫—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –ø—Ä–æ–µ–∫—Ç.",
            reply_markup=CANCEL_KB,
        )
        return
    data = await state.get_data()
    record_id = data.get("film_id")
    article = data.get("article")
    if record_id is None or article is None:
        await _cancel_write_off_film_flow(message, state)
        return
    written_off_by_id = message.from_user.id if message.from_user else None
    written_off_by_name = message.from_user.full_name if message.from_user else None
    try:
        result = await write_off_warehouse_film(
            record_id=record_id,
            project=project,
            written_off_by_id=written_off_by_id,
            written_off_by_name=written_off_by_name,
        )
    except Exception:
        logging.exception("Failed to write off film record")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–ø–∏—Å–∞—Ç—å –ø–ª–µ–Ω–∫—É. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    if result is None:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ù–µ —É–¥–∞–ª–æ—Å—å –Ω–∞–π—Ç–∏ –∑–∞–ø–∏—Å—å –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è. –í–æ–∑–º–æ–∂–Ω–æ, –æ–Ω–∞ —É–∂–µ –±—ã–ª–∞ –∏–∑–º–µ–Ω–µ–Ω–∞.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.clear()
    formatted = format_written_off_film_record(result)
    await message.answer(
        "‚úÖ –ü–ª–µ–Ω–∫–∞ —Å–ø–∏—Å–∞–Ω–∞ —Å–æ —Å–∫–ª–∞–¥–∞.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ü—Ä–æ–µ–∫—Ç: {project}\n\n"
        f"–î–∞–Ω–Ω—ã–µ —Å–ø–∏—Å–∞–Ω–Ω–æ–π –∑–∞–ø–∏—Å–∏:\n{formatted}",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(F.text == WAREHOUSE_FILMS_COMMENT_TEXT)
async def handle_comment_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(CommentWarehouseFilmStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã –ø—Ä–æ—Å–º–æ—Ç—Ä–µ—Ç—å –∏ –∏–∑–º–µ–Ω–∏—Ç—å –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π –∫ –ø–ª–µ–Ω–∫–µ.",
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == WAREHOUSE_FILMS_MOVE_TEXT)
async def handle_move_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    locations = await fetch_film_storage_locations()
    if not locations:
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.set_state(MoveWarehouseFilmStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã –≤—ã–±—Ä–∞—Ç—å –Ω–æ–≤–æ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è.",
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == WAREHOUSE_FILMS_SEARCH_TEXT)
async def handle_search_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(SearchWarehouseFilmStates.choosing_mode)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø–∞—Ä–∞–º–µ—Ç—Ä –ø–æ–∏—Å–∫–∞:",
        reply_markup=WAREHOUSE_FILMS_SEARCH_KB,
    )


@dp.message(F.text == WAREHOUSE_FILMS_EXPORT_TEXT)
async def handle_export_warehouse_film(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer("‚è≥ –§–æ—Ä–º–∏—Ä—É—é —Ñ–∞–π–ª —ç–∫—Å–ø–æ—Ä—Ç–∞. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–æ–∂–¥–∏—Ç–µ...")
    try:
        records = await fetch_all_warehouse_films()
    except Exception:
        logging.exception("Failed to fetch films for export")
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —Å–∫–ª–∞–¥–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return

    if not records:
        await message.answer(
            "‚ÑπÔ∏è –ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return

    try:
        export_file = build_films_export_file(records)
    except Exception:
        logging.exception("Failed to build films export file")
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞—Ç—å —Ñ–∞–π–ª —ç–∫—Å–ø–æ—Ä—Ç–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return

    await message.answer_document(
        document=export_file,
        caption="üìÑ –≠–∫—Å–ø–æ—Ä—Ç –ø–ª–µ–Ω–æ–∫",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(SearchWarehouseFilmStates.choosing_mode)
async def process_search_film_menu(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_film_flow(message, state)
        return
    if text == WAREHOUSE_FILMS_SEARCH_BACK_TEXT:
        await state.clear()
        await message.answer(
            "–í—ã –≤–µ—Ä–Ω—É–ª–∏—Å—å –≤ —Ä–∞–∑–¥–µ–ª ¬´–ü–ª–µ–Ω–∫–∏¬ª.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    if text == WAREHOUSE_FILMS_SEARCH_BY_ARTICLE_TEXT:
        await state.set_state(SearchWarehouseFilmStates.waiting_for_article)
        await message.answer(
            "–í–≤–µ–¥–∏—Ç–µ –∞—Ä—Ç–∏–∫—É–ª –ø–ª–µ–Ω–∫–∏ –¥–ª—è –ø–æ–∏—Å–∫–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if text == WAREHOUSE_FILMS_SEARCH_BY_NUMBER_TEXT:
        await state.set_state(SearchWarehouseFilmStates.waiting_for_number)
        await message.answer(
            "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –ø–ª–µ–Ω–∫–∏ (–∫–æ–¥ —Ü–≤–µ—Ç–∞).",
            reply_markup=CANCEL_KB,
        )
        return
    if text == WAREHOUSE_FILMS_SEARCH_BY_COLOR_TEXT:
        await state.set_state(SearchWarehouseFilmStates.waiting_for_color)
        await message.answer(
            "–£–∫–∞–∂–∏—Ç–µ —Ü–≤–µ—Ç –∏–ª–∏ –µ–≥–æ —á–∞—Å—Ç—å –¥–ª—è –ø–æ–∏—Å–∫–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await message.answer(
        "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ –º–µ–Ω—é –Ω–∏–∂–µ.",
        reply_markup=WAREHOUSE_FILMS_SEARCH_KB,
    )


@dp.message(SearchWarehouseFilmStates.waiting_for_article)
async def process_search_film_by_article(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_film_flow(message, state)
        return
    if not text.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_film_by_article(text)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∞ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    formatted = format_film_record_for_message(record)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n" f"{formatted}",
        reply_markup=CANCEL_KB,
    )
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª –¥–ª—è –Ω–æ–≤–æ–≥–æ –ø–æ–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚ùå –û—Ç–º–µ–Ω–∞¬ª.",
        reply_markup=CANCEL_KB,
    )


@dp.message(SearchWarehouseFilmStates.waiting_for_number)
async def process_search_film_by_number(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_film_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ù–æ–º–µ—Ä –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ.",
            reply_markup=CANCEL_KB,
        )
        return
    matches = await search_warehouse_films_by_color_code(
        text, limit=FILM_SEARCH_RESULTS_LIMIT
    )
    if matches:
        if len(matches) == 1:
            formatted = format_film_record_for_message(matches[0])
            await message.answer(
                "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n" f"{formatted}",
                reply_markup=CANCEL_KB,
            )
        else:
            formatted_list = format_film_records_list_for_message(matches)
            header = [f"–ù–∞–π–¥–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {len(matches)}."]
            if len(matches) == FILM_SEARCH_RESULTS_LIMIT:
                header.append(
                    f"–ü–æ–∫–∞–∑–∞–Ω—ã –ø–µ—Ä–≤—ã–µ {FILM_SEARCH_RESULTS_LIMIT} –∑–∞–ø–∏—Å–µ–π. –£—Ç–æ—á–Ω–∏—Ç–µ –∑–∞–ø—Ä–æ—Å –¥–ª—è –±–æ–ª–µ–µ —Ç–æ—á–Ω–æ–≥–æ –ø–æ–∏—Å–∫–∞."
                )
            await message.answer(
                " ".join(header) + "\n\n" + formatted_list,
                reply_markup=CANCEL_KB,
            )
        await message.answer(
            "–í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π –Ω–æ–º–µ—Ä (–∫–æ–¥ —Ü–≤–µ—Ç–∞) –¥–ª—è –ø–æ–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚ùå –û—Ç–º–µ–Ω–∞¬ª.",
            reply_markup=CANCEL_KB,
        )
        return
    if text.isdigit():
        record_id = int(text)
        record = await fetch_warehouse_film_by_id(record_id)
        if record is not None:
            formatted = format_film_record_for_message(record)
            await message.answer(
                "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n" f"{formatted}",
                reply_markup=CANCEL_KB,
            )
            await message.answer(
                "–í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π –Ω–æ–º–µ—Ä (–∫–æ–¥ —Ü–≤–µ—Ç–∞) –¥–ª—è –ø–æ–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚ùå –û—Ç–º–µ–Ω–∞¬ª.",
                reply_markup=CANCEL_KB,
            )
            return
    await message.answer(
        "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∏ —Å —Ç–∞–∫–∏–º –Ω–æ–º–µ—Ä–æ–º (–∫–æ–¥–æ–º —Ü–≤–µ—Ç–∞) –Ω–µ –Ω–∞–π–¥–µ–Ω—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —É–∫–∞–∑–∞—Ç—å –¥—Ä—É–≥–æ–π –Ω–æ–º–µ—Ä.",
        reply_markup=CANCEL_KB,
    )


@dp.message(SearchWarehouseFilmStates.waiting_for_color)
async def process_search_film_by_color(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_film_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –¶–≤–µ—Ç –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –¥–ª—è –ø–æ–∏—Å–∫–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    matches = await search_warehouse_films_by_color(
        text, limit=FILM_SEARCH_RESULTS_LIMIT
    )
    if not matches:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∏ —Å —Ç–∞–∫–∏–º —Ü–≤–µ—Ç–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∑–∞–ø—Ä–æ—Å.",
            reply_markup=CANCEL_KB,
        )
        return
    formatted_list = format_film_records_list_for_message(matches)
    header = [f"–ù–∞–π–¥–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {len(matches)}."]
    if len(matches) == FILM_SEARCH_RESULTS_LIMIT:
        header.append(
            f"–ü–æ–∫–∞–∑–∞–Ω—ã –ø–µ—Ä–≤—ã–µ {FILM_SEARCH_RESULTS_LIMIT} –∑–∞–ø–∏—Å–µ–π. –£—Ç–æ—á–Ω–∏—Ç–µ –∑–∞–ø—Ä–æ—Å –¥–ª—è –±–æ–ª–µ–µ —Ç–æ—á–Ω–æ–≥–æ –ø–æ–∏—Å–∫–∞."
        )
    await message.answer(
        " ".join(header) + "\n\n" + formatted_list,
        reply_markup=CANCEL_KB,
    )
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π —Ü–≤–µ—Ç –¥–ª—è –ø–æ–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚ùå –û—Ç–º–µ–Ω–∞¬ª.",
        reply_markup=CANCEL_KB,
    )


@dp.message(CommentWarehouseFilmStates.waiting_for_article)
async def process_film_comment_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_comment_film_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_film_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∞ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    previous_comment = record.get("comment")
    formatted = format_film_record_for_message(record)
    await state.update_data(
        film_id=record["id"],
        article=record.get("article"),
        previous_comment=previous_comment,
    )
    await state.set_state(CommentWarehouseFilmStates.waiting_for_comment)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{formatted}\n\n"
        f"–¢–µ–∫—É—â–∏–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {previous_comment or '‚Äî'}",
        reply_markup=CANCEL_KB,
    )
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–≤—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π. –ü—É—Å—Ç–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —É–¥–∞–ª–∏—Ç —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π.",
        reply_markup=CANCEL_KB,
    )


@dp.message(CommentWarehouseFilmStates.waiting_for_comment)
async def process_film_comment_update(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_comment_film_flow(message, state)
        return
    data = await state.get_data()
    record_id = data.get("film_id")
    article = data.get("article")
    previous_comment = data.get("previous_comment")
    if record_id is None or article is None:
        await _cancel_comment_film_flow(message, state)
        return
    new_comment_raw = (message.text or "").strip()
    new_comment: Optional[str]
    if new_comment_raw:
        new_comment = new_comment_raw
    else:
        new_comment = None
    updated = await update_warehouse_film_comment(record_id, new_comment)
    if not updated:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–±–Ω–æ–≤–∏—Ç—å –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        await state.clear()
        return
    await state.clear()
    await message.answer(
        "‚úÖ –ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π –æ–±–Ω–æ–≤–ª—ë–Ω.\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–°—Ç–∞—Ä—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {previous_comment or '‚Äî'}\n"
        f"–ù–æ–≤—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {new_comment or '‚Äî'}",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(MoveWarehouseFilmStates.waiting_for_article)
async def process_move_film_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_move_film_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_film_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∞ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    locations = await fetch_film_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.update_data(
        film_id=record["id"],
        article=record.get("article"),
        previous_location=record.get("warehouse"),
    )
    previous_location = record.get("warehouse") or "‚Äî"
    formatted_record = format_film_record_for_message(record)
    await state.set_state(MoveWarehouseFilmStates.waiting_for_new_location)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{formatted_record}\n\n"
        f"–¢–µ–∫—É—â–µ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è: {previous_location}\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –Ω–æ–≤–æ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –∏–∑ —Å–ø–∏—Å–∫–∞ –Ω–∏–∂–µ.",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(MoveWarehouseFilmStates.waiting_for_new_location)
async def process_move_film_new_location(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_move_film_flow(message, state)
        return
    locations = await fetch_film_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    raw_location = (message.text or "").strip()
    match = next((item for item in locations if item.lower() == raw_location.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–Ω–æ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    data = await state.get_data()
    record_id = data.get("film_id")
    article = data.get("article")
    previous_location_raw = data.get("previous_location")
    previous_location_display = previous_location_raw or "‚Äî"
    if record_id is None or article is None:
        await _cancel_move_film_flow(message, state)
        return
    if previous_location_raw and previous_location_raw.lower() == match.lower():
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–µ–Ω–∫–∞ —É–∂–µ –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ –≤—ã–±—Ä–∞–Ω–Ω–æ–º –º–µ—Å—Ç–µ. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–µ –º–µ—Å—Ç–æ.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    employee_id = message.from_user.id if message.from_user else None
    employee_nick: Optional[str] = None
    if message.from_user:
        employee_nick = message.from_user.username or message.from_user.full_name
    updated_record = await update_warehouse_film_location(
        record_id=record_id,
        new_location=match,
        employee_id=employee_id,
        employee_nick=employee_nick,
    )
    if updated_record is None:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–±–Ω–æ–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.clear()
    formatted = format_film_record_for_message(updated_record)
    await message.answer(
        "‚úÖ –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –æ–±–Ω–æ–≤–ª–µ–Ω–æ.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ü—Ä–µ–¥—ã–¥—É—â–µ–µ –º–µ—Å—Ç–æ: {previous_location_display}\n"
        f"–ù–æ–≤–æ–µ –º–µ—Å—Ç–æ: {match}\n\n"
        f"–ê–∫—Ç—É–∞–ª—å–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ:\n{formatted}",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_article)
async def process_film_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        data = await state.get_data()
        suggestion = data.get("article_suggestion")
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_article_input_keyboard(suggestion),
        )
        return
    manufacturers = await fetch_film_manufacturers()
    if not manufacturers:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –¥–∞–Ω–Ω—ã–µ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.update_data(article=article, article_suggestion=None)
    await state.set_state(AddWarehouseFilmStates.waiting_for_manufacturer)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(AddWarehouseFilmStates.waiting_for_manufacturer)
async def process_film_manufacturer(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    manufacturers = await fetch_film_manufacturers()
    raw = (message.text or "").strip()
    match = next((item for item in manufacturers if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturers),
        )
        return
    series_list = await fetch_film_series_by_manufacturer(match)
    if not series_list:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Å–µ—Ä–∏–∏. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.update_data(manufacturer=match)
    await state.set_state(AddWarehouseFilmStates.waiting_for_series)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Å–µ—Ä–∏—é:",
        reply_markup=build_series_keyboard(series_list),
    )


@dp.message(AddWarehouseFilmStates.waiting_for_series)
async def process_film_series(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    data = await state.get_data()
    manufacturer = data.get("manufacturer")
    if not manufacturer:
        await _cancel_add_film_flow(message, state)
        return
    series_list = await fetch_film_series_by_manufacturer(manufacturer)
    raw = (message.text or "").strip()
    match = next((item for item in series_list if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –°–µ—Ä–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_series_keyboard(series_list),
        )
        return
    await state.update_data(series=match)
    await state.set_state(AddWarehouseFilmStates.waiting_for_color_code)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∫–æ–¥ —Ü–≤–µ—Ç–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä, 3-45).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_color_code)
async def process_film_color_code(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –ö–æ–¥ —Ü–≤–µ—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(color_code=text)
    await state.set_state(AddWarehouseFilmStates.waiting_for_color)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ —Ü–≤–µ—Ç (–Ω–∞–ø—Ä–∏–º–µ—Ä, –ë–µ–ª—ã–π).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_color)
async def process_film_color(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    if not text:
        await message.answer(
            "‚ö†Ô∏è –¶–≤–µ—Ç –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(color=text)
    await state.set_state(AddWarehouseFilmStates.waiting_for_width)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ —à–∏—Ä–∏–Ω—É –ø–ª–µ–Ω–∫–∏ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (–º–æ–∂–Ω–æ —á–µ—Ä–µ–∑ —Ç–æ—á–∫—É –∏–ª–∏ –∑–∞–ø—è—Ç—É—é).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_width)
async def process_film_width(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    value = parse_positive_decimal(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –®–∏—Ä–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(width=value)
    await state.set_state(AddWarehouseFilmStates.waiting_for_length)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –¥–ª–∏–Ω—É –ø–ª–µ–Ω–∫–∏ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (–º–æ–∂–Ω–æ —á–µ—Ä–µ–∑ —Ç–æ—á–∫—É –∏–ª–∏ –∑–∞–ø—è—Ç—É—é).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_length)
async def process_film_length(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    value = parse_positive_decimal(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –î–ª–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    locations = await fetch_film_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö.",
            reply_markup=WAREHOUSE_FILMS_KB,
        )
        return
    await state.update_data(length=value)
    await state.set_state(AddWarehouseFilmStates.waiting_for_storage)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(AddWarehouseFilmStates.waiting_for_storage)
async def process_film_storage(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    locations = await fetch_film_storage_locations()
    raw = (message.text or "").strip()
    match = next((item for item in locations if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    await state.update_data(storage=match)
    await state.set_state(AddWarehouseFilmStates.waiting_for_comment)
    await message.answer(
        "–î–æ–±–∞–≤—å—Ç–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π (–Ω–µ–æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ) –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddWarehouseFilmStates.waiting_for_comment)
async def process_film_comment(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_film_flow(message, state)
        return
    comment: Optional[str]
    if text == SKIP_TEXT:
        comment = None
    else:
        comment = text or None
    data = await state.get_data()
    article = data.get("article")
    manufacturer = data.get("manufacturer")
    series = data.get("series")
    color_code = data.get("color_code")
    color = data.get("color")
    width: Optional[Decimal] = data.get("width")
    length: Optional[Decimal] = data.get("length")
    storage = data.get("storage")
    if not all([article, manufacturer, series, color_code, color, width, length, storage]):
        await _cancel_add_film_flow(message, state)
        return
    employee_id = message.from_user.id if message.from_user else None
    employee_nick: Optional[str] = None
    if message.from_user:
        employee_nick = message.from_user.username or message.from_user.full_name
    record = await insert_warehouse_film_record(
        article=article,
        manufacturer=manufacturer,
        series=series,
        color_code=color_code,
        color=color,
        width_mm=width,
        length_mm=length,
        warehouse=storage,
        comment=comment,
        employee_id=employee_id,
        employee_nick=employee_nick,
    )
    await state.clear()
    formatted = format_film_record_for_message(record)
    await message.answer(
        "‚úÖ –ü–ª–µ–Ω–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∞ –Ω–∞ —Å–∫–ª–∞–¥.\n\n"
        f"{formatted}",
        reply_markup=WAREHOUSE_FILMS_KB,
    )


@dp.message(F.text == "üì§ –≠–∫—Å–ø–æ—Ä—Ç")
async def handle_export_warehouse_plastics(message: Message) -> None:
    await message.answer("‚è≥ –§–æ—Ä–º–∏—Ä—É—é —Ñ–∞–π–ª —ç–∫—Å–ø–æ—Ä—Ç–∞. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–æ–∂–¥–∏—Ç–µ...")
    try:
        records = await fetch_all_warehouse_plastics()
    except Exception:
        logging.exception("Failed to fetch plastics for export")
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —Å–∫–ª–∞–¥–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return

    if not records:
        await message.answer(
            "‚ÑπÔ∏è –ù–µ—Ç –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —ç–∫—Å–ø–æ—Ä—Ç–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return

    try:
        export_file = build_plastics_export_file(records)
    except Exception:
        logging.exception("Failed to build plastics export file")
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å—Ñ–æ—Ä–º–∏—Ä–æ–≤–∞—Ç—å —Ñ–∞–π–ª —ç–∫—Å–ø–æ—Ä—Ç–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return

    await message.answer_document(
        document=export_file,
        caption="üìÑ –≠–∫—Å–ø–æ—Ä—Ç –ø–ª–∞—Å—Ç–∏–∫–æ–≤",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(F.text == "üîç –ù–∞–π—Ç–∏")
async def handle_search_warehouse_plastic(message: Message, state: FSMContext) -> None:
    await state.set_state(SearchWarehousePlasticStates.choosing_mode)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –ø–æ–∏—Å–∫–∞:",
        reply_markup=WAREHOUSE_PLASTICS_SEARCH_KB,
    )


@dp.message(SearchWarehousePlasticStates.choosing_mode)
async def process_search_menu_choice(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == BACK_TO_PLASTICS_MENU_TEXT:
        await state.clear()
        await message.answer(
            "–í—ã –≤–µ—Ä–Ω—É–ª–∏—Å—å –≤ –º–µ–Ω—é –ø–ª–∞—Å—Ç–∏–∫–æ–≤.", reply_markup=WAREHOUSE_PLASTICS_KB
        )
        return
    if text == SEARCH_BY_ARTICLE_TEXT:
        await state.set_state(SearchWarehousePlasticStates.waiting_for_article)
        await message.answer(
            "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞ –¥–ª—è –ø–æ–∏—Å–∫–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if text == ADVANCED_SEARCH_TEXT:
        await _start_advanced_search_flow(message, state)
        return
    await message.answer(
        "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ –Ω–∏–∂–µ.",
        reply_markup=WAREHOUSE_PLASTICS_SEARCH_KB,
    )


@dp.message(F.text == "üí¨ –ö–æ–º–º–µ–Ω—Ç–∏—Ä–æ–≤–∞—Ç—å")
async def handle_comment_warehouse_plastic(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(CommentWarehousePlasticStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã –ø—Ä–æ—Å–º–æ—Ç—Ä–µ—Ç—å –∏ –∏–∑–º–µ–Ω–∏—Ç—å –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π.",
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == "üîÅ –ü–µ—Ä–µ–º–µ—Å—Ç–∏—Ç—å")
async def handle_move_warehouse_plastic(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(MoveWarehousePlasticStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã –≤—ã–±—Ä–∞—Ç—å –Ω–æ–≤–æ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è.",
        reply_markup=CANCEL_KB,
    )


@dp.message(F.text == "‚ûñ –°–ø–∏—Å–∞—Ç—å")
async def handle_write_off_warehouse_plastic(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(WriteOffWarehousePlasticStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞, —á—Ç–æ–±—ã —Å–ø–∏—Å–∞—Ç—å –ø–ª–∞—Å—Ç–∏–∫ —Å–æ —Å–∫–ª–∞–¥–∞.",
        reply_markup=CANCEL_KB,
    )


@dp.message(SearchWarehousePlasticStates.waiting_for_article)
async def process_search_plastic_by_article(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if not text.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_plastic_by_article(text)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–∞—Å—Ç–∏–∫ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –Ω–æ–º–µ—Ä.",
            reply_markup=CANCEL_KB,
        )
        return
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n" f"{format_plastic_record_for_message(record)}",
        reply_markup=CANCEL_KB,
    )
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª –¥–ª—è –Ω–æ–≤–æ–≥–æ –ø–æ–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚ùå –û—Ç–º–µ–Ω–∞¬ª.",
        reply_markup=CANCEL_KB,
    )


async def _start_advanced_search_flow(message: Message, state: FSMContext) -> None:
    materials = await fetch_plastic_material_types()
    await state.update_data(
        advanced_material=None,
        advanced_thickness=None,
        advanced_color=None,
        advanced_min_length=None,
        advanced_min_width=None,
    )
    if not materials:
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ –ø—É—Å—Ç. –ü–æ–∏—Å–∫ –±—É–¥–µ—Ç –≤—ã–ø–æ–ª–Ω–µ–Ω –ø–æ –≤—Å–µ–º –º–∞—Ç–µ—Ä–∏–∞–ª–∞–º."
        )
        await _prompt_advanced_thickness_choice(message, state, None)
        return
    await state.set_state(SearchWarehousePlasticStates.waiting_for_material)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚û°Ô∏è –î–∞–ª–µ–µ¬ª, —á—Ç–æ–±—ã –∏—Å–∫–∞—Ç—å –ø–æ –≤—Å–µ–º –º–∞—Ç–µ—Ä–∏–∞–ª–∞–º.",
        reply_markup=build_advanced_materials_keyboard(materials),
    )


async def _prompt_advanced_thickness_choice(
    message: Message, state: FSMContext, material: Optional[str]
) -> None:
    if material:
        thicknesses = await fetch_material_thicknesses(material)
    else:
        thicknesses = await fetch_all_material_thicknesses()
    if not thicknesses:
        await state.update_data(advanced_thickness=None)
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ç–æ–ª—â–∏–Ω—ã. –ü–æ–∏—Å–∫ –±—É–¥–µ—Ç –≤—ã–ø–æ–ª–Ω–µ–Ω –ø–æ –≤—Å–µ–º —Ç–æ–ª—â–∏–Ω–∞–º."
        )
        await _prompt_advanced_color_choice(message, state, material)
        return
    await state.set_state(SearchWarehousePlasticStates.waiting_for_thickness)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´üìè –í—Å–µ —Ç–æ–ª—â–∏–Ω—ã¬ª.",
        reply_markup=build_advanced_thickness_keyboard(thicknesses),
    )


async def _prompt_advanced_color_choice(
    message: Message, state: FSMContext, material: Optional[str]
) -> None:
    if material:
        colors = await fetch_material_colors(material)
    else:
        colors = await fetch_all_material_colors()
    if not colors:
        await state.update_data(advanced_color=None)
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ü–≤–µ—Ç–∞. –ü–æ–∏—Å–∫ –±—É–¥–µ—Ç –≤—ã–ø–æ–ª–Ω–µ–Ω –ø–æ –≤—Å–µ–º —Ü–≤–µ—Ç–∞–º."
        )
        await _prompt_advanced_min_length(message, state)
        return
    await state.set_state(SearchWarehousePlasticStates.waiting_for_color)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´üé® –í—Å–µ —Ü–≤–µ—Ç–∞¬ª.",
        reply_markup=build_advanced_colors_keyboard(colors),
    )


async def _prompt_advanced_min_length(message: Message, state: FSMContext) -> None:
    await state.set_state(SearchWarehousePlasticStates.waiting_for_min_length)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –º–∏–Ω–∏–º–∞–ª—å–Ω—É—é –¥–ª–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


async def _prompt_advanced_min_width(message: Message, state: FSMContext) -> None:
    await state.set_state(SearchWarehousePlasticStates.waiting_for_min_width)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –º–∏–Ω–∏–º–∞–ª—å–Ω—É—é —à–∏—Ä–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


async def _perform_advanced_search(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    material = data.get("advanced_material")
    thickness = data.get("advanced_thickness")
    color = data.get("advanced_color")
    min_length = data.get("advanced_min_length")
    min_width = data.get("advanced_min_width")
    try:
        records = await search_warehouse_plastics_advanced(
            material=material,
            thickness=thickness,
            color=color,
            min_length=min_length,
            min_width=min_width,
        )
    except Exception:
        logging.exception("Failed to run advanced search for plastics")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –≤—ã–ø–æ–ª–Ω–∏—Ç—å —Ä–∞—Å—à–∏—Ä–µ–Ω–Ω—ã–π –ø–æ–∏—Å–∫. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.clear()
    if not records:
        await message.answer(
            "–ü–æ –∑–∞–¥–∞–Ω–Ω—ã–º –ø–∞—Ä–∞–º–µ—Ç—Ä–∞–º –Ω–∏—á–µ–≥–æ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
    else:
        header_parts = ["–†–µ–∑—É–ª—å—Ç–∞—Ç—ã —Ä–∞—Å—à–∏—Ä–µ–Ω–Ω–æ–≥–æ –ø–æ–∏—Å–∫–∞:"]
        if material:
            header_parts.append(f"–ú–∞—Ç–µ—Ä–∏–∞–ª: {material}")
        if thickness is not None:
            header_parts.append(f"–¢–æ–ª—â–∏–Ω–∞: {format_thickness_value(thickness)}")
        if color:
            header_parts.append(f"–¶–≤–µ—Ç: {color}")
        if min_length is not None:
            header_parts.append(f"–ú–∏–Ω. –¥–ª–∏–Ω–∞: {format_dimension_value(min_length)}")
        if min_width is not None:
            header_parts.append(f"–ú–∏–Ω. —à–∏—Ä–∏–Ω–∞: {format_dimension_value(min_width)}")
        header_text = "\n".join(header_parts)
        records_text = []
        for index, record in enumerate(records, start=1):
            records_text.append(
                f"{index}.\n{format_plastic_record_for_message(record)}"
            )
        full_text = f"{header_text}\n\n" + "\n\n".join(records_text)
        chunks = split_text_into_messages(full_text)
        for idx, chunk in enumerate(chunks):
            if idx == 0:
                await message.answer(chunk, reply_markup=WAREHOUSE_PLASTICS_KB)
            else:
                await message.answer(chunk)
    await state.set_state(SearchWarehousePlasticStates.choosing_mode)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –ø–æ–∏—Å–∫–∞:",
        reply_markup=WAREHOUSE_PLASTICS_SEARCH_KB,
    )


@dp.message(SearchWarehousePlasticStates.waiting_for_material)
async def process_advanced_search_material(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == ADVANCED_SEARCH_SKIP_MATERIAL_TEXT:
        await state.update_data(advanced_material=None)
        await _prompt_advanced_thickness_choice(message, state, None)
        return
    materials = await fetch_plastic_material_types()
    match = next((item for item in materials if item.lower() == text.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´‚û°Ô∏è –î–∞–ª–µ–µ¬ª.",
            reply_markup=build_advanced_materials_keyboard(materials),
        )
        return
    await state.update_data(advanced_material=match)
    await _prompt_advanced_thickness_choice(message, state, match)


@dp.message(SearchWarehousePlasticStates.waiting_for_thickness)
async def process_advanced_search_thickness(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == ADVANCED_SEARCH_ALL_THICKNESSES_TEXT:
        await state.update_data(advanced_thickness=None)
        data = await state.get_data()
        material = data.get("advanced_material")
        await _prompt_advanced_color_choice(message, state, material)
        return
    data = await state.get_data()
    material = data.get("advanced_material")
    if material:
        thicknesses = await fetch_material_thicknesses(material)
    else:
        thicknesses = await fetch_all_material_thicknesses()
    value = parse_thickness_input(text)
    if value is None or all(item != value for item in thicknesses):
        await message.answer(
            "‚ÑπÔ∏è –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –¥–ª—è –≤—ã–±–æ—Ä–∞ —Ç–æ–ª—â–∏–Ω—ã –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´üìè –í—Å–µ —Ç–æ–ª—â–∏–Ω—ã¬ª.",
            reply_markup=build_advanced_thickness_keyboard(thicknesses),
        )
        return
    await state.update_data(advanced_thickness=value)
    await _prompt_advanced_color_choice(message, state, material)


@dp.message(SearchWarehousePlasticStates.waiting_for_color)
async def process_advanced_search_color(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == ADVANCED_SEARCH_ALL_COLORS_TEXT:
        await state.update_data(advanced_color=None)
        await _prompt_advanced_min_length(message, state)
        return
    data = await state.get_data()
    material = data.get("advanced_material")
    if material:
        colors = await fetch_material_colors(material)
    else:
        colors = await fetch_all_material_colors()
    match = next((item for item in colors if item.lower() == text.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¶–≤–µ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´üé® –í—Å–µ —Ü–≤–µ—Ç–∞¬ª.",
            reply_markup=build_advanced_colors_keyboard(colors),
        )
        return
    await state.update_data(advanced_color=match)
    await _prompt_advanced_min_length(message, state)


@dp.message(SearchWarehousePlasticStates.waiting_for_min_length)
async def process_advanced_search_min_length(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == SKIP_TEXT:
        await state.update_data(advanced_min_length=None)
        await _prompt_advanced_min_width(message, state)
        return
    value = parse_dimension_filter_value(text)
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –í–≤–µ–¥–∏—Ç–µ –¥–ª–∏–Ω—É —á–∏—Å–ª–æ–º –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
            reply_markup=SKIP_OR_CANCEL_KB,
        )
        return
    await state.update_data(advanced_min_length=value)
    await _prompt_advanced_min_width(message, state)


@dp.message(SearchWarehousePlasticStates.waiting_for_min_width)
async def process_advanced_search_min_width(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_search_plastic_flow(message, state)
        return
    if text == SKIP_TEXT:
        await state.update_data(advanced_min_width=None)
        await _perform_advanced_search(message, state)
        return
    value = parse_dimension_filter_value(text)
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –í–≤–µ–¥–∏—Ç–µ —à–∏—Ä–∏–Ω—É —á–∏—Å–ª–æ–º –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
            reply_markup=SKIP_OR_CANCEL_KB,
        )
        return
    await state.update_data(advanced_min_width=value)
    await _perform_advanced_search(message, state)


@dp.message(CommentWarehousePlasticStates.waiting_for_article)
async def process_comment_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_comment_plastic_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_plastic_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–∞—Å—Ç–∏–∫ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    previous_comment = record.get("comment")
    await state.update_data(
        plastic_id=record["id"],
        article=record.get("article"),
        previous_comment=previous_comment,
    )
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{format_plastic_record_for_message(record)}\n\n"
        f"–¢–µ–∫—É—â–∏–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {previous_comment or '‚Äî'}",
        reply_markup=CANCEL_KB,
    )
    await state.set_state(CommentWarehousePlasticStates.waiting_for_comment)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–æ–≤—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π. –ü—É—Å—Ç–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —É–¥–∞–ª–∏—Ç —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π.",
        reply_markup=CANCEL_KB,
    )


@dp.message(CommentWarehousePlasticStates.waiting_for_comment)
async def process_comment_update(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_comment_plastic_flow(message, state)
        return
    data = await state.get_data()
    record_id = data.get("plastic_id")
    article = data.get("article")
    previous_comment = data.get("previous_comment")
    if record_id is None or article is None:
        await _cancel_comment_plastic_flow(message, state)
        return
    new_comment_raw = (message.text or "").strip()
    new_comment: Optional[str]
    if new_comment_raw:
        new_comment = new_comment_raw
    else:
        new_comment = None
    updated = await update_warehouse_plastic_comment(record_id, new_comment)
    if not updated:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–±–Ω–æ–≤–∏—Ç—å –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        await state.clear()
        return
    await state.clear()
    await message.answer(
        "‚úÖ –ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π –æ–±–Ω–æ–≤–ª—ë–Ω.\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–°—Ç–∞—Ä—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {previous_comment or '‚Äî'}\n"
        f"–ù–æ–≤—ã–π –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {new_comment or '‚Äî'}",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(MoveWarehousePlasticStates.waiting_for_article)
async def process_move_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_move_plastic_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_plastic_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–∞—Å—Ç–∏–∫ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    locations = await fetch_plastic_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(
        plastic_id=record["id"],
        article=record.get("article"),
        previous_location=record.get("warehouse"),
    )
    previous_location = record.get("warehouse") or "‚Äî"
    formatted_record = format_plastic_record_for_message(record)
    await state.set_state(MoveWarehousePlasticStates.waiting_for_new_location)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{formatted_record}\n\n"
        f"–¢–µ–∫—É—â–µ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è: {previous_location}\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –Ω–æ–≤–æ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –∏–∑ —Å–ø–∏—Å–∫–∞ –Ω–∏–∂–µ.",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(MoveWarehousePlasticStates.waiting_for_new_location)
async def process_move_new_location(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_move_plastic_flow(message, state)
        return
    locations = await fetch_plastic_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    raw_location = (message.text or "").strip()
    match = next((item for item in locations if item.lower() == raw_location.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–Ω–æ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    data = await state.get_data()
    record_id = data.get("plastic_id")
    article = data.get("article")
    previous_location_raw = data.get("previous_location")
    previous_location_display = previous_location_raw or "‚Äî"
    if record_id is None or article is None:
        await _cancel_move_plastic_flow(message, state)
        return
    if previous_location_raw and previous_location_raw.lower() == match.lower():
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–∞—Å—Ç–∏–∫ —É–∂–µ –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ –≤—ã–±—Ä–∞–Ω–Ω–æ–º –º–µ—Å—Ç–µ. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–µ –º–µ—Å—Ç–æ.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    employee_id = message.from_user.id if message.from_user else None
    employee_name = message.from_user.full_name if message.from_user else None
    updated_record = await update_warehouse_plastic_location(
        record_id=record_id,
        new_location=match,
        employee_id=employee_id,
        employee_name=employee_name,
    )
    if updated_record is None:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–±–Ω–æ–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.clear()
    formatted = format_plastic_record_for_message(updated_record)
    await message.answer(
        "‚úÖ –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –æ–±–Ω–æ–≤–ª–µ–Ω–æ.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ü—Ä–µ–¥—ã–¥—É—â–µ–µ –º–µ—Å—Ç–æ: {previous_location_display}\n"
        f"–ù–æ–≤–æ–µ –º–µ—Å—Ç–æ: {match}\n\n"
        f"–ê–∫—Ç—É–∞–ª—å–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ:\n{formatted}",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(WriteOffWarehousePlasticStates.waiting_for_article)
async def process_write_off_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_write_off_plastic_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    record = await fetch_warehouse_plastic_by_article(article)
    if record is None:
        await message.answer(
            "‚ÑπÔ∏è –ü–ª–∞—Å—Ç–∏–∫ —Å —Ç–∞–∫–∏–º –∞—Ä—Ç–∏–∫—É–ª–æ–º –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –¥—Ä—É–≥–æ–π –∞—Ä—Ç–∏–∫—É–ª.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(plastic_id=record["id"], article=record.get("article"))
    formatted = format_plastic_record_for_message(record)
    await state.set_state(WriteOffWarehousePlasticStates.waiting_for_project)
    await message.answer(
        "–ù–∞–π–¥–µ–Ω–∞ –∑–∞–ø–∏—Å—å:\n\n"
        f"{formatted}\n\n"
        "–£–∫–∞–∂–∏—Ç–µ –ø—Ä–æ–µ–∫—Ç, –Ω–∞ –∫–æ—Ç–æ—Ä—ã–π –≤—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è —Å–ø–∏—Å–∞–Ω–∏–µ.",
        reply_markup=CANCEL_KB,
    )


@dp.message(WriteOffWarehousePlasticStates.waiting_for_project)
async def process_write_off_project(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_write_off_plastic_flow(message, state)
        return
    project = (message.text or "").strip()
    if not project:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–µ–∫—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –ø—Ä–æ–µ–∫—Ç.",
            reply_markup=CANCEL_KB,
        )
        return
    data = await state.get_data()
    record_id = data.get("plastic_id")
    article = data.get("article")
    if record_id is None or article is None:
        await _cancel_write_off_plastic_flow(message, state)
        return
    written_off_by_id = message.from_user.id if message.from_user else None
    written_off_by_name = message.from_user.full_name if message.from_user else None
    try:
        result = await write_off_warehouse_plastic(
            record_id=record_id,
            project=project,
            written_off_by_id=written_off_by_id,
            written_off_by_name=written_off_by_name,
        )
    except Exception:
        logging.exception("Failed to write off plastic record")
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–ø–∏—Å–∞—Ç—å –ø–ª–∞—Å—Ç–∏–∫. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    if result is None:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ù–µ —É–¥–∞–ª–æ—Å—å –Ω–∞–π—Ç–∏ –∑–∞–ø–∏—Å—å –¥–ª—è —Å–ø–∏—Å–∞–Ω–∏—è. –í–æ–∑–º–æ–∂–Ω–æ, –æ–Ω–∞ —É–∂–µ –±—ã–ª–∞ –∏–∑–º–µ–Ω–µ–Ω–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.clear()
    formatted = format_written_off_plastic_record(result)
    await message.answer(
        "‚úÖ –ü–ª–∞—Å—Ç–∏–∫ —Å–ø–∏—Å–∞–Ω —Å–æ —Å–∫–ª–∞–¥–∞.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ü—Ä–æ–µ–∫—Ç: {project}\n\n"
        f"–î–∞–Ω–Ω—ã–µ —Å–ø–∏—Å–∞–Ω–Ω–æ–π –∑–∞–ø–∏—Å–∏:\n{formatted}",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å")
async def handle_add_warehouse_plastic(message: Message, state: FSMContext) -> None:
    await state.clear()
    await state.set_state(AddWarehousePlasticStates.waiting_for_article)
    suggested_article: Optional[str] = None
    last_article = await fetch_max_plastic_article()
    if last_article is not None:
        suggested_article = str(last_article + 1)
    await state.update_data(article_suggestion=suggested_article)
    prompt_lines = ["–í–≤–µ–¥–∏—Ç–µ –Ω–æ–º–µ—Ä –∞—Ä—Ç–∏–∫—É–ª–∞ (—Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã)."]
    if last_article is not None and suggested_article is not None:
        prompt_lines.append("")
        prompt_lines.append(
            "–ü–æ—Å–ª–µ–¥–Ω–∏–π –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–π –∞—Ä—Ç–∏–∫—É–ª: "
            f"{last_article}. –ù–∞–∂–º–∏—Ç–µ –∫–Ω–æ–ø–∫—É –Ω–∏–∂–µ, —á—Ç–æ–±—ã –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Å–ª–µ–¥—É—é—â–∏–π –Ω–æ–º–µ—Ä."
        )
    await message.answer(
        "\n".join(prompt_lines),
        reply_markup=build_article_input_keyboard(suggested_article),
    )


@dp.message(AddWarehousePlasticStates.waiting_for_article)
async def process_plastic_article(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    article = (message.text or "").strip()
    if not article.isdigit():
        data = await state.get_data()
        suggestion = data.get("article_suggestion")
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –¥–æ–ª–∂–µ–Ω —Å–æ–¥–µ—Ä–∂–∞—Ç—å —Ç–æ–ª—å–∫–æ —Ü–∏—Ñ—Ä—ã. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_article_input_keyboard(suggestion),
        )
        return
    materials = await fetch_plastic_material_types()
    if not materials:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(article=article, article_suggestion=None)
    await state.set_state(AddWarehousePlasticStates.waiting_for_material)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –º–∞—Ç–µ—Ä–∏–∞–ª–∞:",
        reply_markup=build_materials_keyboard(materials),
    )


@dp.message(AddWarehousePlasticStates.waiting_for_material)
async def process_plastic_material(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    materials = await fetch_plastic_material_types()
    raw = (message.text or "").strip()
    match = next((item for item in materials if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¢–∞–∫–æ–π –º–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(materials),
        )
        return
    thicknesses = await fetch_material_thicknesses(match)
    if not thicknesses:
        await state.clear()
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ç–æ–ª—â–∏–Ω—ã. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(material=match)
    await state.set_state(AddWarehousePlasticStates.waiting_for_thickness)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É –∏–∑ —Å–ø–∏—Å–∫–∞:",
        reply_markup=build_thickness_keyboard(thicknesses),
    )


@dp.message(AddWarehousePlasticStates.waiting_for_thickness)
async def process_plastic_thickness(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    data = await state.get_data()
    material = data.get("material")
    if not material:
        await _cancel_add_plastic_flow(message, state)
        return
    thicknesses = await fetch_material_thicknesses(material)
    value = parse_thickness_input(message.text or "")
    if value is None or all(item != value for item in thicknesses):
        await message.answer(
            "‚ö†Ô∏è –í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É, –∏—Å–ø–æ–ª—å–∑—É—è –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ.",
            reply_markup=build_thickness_keyboard(thicknesses),
        )
        return
    colors = await fetch_material_colors(material)
    if not colors:
        await state.clear()
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ü–≤–µ—Ç–∞. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(thickness=value)
    await state.set_state(AddWarehousePlasticStates.waiting_for_color)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç:",
        reply_markup=build_colors_keyboard(colors),
    )


@dp.message(AddWarehousePlasticStates.waiting_for_color)
async def process_plastic_color(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    data = await state.get_data()
    material = data.get("material")
    if not material:
        await _cancel_add_plastic_flow(message, state)
        return
    colors = await fetch_material_colors(material)
    raw = (message.text or "").strip()
    match = next((item for item in colors if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¶–≤–µ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_colors_keyboard(colors),
        )
        return
    await state.update_data(color=match)
    await state.set_state(AddWarehousePlasticStates.waiting_for_length)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –¥–ª–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehousePlasticStates.waiting_for_length)
async def process_plastic_length(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –î–ª–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(length=value)
    await state.set_state(AddWarehousePlasticStates.waiting_for_width)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ —à–∏—Ä–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehousePlasticStates.waiting_for_width)
async def process_plastic_width(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –®–∏—Ä–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    locations = await fetch_plastic_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(width=value)
    await state.set_state(AddWarehousePlasticStates.waiting_for_storage)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(AddWarehousePlasticStates.waiting_for_storage)
async def process_plastic_storage(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    locations = await fetch_plastic_storage_locations()
    raw = (message.text or "").strip()
    match = next((item for item in locations if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–Ω–æ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    await state.update_data(storage=match)
    await state.set_state(AddWarehousePlasticStates.waiting_for_comment)
    await message.answer(
        "–î–æ–±–∞–≤—å—Ç–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π (–Ω–µ–æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ) –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddWarehousePlasticStates.waiting_for_comment)
async def process_plastic_comment(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_plastic_flow(message, state)
        return
    comment: Optional[str]
    if text == SKIP_TEXT:
        comment = None
    else:
        comment = text or None
    data = await state.get_data()
    article = data.get("article")
    material = data.get("material")
    thickness: Optional[Decimal] = data.get("thickness")
    color = data.get("color")
    length = data.get("length")
    width = data.get("width")
    storage = data.get("storage")
    if not all([article, material, thickness, color, length, width, storage]):
        await _cancel_add_plastic_flow(message, state)
        return
    employee_id = message.from_user.id if message.from_user else None
    employee_name = message.from_user.full_name if message.from_user else None
    record = await insert_warehouse_plastic_record(
        article=article,
        material=material,
        thickness=thickness,
        color=color,
        length_mm=Decimal(length),
        width_mm=Decimal(width),
        warehouse=storage,
        comment=comment,
        employee_id=employee_id,
        employee_name=employee_name,
    )
    await state.clear()
    summary_comment = (record.get("comment") if record else comment) or "‚Äî"
    if record and record.get("employee_name"):
        summary_employee = record.get("employee_name") or "‚Äî"
    else:
        summary_employee = employee_name or "‚Äî"
    arrival_at = record.get("arrival_at") if record else None
    if arrival_at:
        try:
            arrival_local = arrival_at.astimezone(WARSAW_TZ)
        except Exception:
            arrival_local = arrival_at
        arrival_formatted = arrival_local.strftime("%Y-%m-%d %H:%M")
    else:
        arrival_formatted = datetime.now(WARSAW_TZ).strftime("%Y-%m-%d %H:%M")
    await message.answer(
        "‚úÖ –ü–ª–∞—Å—Ç–∏–∫ –¥–æ–±–∞–≤–ª–µ–Ω –Ω–∞ —Å–∫–ª–∞–¥.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ú–∞—Ç–µ—Ä–∏–∞–ª: {material}\n"
        f"–¢–æ–ª—â–∏–Ω–∞: {format_thickness_value(thickness)}\n"
        f"–¶–≤–µ—Ç: {color}\n"
        f"–î–ª–∏–Ω–∞: {length} –º–º\n"
        f"–®–∏—Ä–∏–Ω–∞: {width} –º–º\n"
        f"–ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è: {storage}\n"
        f"–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {summary_comment}\n"
        f"–î–æ–±–∞–≤–∏–ª: {summary_employee}\n"
        f"–î–æ–±–∞–≤–ª–µ–Ω–æ: {arrival_formatted}",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(F.text == "++–¥–æ–±–∞–≤–∏—Ç—å –ø–∞—á–∫—É")
async def handle_add_warehouse_plastic_batch(
    message: Message, state: FSMContext
) -> None:
    await state.clear()
    last_article = await fetch_max_plastic_article()
    await state.update_data(batch_last_article=last_article)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_quantity)
    prompt_lines = ["–°–∫–æ–ª—å–∫–æ –ª–∏—Å—Ç–æ–≤ –ø–ª–∞—Å—Ç–∏–∫–∞ –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å? –í–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ."]
    if last_article is None:
        prompt_lines.append("")
        prompt_lines.append("–°–µ–π—á–∞—Å –Ω–µ—Ç –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã—Ö –∞—Ä—Ç–∏–∫—É–ª–æ–≤. –ù—É–º–µ—Ä–∞—Ü–∏—è –Ω–∞—á–Ω—ë—Ç—Å—è —Å 1.")
    else:
        prompt_lines.append("")
        prompt_lines.append(
            "–ü–æ—Å–ª–µ–¥–Ω–∏–π –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–π –∞—Ä—Ç–∏–∫—É–ª: "
            f"{last_article}. –ù–æ–≤—ã–µ –ª–∏—Å—Ç—ã –ø–æ–ª—É—á–∞—Ç –Ω–æ–º–µ—Ä–∞ –Ω–∞—á–∏–Ω–∞—è —Å {last_article + 1}."
        )
    await message.answer("\n".join(prompt_lines), reply_markup=CANCEL_KB)


@dp.message(AddWarehousePlasticBatchStates.waiting_for_quantity)
async def process_plastic_batch_quantity(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    quantity = parse_positive_integer(message.text or "")
    if quantity is None:
        await message.answer(
            "‚ö†Ô∏è –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–æ–ª–∂–Ω–æ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    materials = await fetch_plastic_material_types()
    if not materials:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(batch_quantity=quantity)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_material)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø –º–∞—Ç–µ—Ä–∏–∞–ª–∞:",
        reply_markup=build_materials_keyboard(materials),
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_material)
async def process_plastic_batch_material(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    materials = await fetch_plastic_material_types()
    raw = (message.text or "").strip()
    match = next((item for item in materials if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¢–∞–∫–æ–π –º–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(materials),
        )
        return
    thicknesses = await fetch_material_thicknesses(match)
    if not thicknesses:
        await state.clear()
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ç–æ–ª—â–∏–Ω—ã. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(batch_material=match)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_thickness)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É –∏–∑ —Å–ø–∏—Å–∫–∞:",
        reply_markup=build_thickness_keyboard(thicknesses),
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_thickness)
async def process_plastic_batch_thickness(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    data = await state.get_data()
    material = data.get("batch_material")
    if not material:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    thicknesses = await fetch_material_thicknesses(material)
    value = parse_thickness_input(message.text or "")
    if value is None or all(item != value for item in thicknesses):
        await message.answer(
            "‚ö†Ô∏è –í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É, –∏—Å–ø–æ–ª—å–∑—É—è –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ.",
            reply_markup=build_thickness_keyboard(thicknesses),
        )
        return
    colors = await fetch_material_colors(material)
    if not colors:
        await state.clear()
        await message.answer(
            "–î–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ –Ω–µ —É–∫–∞–∑–∞–Ω—ã —Ü–≤–µ—Ç–∞. –î–æ–±–∞–≤—å—Ç–µ –∏—Ö –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(batch_thickness=value)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_color)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç:",
        reply_markup=build_colors_keyboard(colors),
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_color)
async def process_plastic_batch_color(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    data = await state.get_data()
    material = data.get("batch_material")
    if not material:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    colors = await fetch_material_colors(material)
    raw = (message.text or "").strip()
    match = next((item for item in colors if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¶–≤–µ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_colors_keyboard(colors),
        )
        return
    await state.update_data(batch_color=match)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_length)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ –¥–ª–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_length)
async def process_plastic_batch_length(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –î–ª–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    await state.update_data(batch_length=value)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_width)
    await message.answer(
        "–£–∫–∞–∂–∏—Ç–µ —à–∏—Ä–∏–Ω—É –ª–∏—Å—Ç–∞ –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ).",
        reply_markup=CANCEL_KB,
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_width)
async def process_plastic_batch_width(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –®–∏—Ä–∏–Ω–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω—ã–º —á–∏—Å–ª–æ–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    locations = await fetch_plastic_storage_locations()
    if not locations:
        await state.clear()
        await message.answer(
            "–°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–∫–ª–∞–¥–∞.",
            reply_markup=WAREHOUSE_PLASTICS_KB,
        )
        return
    await state.update_data(batch_width=value)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_storage)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_storage)
async def process_plastic_batch_storage(message: Message, state: FSMContext) -> None:
    if (message.text or "").strip() == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    locations = await fetch_plastic_storage_locations()
    raw = (message.text or "").strip()
    match = next((item for item in locations if item.lower() == raw.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¢–∞–∫–æ–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–Ω–æ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_storage_locations_keyboard(locations),
        )
        return
    await state.update_data(batch_storage=match)
    await state.set_state(AddWarehousePlasticBatchStates.waiting_for_comment)
    await message.answer(
        "–î–æ–±–∞–≤—å—Ç–µ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π (–Ω–µ–æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ) –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ ¬´–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å¬ª.",
        reply_markup=SKIP_OR_CANCEL_KB,
    )


@dp.message(AddWarehousePlasticBatchStates.waiting_for_comment)
async def process_plastic_batch_comment(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    if text == CANCEL_TEXT:
        await _cancel_add_plastic_batch_flow(message, state)
        return
    if text == SKIP_TEXT:
        comment: Optional[str] = None
    else:
        comment = text or None
    data = await state.get_data()
    quantity = data.get("batch_quantity")
    material = data.get("batch_material")
    thickness: Optional[Decimal] = data.get("batch_thickness")
    color = data.get("batch_color")
    length = data.get("batch_length")
    width = data.get("batch_width")
    storage = data.get("batch_storage")
    last_article = data.get("batch_last_article")
    if not all([quantity, material, thickness, color, length, width, storage]):
        await _cancel_add_plastic_batch_flow(message, state)
        return
    if not isinstance(quantity, int):
        try:
            quantity = int(quantity)
        except (TypeError, ValueError):
            await _cancel_add_plastic_batch_flow(message, state)
            return
    start_article = 1 if last_article is None else int(last_article) + 1
    articles = [str(start_article + idx) for idx in range(quantity)]
    employee_id = message.from_user.id if message.from_user else None
    employee_name = message.from_user.full_name if message.from_user else None
    records: list[Dict[str, Any]] = []
    for article in articles:
        record = await insert_warehouse_plastic_record(
            article=article,
            material=material,
            thickness=thickness,
            color=color,
            length_mm=Decimal(length),
            width_mm=Decimal(width),
            warehouse=storage,
            comment=comment,
            employee_id=employee_id,
            employee_name=employee_name,
        )
        if not record:
            await state.clear()
            await message.answer(
                "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –¥–æ–±–∞–≤–∏—Ç—å –ø–ª–∞—Å—Ç–∏–∫. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
                reply_markup=WAREHOUSE_PLASTICS_KB,
            )
            return
        records.append(record)
    await state.clear()
    summary_comment = (records[0].get("comment") if records else comment) or "‚Äî"
    if records and records[0].get("employee_name"):
        summary_employee = records[0].get("employee_name") or "‚Äî"
    else:
        summary_employee = employee_name or "‚Äî"
    arrival_at = records[0].get("arrival_at") if records else None
    if arrival_at:
        try:
            arrival_local = arrival_at.astimezone(WARSAW_TZ)
        except Exception:
            arrival_local = arrival_at
        arrival_formatted = arrival_local.strftime("%Y-%m-%d %H:%M")
    else:
        arrival_formatted = datetime.now(WARSAW_TZ).strftime("%Y-%m-%d %H:%M")
    articles_text = ", ".join(articles)
    await message.answer(
        "‚úÖ –ü–∞—á–∫–∞ –ø–ª–∞—Å—Ç–∏–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∞ –Ω–∞ —Å–∫–ª–∞–¥.\n\n"
        f"–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ: {quantity}\n"
        f"–ê—Ä—Ç–∏–∫—É–ª—ã: {articles_text}\n"
        f"–ú–∞—Ç–µ—Ä–∏–∞–ª: {material}\n"
        f"–¢–æ–ª—â–∏–Ω–∞: {format_thickness_value(thickness)}\n"
        f"–¶–≤–µ—Ç: {color}\n"
        f"–î–ª–∏–Ω–∞: {length} –º–º\n"
        f"–®–∏—Ä–∏–Ω–∞: {width} –º–º\n"
        f"–ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è: {storage}\n"
        f"–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π: {summary_comment}\n"
        f"–î–æ–±–∞–≤–∏–ª: {summary_employee}\n"
        f"–î–æ–±–∞–≤–ª–µ–Ω–æ: {arrival_formatted}",
        reply_markup=WAREHOUSE_PLASTICS_KB,
    )


@dp.message(F.text == "üéûÔ∏è –ü–ª–µ–Ω–∫–∏ ‚öôÔ∏è")
async def handle_warehouse_settings_films(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "üè≠ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å")
async def handle_film_manufacturers_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_film_manufacturers_menu(message)


@dp.message(F.text == "üè¨ –°–∫–ª–∞–¥")
async def handle_film_storage_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_film_storage_overview(message)


@dp.message(F.text == "üé¨ –°–µ—Ä–∏—è")
async def handle_film_series_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    manufacturers = await fetch_film_manufacturers_with_series()
    if manufacturers:
        lines = []
        for manufacturer in manufacturers:
            name = manufacturer["name"]
            formatted_series = format_series_list(manufacturer.get("series") or [])
            lines.append(
                "\n".join(
                    [
                        f"‚Ä¢ {name}",
                        f"   –°–µ—Ä–∏–∏: {formatted_series}",
                    ]
                )
            )
        formatted = "\n".join(lines)
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–µ–Ω–∫–∏ ‚Üí –°–µ—Ä–∏—è.\n\n"
            "–î–æ—Å—Ç—É–ø–Ω—ã–µ —Å–µ—Ä–∏–∏ –ø–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è–º:\n"
            f"{formatted}\n\n"
            "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫–∏ –Ω–∏–∂–µ, —á—Ç–æ–±—ã –¥–æ–±–∞–≤–∏—Ç—å –∏–ª–∏ —É–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é."
        )
    else:
        text = (
            "‚öôÔ∏è –ù–∞—Å—Ç—Ä–æ–π–∫–∏ —Å–∫–ª–∞–¥–∞ ‚Üí –ü–ª–µ–Ω–∫–∏ ‚Üí –°–µ—Ä–∏—è.\n\n"
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π, —á—Ç–æ–±—ã —Å–æ–∑–¥–∞–≤–∞—Ç—å —Å–µ—Ä–∏–∏."
        )
    await message.answer(text, reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB)


@dp.message(F.text == "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–µ–Ω–∫–∞–º")
async def handle_back_to_film_settings(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è")
async def handle_add_film_manufacturer_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManageFilmManufacturerStates.waiting_for_new_manufacturer_name
    )
    manufacturers = await fetch_film_manufacturers()
    existing_text = format_materials_list(manufacturers)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageFilmManufacturerStates.waiting_for_new_manufacturer_name)
async def process_new_film_manufacturer(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_film_manufacturer(name):
        await message.answer(f"‚úÖ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è")
async def handle_remove_film_manufacturer_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_film_manufacturers()
    if not manufacturers:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_FILM_MANUFACTURERS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageFilmManufacturerStates.waiting_for_manufacturer_name_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(ManageFilmManufacturerStates.waiting_for_manufacturer_name_to_delete)
async def process_remove_film_manufacturer(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_film_manufacturer(name):
        await message.answer(f"üóë –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–ª–µ–Ω–∫–∏")
async def handle_add_film_storage_location_button(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManageFilmStorageStates.waiting_for_new_storage_location_name
    )
    locations = await fetch_film_storage_locations()
    existing_text = format_storage_locations_list(locations)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–ª—è –ø–ª–µ–Ω–∫–∏.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageFilmStorageStates.waiting_for_new_storage_location_name)
async def process_new_film_storage_location(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_film_storage_location(name):
        await message.answer(f"‚úÖ –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_film_storage_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–ª–µ–Ω–∫–∏")
async def handle_remove_film_storage_location_button(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    locations = await fetch_film_storage_locations()
    if not locations:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_FILM_STORAGE_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageFilmStorageStates.waiting_for_storage_location_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è, –∫–æ—Ç–æ—Ä–æ–µ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(ManageFilmStorageStates.waiting_for_storage_location_to_delete)
async def process_remove_film_storage_location(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_film_storage_location(name):
        await message.answer(f"üóë –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–¥–∞–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_film_storage_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é")
async def handle_add_film_series_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_film_manufacturers()
    if not manufacturers:
        await message.answer(
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π, —á—Ç–æ–±—ã —É–∫–∞–∑–∞—Ç—å –∏—Ö —Å–µ—Ä–∏–∏.",
            reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageFilmSeriesStates.waiting_for_manufacturer_for_new_series
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –¥–ª—è –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(ManageFilmSeriesStates.waiting_for_manufacturer_for_new_series)
async def process_choose_manufacturer_for_new_series(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    manufacturer_name = (message.text or "").strip()
    manufacturer = await get_film_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        manufacturers = await fetch_film_manufacturers()
        if not manufacturers:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π, —á—Ç–æ–±—ã –ø—Ä–æ–¥–æ–ª–∂–∏—Ç—å.",
                reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturers),
        )
        return
    await state.update_data(selected_manufacturer=manufacturer["name"])
    await state.set_state(ManageFilmSeriesStates.waiting_for_new_series_name)
    existing_series = await fetch_film_series_by_manufacturer(manufacturer["name"])
    formatted_series = format_series_list(existing_series)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –Ω–æ–≤–æ–π —Å–µ—Ä–∏–∏.\n\n"
        f"–¢–µ–∫—É—â–∏–µ —Å–µ—Ä–∏–∏ —É ¬´{manufacturer['name']}¬ª: {formatted_series}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageFilmSeriesStates.waiting_for_new_series_name)
async def process_new_series_name(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    series_name = (message.text or "").strip()
    if not series_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Å–µ—Ä–∏–∏ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    data = await state.get_data()
    manufacturer_name = data.get("selected_manufacturer")
    if not manufacturer_name:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
        )
        await send_film_settings_overview(message)
        return
    status = await insert_film_series(manufacturer_name, series_name)
    if status == "manufacturer_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –û–±–Ω–æ–≤–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –∏ –ø–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
    elif status == "already_exists":
        await message.answer(
            f"‚ÑπÔ∏è –°–µ—Ä–∏—è ¬´{series_name}¬ª —É–∂–µ —É–∫–∞–∑–∞–Ω–∞ –¥–ª—è ¬´{manufacturer_name}¬ª."
        )
    elif status == "inserted":
        await message.answer(
            f"‚úÖ –°–µ—Ä–∏—è ¬´{series_name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–∞ –¥–ª—è ¬´{manufacturer_name}¬ª."
        )
    else:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –¥–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ."
        )
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é")
async def handle_remove_film_series_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_film_manufacturers_with_series()
    manufacturers_with_series = [
        item["name"] for item in manufacturers if item.get("series")
    ]
    if not manufacturers_with_series:
        await message.answer(
            "–î–ª—è —É–¥–∞–ª–µ–Ω–∏—è –ø–æ–∫–∞ –Ω–µ—Ç –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã—Ö —Å–µ—Ä–∏–π.",
            reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageFilmSeriesStates.waiting_for_manufacturer_for_series_deletion
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, —É –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é:",
        reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
    )


@dp.message(ManageFilmSeriesStates.waiting_for_manufacturer_for_series_deletion)
async def process_choose_manufacturer_for_series_deletion(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    manufacturer_name = (message.text or "").strip()
    manufacturer = await get_film_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        manufacturers = await fetch_film_manufacturers_with_series()
        manufacturers_with_series = [
            item["name"] for item in manufacturers if item.get("series")
        ]
        if not manufacturers_with_series:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ —Å–µ—Ä–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏–∏, —á—Ç–æ–±—ã –∏—Ö —É–¥–∞–ª–∏—Ç—å.",
                reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
        )
        return
    series = await fetch_film_series_by_manufacturer(manufacturer["name"])
    if not series:
        manufacturers = await fetch_film_manufacturers_with_series()
        manufacturers_with_series = [
            item["name"] for item in manufacturers if item.get("series")
        ]
        if not manufacturers_with_series:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ —Å–µ—Ä–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏–∏, —á—Ç–æ–±—ã –∏—Ö —É–¥–∞–ª–∏—Ç—å.",
                reply_markup=WAREHOUSE_SETTINGS_FILM_SERIES_KB,
            )
            return
        await message.answer(
            "‚ÑπÔ∏è –£ –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –Ω–µ—Ç —Å–µ—Ä–∏–π. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
            reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
        )
        return
    await state.update_data(selected_manufacturer=manufacturer["name"])
    await state.set_state(ManageFilmSeriesStates.waiting_for_series_name_to_delete)
    await message.answer(
        f"–í—ã–±–µ—Ä–∏—Ç–µ —Å–µ—Ä–∏—é –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è —É ¬´{manufacturer['name']}¬ª:",
        reply_markup=build_series_keyboard(series),
    )


@dp.message(ManageFilmSeriesStates.waiting_for_series_name_to_delete)
async def process_remove_film_series(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    series_name = (message.text or "").strip()
    if not series_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Å–µ—Ä–∏–∏ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    data = await state.get_data()
    manufacturer_name = data.get("selected_manufacturer")
    if not manufacturer_name:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
        await send_film_settings_overview(message)
        return
    status = await delete_film_series(manufacturer_name, series_name)
    if status == "manufacturer_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –û–±–Ω–æ–≤–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –∏ –ø–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
    elif status == "deleted":
        await message.answer(
            f"üóë –°–µ—Ä–∏—è ¬´{series_name}¬ª —É–¥–∞–ª–µ–Ω–∞ —É ¬´{manufacturer_name}¬ª."
        )
    else:
        await message.answer(
            f"‚ÑπÔ∏è –°–µ—Ä–∏—è ¬´{series_name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ —É ¬´{manufacturer_name}¬ª."
        )
    await state.clear()
    await send_film_settings_overview(message)


@dp.message(F.text == "üß± –ü–ª–∞—Å—Ç–∏–∫")
async def handle_warehouse_settings_plastic(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await send_plastic_settings_overview(message)


@dp.message(F.text == WAREHOUSE_SETTINGS_ELECTRICS_TEXT)
async def handle_warehouse_settings_electrics(message: Message) -> None:
    if not await ensure_admin_access(message):
        return
    await send_electrics_settings_overview(message)


@dp.message(F.text == WAREHOUSE_SETTINGS_ELECTRICS_LED_STRIPS_TEXT)
async def handle_warehouse_settings_led_strips(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_strips_settings_overview(message)


@dp.message(F.text == WAREHOUSE_SETTINGS_ELECTRICS_LED_MODULES_TEXT)
async def handle_warehouse_settings_led_modules(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == LED_MODULES_MANUFACTURERS_MENU_TEXT)
async def handle_led_module_manufacturers_menu(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_manufacturers_menu(message)


@dp.message(F.text == LED_MODULES_COLORS_MENU_TEXT)
async def handle_led_module_colors_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_colors_menu(message)


@dp.message(F.text == LED_MODULES_POWER_MENU_TEXT)
async def handle_led_module_power_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_power_menu(message)


@dp.message(F.text == LED_MODULES_VOLTAGE_MENU_TEXT)
async def handle_led_module_voltage_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_voltage_menu(message)


@dp.message(F.text == LED_MODULES_LENS_MENU_TEXT)
async def handle_led_module_lens_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_lens_menu(message)


@dp.message(F.text == LED_MODULES_SERIES_MENU_TEXT)
async def handle_led_module_series_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_series_menu(message)


@dp.message(F.text == LED_MODULES_STORAGE_MENU_TEXT)
async def handle_led_module_storage_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_storage_overview(message)


@dp.message(F.text == LED_MODULES_BASE_MENU_TEXT)
async def handle_led_module_base_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_module_base_menu(message)


@dp.message(F.text == LED_MODULES_GENERATE_TEXT)
async def handle_generate_led_module(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    manufacturers_with_series = [
        item
        for item in await fetch_led_module_manufacturers_with_series()
        if item.get("series")
    ]
    colors = await fetch_led_module_colors()
    lens_counts = await fetch_led_module_lens_counts()
    power_options = await fetch_led_module_power_options()
    voltage_options = await fetch_led_module_voltage_options()
    missing: list[str] = []
    if not manufacturers_with_series:
        missing.append("‚Ä¢ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–∏ –∏ —Å–µ—Ä–∏–∏")
    if not colors:
        missing.append("‚Ä¢ —Ü–≤–µ—Ç–∞ –º–æ–¥—É–ª–µ–π")
    if not lens_counts:
        missing.append("‚Ä¢ –∑–Ω–∞—á–µ–Ω–∏—è –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ª–∏–Ω–∑")
    if not power_options:
        missing.append("‚Ä¢ –∑–Ω–∞—á–µ–Ω–∏—è –º–æ—â–Ω–æ—Å—Ç–∏")
    if not voltage_options:
        missing.append("‚Ä¢ –∑–Ω–∞—á–µ–Ω–∏—è –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è")
    if missing:
        details = "\n".join(missing)
        await message.answer(
            "‚ö†Ô∏è –ù–µ–≤–æ–∑–º–æ–∂–Ω–æ –Ω–∞—á–∞—Ç—å –≥–µ–Ω–µ—Ä–∞—Ü–∏—é Led –º–æ–¥—É–ª—è.\n\n"
            "–ó–∞–ø–æ–ª–Ω–∏—Ç–µ —Å–ª–µ–¥—É—é—â–∏–µ —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∏ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö:\n"
            f"{details}",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    await state.set_state(GenerateLedModuleStates.waiting_for_article)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∞—Ä—Ç–∏–∫—É–ª –¥–ª—è –Ω–æ–≤–æ–≥–æ Led –º–æ–¥—É–ª—è.",
        reply_markup=build_article_input_keyboard(),
    )


@dp.message(GenerateLedModuleStates.waiting_for_article)
async def process_generate_led_module_article(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    article = (message.text or "").strip()
    if not article:
        await message.answer(
            "‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_article_input_keyboard(),
        )
        return
    existing = await get_generated_led_module_by_article(article)
    if existing:
        await message.answer(
            f"‚ö†Ô∏è –ê—Ä—Ç–∏–∫—É–ª ¬´{article}¬ª —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –£–∫–∞–∂–∏—Ç–µ –¥—Ä—É–≥–æ–π.",
            reply_markup=build_article_input_keyboard(),
        )
        return
    manufacturers_with_series = [
        item
        for item in await fetch_led_module_manufacturers_with_series()
        if item.get("series")
    ]
    if not manufacturers_with_series:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π —Å —Å–µ—Ä–∏—è–º–∏ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –¥–∞–Ω–Ω—ã–µ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö Led –º–æ–¥—É–ª–µ–π.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    manufacturer_names = [item["name"] for item in manufacturers_with_series]
    await state.update_data(generated_led_module_article=article)
    await state.set_state(GenerateLedModuleStates.waiting_for_manufacturer)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –º–æ–¥—É–ª—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã:\n"
        f"{format_materials_list(manufacturer_names)}",
        reply_markup=build_manufacturers_keyboard(manufacturer_names),
    )


@dp.message(GenerateLedModuleStates.waiting_for_manufacturer)
async def process_generate_led_module_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    manufacturers_with_series = [
        item
        for item in await fetch_led_module_manufacturers_with_series()
        if item.get("series")
    ]
    if not manufacturers_with_series:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π —Å —Å–µ—Ä–∏—è–º–∏ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –¥–∞–Ω–Ω—ã–µ –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö Led –º–æ–¥—É–ª–µ–π.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    manufacturer_names = [item["name"] for item in manufacturers_with_series]
    raw = (message.text or "").strip()
    match = next(
        (item for item in manufacturers_with_series if item["name"].lower() == raw.lower()),
        None,
    )
    if match is None:
        await message.answer(
            "‚ö†Ô∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturer_names),
        )
        return
    series_names = await fetch_led_module_series_by_manufacturer(match["name"])
    if not series_names:
        await message.answer(
            "‚ÑπÔ∏è –£ –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –ø–æ–∫–∞ –Ω–µ—Ç —Å–µ—Ä–∏–π. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –∏–ª–∏ –¥–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏—é –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö.",
            reply_markup=build_manufacturers_keyboard(manufacturer_names),
        )
        return
    await state.update_data(
        generated_led_module_manufacturer={"id": match["id"], "name": match["name"]}
    )
    await state.set_state(GenerateLedModuleStates.waiting_for_series)
    await message.answer(
        f"–í—ã–±–µ—Ä–∏—Ç–µ —Å–µ—Ä–∏—é –¥–ª—è –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è ¬´{match['name']}¬ª.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ —Å–µ—Ä–∏–∏:\n"
        f"{format_materials_list(series_names)}",
        reply_markup=build_series_keyboard(series_names),
    )


@dp.message(GenerateLedModuleStates.waiting_for_series)
async def process_generate_led_module_series(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    data = await state.get_data()
    manufacturer: Optional[dict[str, Any]] = data.get("generated_led_module_manufacturer")
    if not manufacturer:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è. –ù–∞—á–Ω–∏—Ç–µ –≥–µ–Ω–µ—Ä–∞—Ü–∏—é –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
        )
        return
    series_names = await fetch_led_module_series_by_manufacturer(manufacturer["name"])
    if not series_names:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –£ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –±–æ–ª—å—à–µ –Ω–µ—Ç —Å–µ—Ä–∏–π. –î–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏—é –∏ –Ω–∞—á–Ω–∏—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    raw = (message.text or "").strip()
    series_name = next((item for item in series_names if item.lower() == raw.lower()), None)
    if series_name is None:
        await message.answer(
            "‚ö†Ô∏è –°–µ—Ä–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –í—ã–±–µ—Ä–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_series_keyboard(series_names),
        )
        return
    series = await get_led_module_series_by_name(manufacturer["id"], series_name)
    if series is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—É—é —Å–µ—Ä–∏—é. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_series_keyboard(series_names),
        )
        return
    await state.update_data(generated_led_module_series=series)
    colors = await fetch_led_module_colors()
    if not colors:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ —Ü–≤–µ—Ç–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Ü–≤–µ—Ç–∞ –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    await state.set_state(GenerateLedModuleStates.waiting_for_color)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç Led –º–æ–¥—É–ª—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ —Ü–≤–µ—Ç–∞:\n"
        f"{format_materials_list(colors)}",
        reply_markup=build_colors_keyboard(colors),
    )


@dp.message(GenerateLedModuleStates.waiting_for_color)
async def process_generate_led_module_color(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    colors = await fetch_led_module_colors()
    if not colors:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ —Ü–≤–µ—Ç–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Ü–≤–µ—Ç–∞ –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    raw = (message.text or "").strip()
    color_name = next((item for item in colors if item.lower() == raw.lower()), None)
    if color_name is None:
        await message.answer(
            "‚ö†Ô∏è –¶–≤–µ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_colors_keyboard(colors),
        )
        return
    color = await get_led_module_color_by_name(color_name)
    if color is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Ü–≤–µ—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_colors_keyboard(colors),
        )
        return
    await state.update_data(generated_led_module_color=color)
    lens_counts = await fetch_led_module_lens_counts()
    if not lens_counts:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ª–∏–Ω–∑ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    await state.set_state(GenerateLedModuleStates.waiting_for_lens_count)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ –¥–ª—è Led –º–æ–¥—É–ª—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã:\n"
        f"{format_materials_list([str(value) for value in lens_counts])}",
        reply_markup=build_lens_counts_keyboard(lens_counts),
    )


@dp.message(GenerateLedModuleStates.waiting_for_lens_count)
async def process_generate_led_module_lens_count(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    lens_counts = await fetch_led_module_lens_counts()
    if not lens_counts:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ª–∏–Ω–∑ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    raw = (message.text or "").strip()
    parsed = parse_positive_integer(raw)
    if parsed is None or parsed not in lens_counts:
        await message.answer(
            "‚ö†Ô∏è –£–∫–∞–∂–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑, –¥–æ—Å—Ç—É–ø–Ω–æ–µ –≤ —Å–ø–∏—Å–∫–µ.",
            reply_markup=build_lens_counts_keyboard(lens_counts),
        )
        return
    lens = await get_led_module_lens_count_by_value(parsed)
    if lens is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_lens_counts_keyboard(lens_counts),
        )
        return
    await state.update_data(generated_led_module_lens_count=lens)
    power_options = await fetch_led_module_power_options()
    if not power_options:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–æ—â–Ω–æ—Å—Ç–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    await state.set_state(GenerateLedModuleStates.waiting_for_power)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–æ—â–Ω–æ—Å—Ç—å Led –º–æ–¥—É–ª—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã:\n"
        f"{format_materials_list(power_options)}",
        reply_markup=build_power_values_keyboard(power_options),
    )


@dp.message(GenerateLedModuleStates.waiting_for_power)
async def process_generate_led_module_power(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    power_options = await fetch_led_module_power_options()
    if not power_options:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –º–æ—â–Ω–æ—Å—Ç–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    raw = (message.text or "").strip()
    power_name = next((item for item in power_options if item.lower() == raw.lower()), None)
    if power_name is None:
        await message.answer(
            "‚ö†Ô∏è –ú–æ—â–Ω–æ—Å—Ç—å –Ω–µ –Ω–∞–π–¥–µ–Ω–∞. –í—ã–±–µ—Ä–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_power_values_keyboard(power_options),
        )
        return
    power = await get_led_module_power_option_by_name(power_name)
    if power is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—É—é –º–æ—â–Ω–æ—Å—Ç—å. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_power_values_keyboard(power_options),
        )
        return
    await state.update_data(generated_led_module_power=power)
    voltage_options = await fetch_led_module_voltage_options()
    if not voltage_options:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    await state.set_state(GenerateLedModuleStates.waiting_for_voltage)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ Led –º–æ–¥—É–ª—è.\n\n"
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã:\n"
        f"{format_materials_list(voltage_options)}",
        reply_markup=build_voltage_values_keyboard(voltage_options),
    )


@dp.message(GenerateLedModuleStates.waiting_for_voltage)
async def process_generate_led_module_voltage(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    voltage_options = await fetch_led_module_voltage_options()
    if not voltage_options:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –°–ø—Ä–∞–≤–æ—á–Ω–∏–∫ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –∏ –Ω–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_KB,
        )
        return
    raw = (message.text or "").strip()
    voltage_name = next((item for item in voltage_options if item.lower() == raw.lower()), None)
    if voltage_name is None:
        await message.answer(
            "‚ö†Ô∏è –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ –Ω–µ –Ω–∞–π–¥–µ–Ω–æ. –í—ã–±–µ—Ä–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_voltage_values_keyboard(voltage_options),
        )
        return
    voltage = await get_led_module_voltage_option_by_name(voltage_name)
    if voltage is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω–æ–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=build_voltage_values_keyboard(voltage_options),
        )
        return
    data = await state.get_data()
    article = data.get("generated_led_module_article")
    manufacturer = data.get("generated_led_module_manufacturer")
    series = data.get("generated_led_module_series")
    color = data.get("generated_led_module_color")
    lens = data.get("generated_led_module_lens_count")
    power = data.get("generated_led_module_power")
    if not all([article, manufacturer, series, color, lens, power]):
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è Led –º–æ–¥—É–ª—è. –ù–∞—á–Ω–∏—Ç–µ –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
        )
        return
    record = await insert_generated_led_module(
        article=article,
        manufacturer_id=manufacturer["id"],
        series_id=series["id"],
        color_id=color["id"],
        lens_count_id=lens["id"],
        power_option_id=power["id"],
        voltage_option_id=voltage["id"],
    )
    if record is None:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å Led –º–æ–¥—É–ª—å. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
        )
        return
    await state.clear()
    created_at = record.get("created_at")
    created_text = _format_datetime(created_at)
    await message.answer(
        "‚úÖ Led –º–æ–¥—É–ª—å –¥–æ–±–∞–≤–ª–µ–Ω –≤ –±–∞–∑—É.\n\n"
        f"–ê—Ä—Ç–∏–∫—É–ª: {article}\n"
        f"–ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å: {manufacturer['name']}\n"
        f"–°–µ—Ä–∏—è: {series['name']}\n"
        f"–¶–≤–µ—Ç: {color['name']}\n"
        f"–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑: {lens['value']}\n"
        f"–ú–æ—â–Ω–æ—Å—Ç—å: {power['name']}\n"
        f"–ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ: {voltage['name']}\n"
        f"–°–æ–∑–¥–∞–Ω–æ: {created_text}",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
    )


@dp.message(F.text == LED_MODULES_DELETE_TEXT)
async def handle_delete_led_module(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await message.answer(
        "üóëÔ∏è –£–¥–∞–ª–µ–Ω–∏–µ Led –º–æ–¥—É–ª—è –Ω–∞—Ö–æ–¥–∏—Ç—Å—è –≤ —Ä–∞–∑—Ä–∞–±–æ—Ç–∫–µ.",
        reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_BASE_KB,
    )


@dp.message(F.text == LED_MODULES_BACK_TEXT)
async def handle_back_to_led_module_settings(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == WAREHOUSE_SETTINGS_ELECTRICS_POWER_SUPPLIES_TEXT)
async def handle_warehouse_settings_power_supplies(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_power_supplies_settings_overview(message)


@dp.message(F.text == LED_STRIPS_ADD_MANUFACTURER_TEXT)
async def handle_add_led_strip_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManageLedStripManufacturerStates.waiting_for_new_manufacturer_name
    )
    manufacturers = await fetch_led_strip_manufacturers()
    existing_text = format_materials_list(manufacturers)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –ª–µ–Ω—Ç—ã.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedStripManufacturerStates.waiting_for_new_manufacturer_name)
async def process_new_led_strip_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_led_strip_manufacturer(name):
        await message.answer(f"‚úÖ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_strips_settings_overview(message)


@dp.message(F.text == LED_STRIPS_REMOVE_MANUFACTURER_TEXT)
async def handle_remove_led_strip_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_led_strip_manufacturers()
    if not manufacturers:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_STRIPS_MANUFACTURERS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageLedStripManufacturerStates.waiting_for_manufacturer_name_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(ManageLedStripManufacturerStates.waiting_for_manufacturer_name_to_delete)
async def process_remove_led_strip_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_led_strip_manufacturer(name):
        await message.answer(f"üóë –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_strips_settings_overview(message)


@dp.message(F.text == LED_MODULES_ADD_MANUFACTURER_TEXT)
async def handle_add_led_module_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManageLedModuleManufacturerStates.waiting_for_new_manufacturer_name
    )
    manufacturers = await fetch_led_module_manufacturers()
    existing_text = format_materials_list(manufacturers)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è Led –º–æ–¥—É–ª–µ–π.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleManufacturerStates.waiting_for_new_manufacturer_name)
async def process_new_led_module_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_led_module_manufacturer(name):
        await message.answer(f"‚úÖ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == LED_MODULES_REMOVE_MANUFACTURER_TEXT)
async def handle_remove_led_module_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_led_module_manufacturers()
    if not manufacturers:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_MANUFACTURERS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageLedModuleManufacturerStates.waiting_for_manufacturer_name_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(F.text == LED_MODULES_ADD_STORAGE_TEXT)
async def handle_add_led_module_storage_location(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManageLedModuleStorageStates.waiting_for_new_storage_location_name
    )
    locations = await fetch_led_module_storage_locations()
    existing_text = format_storage_locations_list(locations)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è –¥–ª—è Led –º–æ–¥—É–ª–µ–π.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleStorageStates.waiting_for_new_storage_location_name)
async def process_new_led_module_storage_location(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_led_module_storage_location(name):
        await message.answer(f"‚úÖ –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_storage_overview(message)


@dp.message(F.text == LED_MODULES_REMOVE_STORAGE_TEXT)
async def handle_remove_led_module_storage_location(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    locations = await fetch_led_module_storage_locations()
    if not locations:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_STORAGE_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageLedModuleStorageStates.waiting_for_storage_location_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è, –∫–æ—Ç–æ—Ä–æ–µ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(ManageLedModuleStorageStates.waiting_for_storage_location_to_delete)
async def process_remove_led_module_storage_location(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_led_module_storage_location(name):
        await message.answer(f"üóë –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–¥–∞–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_storage_overview(message)


@dp.message(ManageLedModuleManufacturerStates.waiting_for_manufacturer_name_to_delete)
async def process_remove_led_module_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_led_module_manufacturer(name):
        await message.answer(f"üóë –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == LED_MODULES_ADD_COLOR_TEXT)
async def handle_add_led_module_color(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManageLedModuleColorStates.waiting_for_new_color_name)
    colors = await fetch_led_module_colors()
    existing_text = format_materials_list(colors)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ —Ü–≤–µ—Ç Led –º–æ–¥—É–ª–µ–π.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleColorStates.waiting_for_new_color_name)
async def process_new_led_module_color(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    color_name = (message.text or "").strip()
    if not color_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Ü–≤–µ—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_led_module_color(color_name):
        await message.answer(f"‚úÖ –¶–≤–µ—Ç ¬´{color_name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –¶–≤–µ—Ç ¬´{color_name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_colors_menu(message)


@dp.message(F.text == LED_MODULES_ADD_POWER_TEXT)
async def handle_add_led_module_power_option(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManageLedModulePowerStates.waiting_for_new_power_value)
    existing = await fetch_led_module_power_options()
    existing_text = format_materials_list(existing)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –º–æ—â–Ω–æ—Å—Ç–∏ –¥–ª—è Led –º–æ–¥—É–ª–µ–π.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModulePowerStates.waiting_for_new_power_value)
async def process_new_led_module_power_option(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = (message.text or "").strip()
    if not value:
        await message.answer("‚ö†Ô∏è –ó–Ω–∞—á–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_led_module_power_option(value):
        await message.answer(f"‚úÖ –ú–æ—â–Ω–æ—Å—Ç—å ¬´{value}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–∞.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–æ—â–Ω–æ—Å—Ç—å ¬´{value}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_power_menu(message)


@dp.message(F.text == LED_MODULES_ADD_VOLTAGE_TEXT)
async def handle_add_led_module_voltage_option(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManageLedModuleVoltageStates.waiting_for_new_voltage_value)
    existing = await fetch_led_module_voltage_options()
    existing_text = format_materials_list(existing)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∑–Ω–∞—á–µ–Ω–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è –¥–ª—è Led –º–æ–¥—É–ª–µ–π.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleVoltageStates.waiting_for_new_voltage_value)
async def process_new_led_module_voltage_option(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = (message.text or "").strip()
    if not value:
        await message.answer(
            "‚ö†Ô∏è –ó–Ω–∞—á–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if await insert_led_module_voltage_option(value):
        await message.answer(f"‚úÖ –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ ¬´{value}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ ¬´{value}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_voltage_menu(message)


@dp.message(F.text == LED_MODULES_REMOVE_COLOR_TEXT)
async def handle_remove_led_module_color(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    colors = await fetch_led_module_colors()
    if not colors:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ —Ü–≤–µ—Ç–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Ü–≤–µ—Ç–∞ –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_COLORS_KB,
        )
        await state.clear()
        return
    await state.set_state(ManageLedModuleColorStates.waiting_for_color_name_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç, –∫–æ—Ç–æ—Ä—ã–π –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_colors_keyboard(colors),
    )


@dp.message(ManageLedModuleColorStates.waiting_for_color_name_to_delete)
async def process_remove_led_module_color(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    color_name = (message.text or "").strip()
    if not color_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Ü–≤–µ—Ç–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_led_module_color(color_name):
        await message.answer(f"üóë –¶–≤–µ—Ç ¬´{color_name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –¶–≤–µ—Ç ¬´{color_name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_colors_menu(message)


@dp.message(F.text == LED_MODULES_REMOVE_POWER_TEXT)
async def handle_remove_led_module_power_option(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    power_options = await fetch_led_module_power_options()
    if not power_options:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –º–æ—â–Ω–æ—Å—Ç–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_POWER_KB,
        )
        await state.clear()
        return
    await state.set_state(ManageLedModulePowerStates.waiting_for_power_value_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–æ—â–Ω–æ—Å—Ç—å, –∫–æ—Ç–æ—Ä—É—é –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_power_values_keyboard(power_options),
    )


@dp.message(ManageLedModulePowerStates.waiting_for_power_value_to_delete)
async def process_remove_led_module_power_option(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = (message.text or "").strip()
    if not value:
        await message.answer(
            "‚ö†Ô∏è –ó–Ω–∞—á–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if await delete_led_module_power_option(value):
        await message.answer(f"üóë –ú–æ—â–Ω–æ—Å—Ç—å ¬´{value}¬ª —É–¥–∞–ª–µ–Ω–∞.")
        await state.clear()
        await send_led_module_power_menu(message)
    else:
        await message.answer(
            f"‚ÑπÔ∏è –ú–æ—â–Ω–æ—Å—Ç—å ¬´{value}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ —Å–ø–∏—Å–∫–µ.",
            reply_markup=CANCEL_KB,
        )


@dp.message(F.text == LED_MODULES_REMOVE_VOLTAGE_TEXT)
async def handle_remove_led_module_voltage_option(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    voltage_options = await fetch_led_module_voltage_options()
    if not voltage_options:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_VOLTAGE_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManageLedModuleVoltageStates.waiting_for_voltage_value_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ, –∫–æ—Ç–æ—Ä–æ–µ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_voltage_values_keyboard(voltage_options),
    )


@dp.message(ManageLedModuleVoltageStates.waiting_for_voltage_value_to_delete)
async def process_remove_led_module_voltage_option(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = (message.text or "").strip()
    if not value:
        await message.answer(
            "‚ö†Ô∏è –ó–Ω–∞—á–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if await delete_led_module_voltage_option(value):
        await message.answer(f"üóë –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ ¬´{value}¬ª —É–¥–∞–ª–µ–Ω–æ.")
        await state.clear()
        await send_led_module_voltage_menu(message)
    else:
        await message.answer(
            f"‚ÑπÔ∏è –ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ ¬´{value}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –≤ —Å–ø–∏—Å–∫–µ.",
            reply_markup=CANCEL_KB,
        )


@dp.message(F.text == LED_MODULES_ADD_LENS_COUNT_TEXT)
async def handle_add_led_module_lens_count(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManageLedModuleLensStates.waiting_for_new_lens_count)
    existing = await fetch_led_module_lens_counts()
    existing_text = format_materials_list([str(value) for value in existing])
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ (—Ü–µ–ª–æ–µ –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω–æ–µ —á–∏—Å–ª–æ).\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleLensStates.waiting_for_new_lens_count)
async def process_new_led_module_lens_count(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –í–≤–µ–¥–∏—Ç–µ –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω–æ–µ —Ü–µ–ª–æ–µ —á–∏—Å–ª–æ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if await insert_led_module_lens_count(value):
        await message.answer(f"‚úÖ –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ ¬´{value}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–æ.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ ¬´{value}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_led_module_lens_menu(message)


@dp.message(F.text == LED_MODULES_REMOVE_LENS_COUNT_TEXT)
async def handle_remove_led_module_lens_count(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    existing = await fetch_led_module_lens_counts()
    if not existing:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –∫–æ–ª–∏—á–µ—Å—Ç–≤ –ª–∏–Ω–∑ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –∑–Ω–∞—á–µ–Ω–∏—è –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_LENS_KB,
        )
        await state.clear()
        return
    await state.set_state(ManageLedModuleLensStates.waiting_for_lens_count_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑, –∫–æ—Ç–æ—Ä–æ–µ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_lens_counts_keyboard(existing),
    )


@dp.message(ManageLedModuleLensStates.waiting_for_lens_count_to_delete)
async def process_remove_led_module_lens_count(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    value = parse_positive_integer(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –í–≤–µ–¥–∏—Ç–µ –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω–æ–µ —Ü–µ–ª–æ–µ —á–∏—Å–ª–æ. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    if await delete_led_module_lens_count(value):
        await message.answer(f"üóë –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ ¬´{value}¬ª —É–¥–∞–ª–µ–Ω–æ.")
        await state.clear()
        await send_led_module_lens_menu(message)
    else:
        await message.answer(
            f"‚ÑπÔ∏è –ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –ª–∏–Ω–∑ ¬´{value}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –≤ —Å–ø–∏—Å–∫–µ.",
            reply_markup=CANCEL_KB,
        )


@dp.message(F.text == LED_MODULES_ADD_SERIES_TEXT)
async def handle_add_led_module_series(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_led_module_manufacturers()
    if not manufacturers:
        await state.clear()
        await message.answer(
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π, —á—Ç–æ–±—ã —É–∫–∞–∑—ã–≤–∞—Ç—å —Å–µ—Ä–∏–∏.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB,
        )
        return
    await state.set_state(
        ManageLedModuleSeriesStates.waiting_for_manufacturer_for_new_series
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –¥–ª—è –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(ManageLedModuleSeriesStates.waiting_for_manufacturer_for_new_series)
async def process_choose_led_module_manufacturer_for_series(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    manufacturer_name = (message.text or "").strip()
    manufacturer = await get_led_module_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        manufacturers = await fetch_led_module_manufacturers()
        if not manufacturers:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, —á—Ç–æ–±—ã —É–∫–∞–∑–∞—Ç—å —Å–µ—Ä–∏—é.",
                reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturers),
        )
        return
    existing_series = await fetch_led_module_series_by_manufacturer(
        manufacturer["name"]
    )
    formatted_series = format_series_list(existing_series)
    await state.update_data(selected_manufacturer=manufacturer["name"])
    await state.set_state(ManageLedModuleSeriesStates.waiting_for_new_series_name)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –Ω–æ–≤–æ–π —Å–µ—Ä–∏–∏.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{formatted_series}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManageLedModuleSeriesStates.waiting_for_new_series_name)
async def process_new_led_module_series(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    series_name = (message.text or "").strip()
    if not series_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Å–µ—Ä–∏–∏ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    data = await state.get_data()
    manufacturer_name = data.get("selected_manufacturer")
    if not manufacturer_name:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
        await send_led_modules_settings_overview(message)
        return
    status = await insert_led_module_series(manufacturer_name, series_name)
    if status == "manufacturer_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –û–±–Ω–æ–≤–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –∏ –ø–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
    elif status == "already_exists":
        await message.answer(
            f"‚ÑπÔ∏è –°–µ—Ä–∏—è ¬´{series_name}¬ª —É–∂–µ —É–∫–∞–∑–∞–Ω–∞ –¥–ª—è ¬´{manufacturer_name}¬ª."
        )
    elif status == "inserted":
        await message.answer(
            f"‚úÖ –°–µ—Ä–∏—è ¬´{series_name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–∞ –¥–ª—è ¬´{manufacturer_name}¬ª."
        )
    else:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –¥–æ–±–∞–≤–∏—Ç—å —Å–µ—Ä–∏—é. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ."
        )
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == LED_MODULES_REMOVE_SERIES_TEXT)
async def handle_remove_led_module_series(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_led_module_manufacturers_with_series()
    manufacturers_with_series = [
        item["name"] for item in manufacturers if item.get("series")
    ]
    if not manufacturers_with_series:
        await state.clear()
        await message.answer(
            "–î–ª—è —É–¥–∞–ª–µ–Ω–∏—è –ø–æ–∫–∞ –Ω–µ—Ç –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã—Ö —Å–µ—Ä–∏–π.",
            reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB,
        )
        return
    await state.set_state(
        ManageLedModuleSeriesStates.waiting_for_manufacturer_for_series_deletion
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, —É –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å —Å–µ—Ä–∏—é:",
        reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
    )


@dp.message(ManageLedModuleSeriesStates.waiting_for_manufacturer_for_series_deletion)
async def process_choose_led_module_manufacturer_for_series_deletion(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    manufacturer_name = (message.text or "").strip()
    manufacturer = await get_led_module_manufacturer_by_name(manufacturer_name)
    if manufacturer is None:
        manufacturers = await fetch_led_module_manufacturers_with_series()
        manufacturers_with_series = [
            item["name"] for item in manufacturers if item.get("series")
        ]
        if not manufacturers_with_series:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ —Å–µ—Ä–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏–∏, —á—Ç–æ–±—ã –∏—Ö —É–¥–∞–ª–∏—Ç—å.",
                reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB,
            )
            return
        await message.answer(
            "‚ö†Ô∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
        )
        return
    series = await fetch_led_module_series_by_manufacturer(manufacturer["name"])
    if not series:
        manufacturers = await fetch_led_module_manufacturers_with_series()
        manufacturers_with_series = [
            item["name"] for item in manufacturers if item.get("series")
        ]
        if not manufacturers_with_series:
            await state.clear()
            await message.answer(
                "–°–ø–∏—Å–æ–∫ —Å–µ—Ä–∏–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ —Å–µ—Ä–∏–∏, —á—Ç–æ–±—ã –∏—Ö —É–¥–∞–ª–∏—Ç—å.",
                reply_markup=WAREHOUSE_SETTINGS_LED_MODULES_SERIES_KB,
            )
            return
        await message.answer(
            "‚ÑπÔ∏è –£ –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –Ω–µ—Ç —Å–µ—Ä–∏–π. –í—ã–±–µ—Ä–∏—Ç–µ –¥—Ä—É–≥–æ–≥–æ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è.",
            reply_markup=build_manufacturers_keyboard(manufacturers_with_series),
        )
        return
    await state.update_data(selected_manufacturer=manufacturer["name"])
    await state.set_state(ManageLedModuleSeriesStates.waiting_for_series_name_to_delete)
    await message.answer(
        f"–í—ã–±–µ—Ä–∏—Ç–µ —Å–µ—Ä–∏—é –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è —É ¬´{manufacturer['name']}¬ª:",
        reply_markup=build_series_keyboard(series),
    )


@dp.message(ManageLedModuleSeriesStates.waiting_for_series_name_to_delete)
async def process_remove_led_module_series(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    series_name = (message.text or "").strip()
    if not series_name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ —Å–µ—Ä–∏–∏ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    data = await state.get_data()
    manufacturer_name = data.get("selected_manufacturer")
    if not manufacturer_name:
        await state.clear()
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
        await send_led_modules_settings_overview(message)
        return
    status = await delete_led_module_series(manufacturer_name, series_name)
    if status == "manufacturer_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –û–±–Ω–æ–≤–∏—Ç–µ —Å–ø–∏—Å–æ–∫ –∏ –ø–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞."
        )
    elif status == "deleted":
        await message.answer(
            f"üóë –°–µ—Ä–∏—è ¬´{series_name}¬ª —É–¥–∞–ª–µ–Ω–∞ —É ¬´{manufacturer_name}¬ª."
        )
    else:
        await message.answer(
            f"‚ÑπÔ∏è –°–µ—Ä–∏—è ¬´{series_name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ —É ¬´{manufacturer_name}¬ª."
        )
    await state.clear()
    await send_led_modules_settings_overview(message)


@dp.message(F.text == POWER_SUPPLIES_ADD_MANUFACTURER_TEXT)
async def handle_add_power_supply_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManagePowerSupplyManufacturerStates.waiting_for_new_manufacturer_name
    )
    manufacturers = await fetch_power_supply_manufacturers()
    existing_text = format_materials_list(manufacturers)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è –±–ª–æ–∫–æ–≤ –ø–∏—Ç–∞–Ω–∏—è.\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManagePowerSupplyManufacturerStates.waiting_for_new_manufacturer_name)
async def process_new_power_supply_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_power_supply_manufacturer(name):
        await message.answer(f"‚úÖ –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_power_supplies_settings_overview(message)


@dp.message(F.text == POWER_SUPPLIES_REMOVE_MANUFACTURER_TEXT)
async def handle_remove_power_supply_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    manufacturers = await fetch_power_supply_manufacturers()
    if not manufacturers:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª–µ–π –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_POWER_SUPPLIES_MANUFACTURERS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManagePowerSupplyManufacturerStates.waiting_for_manufacturer_name_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—è, –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_manufacturers_keyboard(manufacturers),
    )


@dp.message(
    ManagePowerSupplyManufacturerStates.waiting_for_manufacturer_name_to_delete
)
async def process_remove_power_supply_manufacturer(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_power_supply_manufacturer(name):
        await message.answer(f"üóë –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ü—Ä–æ–∏–∑–≤–æ–¥–∏—Ç–µ–ª—å ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_power_supplies_settings_overview(message)


@dp.message(F.text == WAREHOUSE_SETTINGS_BACK_TO_ELECTRICS_TEXT)
async def handle_back_to_electrics_settings(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_electrics_settings_overview(message)


@dp.message(F.text == "üì¶ –ú–∞—Ç–µ—Ä–∏–∞–ª")
async def handle_plastic_materials_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ —Å –º–∞—Ç–µ—Ä–∏–∞–ª–∞–º–∏:",
        reply_markup=WAREHOUSE_SETTINGS_PLASTIC_MATERIALS_KB,
    )


@dp.message(F.text == "üìè –¢–æ–ª—â–∏–Ω–∞")
async def handle_plastic_thickness_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ —Å —Ç–æ–ª—â–∏–Ω–∞–º–∏:",
        reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
    )


@dp.message(F.text == "üé® –¶–≤–µ—Ç")
async def handle_plastic_colors_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ —Å —Ü–≤–µ—Ç–∞–º–∏:",
        reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
    )


@dp.message(F.text == "üè∑Ô∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")
async def handle_plastic_storage_menu(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_storage_locations_overview(message)


@dp.message(F.text == "‚¨ÖÔ∏è –ù–∞–∑–∞–¥ –∫ –ø–ª–∞—Å—Ç–∏–∫—É")
async def handle_back_to_plastic_settings(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–∞—Ç–µ—Ä–∏–∞–ª")
async def handle_add_plastic_material_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(ManagePlasticMaterialStates.waiting_for_new_material_name)
    materials = await fetch_plastic_material_types()
    existing_text = format_materials_list(materials)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–∞—Ç–µ—Ä–∏–∞–ª–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä, –î–∏–±–æ–Ω–¥, –ê–∫—Ä–∏–ª, –ü–í–•).\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_new_material_name)
async def process_new_plastic_material(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_plastic_material_type(name):
        await message.answer(f"‚úÖ –ú–∞—Ç–µ—Ä–∏–∞–ª ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å –º–∞—Ç–µ—Ä–∏–∞–ª")
async def handle_remove_plastic_material_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    materials = await fetch_plastic_material_types()
    if not materials:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_MATERIALS_KB,
        )
        await state.clear()
        return
    await state.set_state(ManagePlasticMaterialStates.waiting_for_material_name_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, –∫–æ—Ç–æ—Ä—ã–π –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_materials_keyboard(materials),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_material_name_to_delete)
async def process_remove_plastic_material(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_plastic_material_type(name):
        await message.answer(f"üóë –ú–∞—Ç–µ—Ä–∏–∞–ª ¬´{name}¬ª —É–¥–∞–ª—ë–Ω.")
    else:
        await message.answer(f"‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Å–ø–∏—Å–∫–µ.")
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")
async def handle_add_storage_location_button(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    await state.set_state(
        ManagePlasticMaterialStates.waiting_for_new_storage_location_name
    )
    locations = await fetch_plastic_storage_locations()
    existing_text = format_storage_locations_list(locations)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–µ—Å—Ç–∞ —Ö—Ä–∞–Ω–µ–Ω–∏—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, –ü–æ–ª–∫–∞ –ê1, –°—Ç–µ–ª–ª–∞–∂ 3).\n\n"
        f"–£–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω—ã:\n{existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_new_storage_location_name)
async def process_new_storage_location(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await insert_plastic_storage_location(name):
        await message.answer(f"‚úÖ –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –¥–æ–±–∞–≤–ª–µ–Ω–æ.")
    else:
        await message.answer(
            f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–∂–µ –µ—Å—Ç—å –≤ —Å–ø–∏—Å–∫–µ."
        )
    await state.clear()
    await send_storage_locations_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è")
async def handle_remove_storage_location_button(
    message: Message, state: FSMContext
) -> None:
    if not await ensure_admin_access(message, state):
        return
    locations = await fetch_plastic_storage_locations()
    if not locations:
        await message.answer(
            "–°–ø–∏—Å–æ–∫ –º–µ—Å—Ç —Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—É—Å—Ç. –î–æ–±–∞–≤—å—Ç–µ –º–µ—Å—Ç–∞ –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_STORAGE_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManagePlasticMaterialStates.waiting_for_storage_location_to_delete
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è, –∫–æ—Ç–æ—Ä–æ–µ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_storage_locations_keyboard(locations),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_storage_location_to_delete)
async def process_remove_storage_location(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    if await delete_plastic_storage_location(name):
        await message.answer(f"üóë –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª —É–¥–∞–ª–µ–Ω–æ.")
    else:
        await message.answer(
            f"‚ÑπÔ∏è –ú–µ—Å—Ç–æ —Ö—Ä–∞–Ω–µ–Ω–∏—è ¬´{name}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω–æ –≤ —Å–ø–∏—Å–∫–µ."
        )
    await state.clear()
    await send_storage_locations_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É")
async def handle_add_thickness_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    materials = await fetch_plastic_material_types()
    if not materials:
        await message.answer(
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, —á—Ç–æ–±—ã –º–æ–∂–Ω–æ –±—ã–ª–æ —É–∫–∞–∑–∞—Ç—å —Ç–æ–ª—â–∏–Ω—ã.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManagePlasticMaterialStates.waiting_for_material_name_to_add_thickness
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, –¥–ª—è –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É:",
        reply_markup=build_materials_keyboard(materials),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_material_name_to_add_thickness)
async def process_add_thickness_material_selection(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    materials = await fetch_plastic_material_types()
    match = next((item for item in materials if item.lower() == name.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¢–∞–∫–æ–π –º–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(materials),
        )
        return
    await state.update_data(selected_material=match)
    await state.set_state(ManagePlasticMaterialStates.waiting_for_thickness_value_to_add)
    existing_thicknesses = await fetch_material_thicknesses(match)
    existing_text = format_thicknesses_list(existing_thicknesses)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É –≤ –º–∏–ª–ª–∏–º–µ—Ç—Ä–∞—Ö (–Ω–∞–ø—Ä–∏–º–µ—Ä, 3 –∏–ª–∏ 3.5).\n"
        "–î–æ–ø—É—Å—Ç–∏–º—ã –∑–Ω–∞—á–µ–Ω–∏—è —Å —Ç–æ—á–∫–æ–π –∏–ª–∏ –∑–∞–ø—è—Ç–æ–π, –º–æ–∂–Ω–æ —É–∫–∞–∑—ã–≤–∞—Ç—å '–º–º'.\n\n"
        f"–¢–µ–∫—É—â–∏–µ —Ç–æ–ª—â–∏–Ω—ã –¥–ª—è ¬´{match}¬ª: {existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_thickness_value_to_add)
async def process_add_thickness_value(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    data = await state.get_data()
    material = data.get("selected_material")
    if not material:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
        return
    value = parse_thickness_input(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ç–æ–ª—â–∏–Ω—É. –£–∫–∞–∂–∏—Ç–µ —á–∏—Å–ª–æ, –Ω–∞–ø—Ä–∏–º–µ—Ä 3 –∏–ª–∏ 3.5 –º–º.",
            reply_markup=CANCEL_KB,
        )
        return
    status = await insert_material_thickness(material, value)
    if status == "material_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    elif status == "exists":
        await message.answer(
            f"‚ÑπÔ∏è –¢–æ–ª—â–∏–Ω–∞ {format_thickness_value(value)} —É–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω–∞ –¥–ª—è ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    else:
        await message.answer(
            f"‚úÖ –¢–æ–ª—â–∏–Ω–∞ {format_thickness_value(value)} –¥–æ–±–∞–≤–ª–µ–Ω–∞ –¥–ª—è ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûï –î–æ–±–∞–≤–∏—Ç—å —Ü–≤–µ—Ç")
async def handle_add_color_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    materials = await fetch_plastic_material_types()
    if not materials:
        await message.answer(
            "–°–Ω–∞—á–∞–ª–∞ –¥–æ–±–∞–≤—å—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã, —á—Ç–æ–±—ã —É–∫–∞–∑–∞—Ç—å –¥–ª—è –Ω–∏—Ö —Ü–≤–µ—Ç–∞.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
        await state.clear()
        return
    await state.set_state(ManagePlasticMaterialStates.waiting_for_material_name_to_add_color)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, –¥–ª—è –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ –¥–æ–±–∞–≤–∏—Ç—å —Ü–≤–µ—Ç:",
        reply_markup=build_materials_keyboard(materials),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_material_name_to_add_color)
async def process_add_color_material_selection(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    materials = await fetch_plastic_material_types()
    match = next((item for item in materials if item.lower() == name.lower()), None)
    if match is None:
        await message.answer(
            "‚ÑπÔ∏è –¢–∞–∫–æ–π –º–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –æ–¥–∏–Ω –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(materials),
        )
        return
    await state.update_data(selected_material=match)
    await state.set_state(ManagePlasticMaterialStates.waiting_for_color_value_to_add)
    existing_colors = await fetch_material_colors(match)
    existing_text = format_colors_list(existing_colors)
    await message.answer(
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ —Ü–≤–µ—Ç–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä, –ë–µ–ª—ã–π, –ß—ë—Ä–Ω—ã–π, –ö—Ä–∞—Å–Ω—ã–π).\n\n"
        f"–¢–µ–∫—É—â–∏–µ —Ü–≤–µ—Ç–∞ –¥–ª—è ¬´{match}¬ª: {existing_text}",
        reply_markup=CANCEL_KB,
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_color_value_to_add)
async def process_add_color_value(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    data = await state.get_data()
    material = data.get("selected_material")
    if not material:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
        return
    color = (message.text or "").strip()
    if not color:
        await message.answer(
            "‚ö†Ô∏è –¶–≤–µ—Ç –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –£–∫–∞–∂–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ —Ü–≤–µ—Ç–∞.",
            reply_markup=CANCEL_KB,
        )
        return
    status = await insert_material_color(material, color)
    if status == "material_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    elif status == "exists":
        await message.answer(
            f"‚ÑπÔ∏è –¶–≤–µ—Ç ¬´{color}¬ª —É–∂–µ –¥–æ–±–∞–≤–ª–µ–Ω –¥–ª—è ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    else:
        await message.answer(
            f"‚úÖ –¶–≤–µ—Ç ¬´{color}¬ª –¥–æ–±–∞–≤–ª–µ–Ω –¥–ª—è ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É")
async def handle_remove_thickness_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    materials = await fetch_materials_with_thicknesses()
    materials_with_data = [
        item["name"]
        for item in materials
        if item.get("thicknesses") and len(item["thicknesses"]) > 0
    ]
    if not materials_with_data:
        await message.answer(
            "–ü–æ–∫–∞ –Ω–µ—Ç –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ —Å —Ç–æ–ª—â–∏–Ω–∞–º–∏ –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManagePlasticMaterialStates.waiting_for_material_name_to_delete_thickness
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, —É –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å —Ç–æ–ª—â–∏–Ω—É:",
        reply_markup=build_materials_keyboard(materials_with_data),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_material_name_to_delete_thickness)
async def process_remove_thickness_material_selection(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    materials = await fetch_materials_with_thicknesses()
    match = next(
        (
            item
            for item in materials
            if item["name"].lower() == name.lower()
            and item.get("thicknesses")
            and len(item["thicknesses"]) > 0
        ),
        None,
    )
    if match is None:
        options = [
            item["name"]
            for item in materials
            if item.get("thicknesses") and len(item["thicknesses"]) > 0
        ]
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω –∏–ª–∏ —É –Ω–µ–≥–æ –Ω–µ—Ç —Ç–æ–ª—â–∏–Ω. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(options),
        )
        return
    await state.update_data(selected_material=match["name"])
    await state.set_state(ManagePlasticMaterialStates.waiting_for_thickness_value_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ç–æ–ª—â–∏–Ω—É, –∫–æ—Ç–æ—Ä—É—é –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_thickness_keyboard(match["thicknesses"]),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_thickness_value_to_delete)
async def process_remove_thickness_value(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    data = await state.get_data()
    material = data.get("selected_material")
    if not material:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
        return
    value = parse_thickness_input(message.text or "")
    if value is None:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ç–æ–ª—â–∏–Ω—É. –£–∫–∞–∂–∏—Ç–µ —á–∏—Å–ª–æ, –Ω–∞–ø—Ä–∏–º–µ—Ä 3 –∏–ª–∏ 3.5 –º–º.",
            reply_markup=build_thickness_keyboard(await fetch_material_thicknesses(material)),
        )
        return
    status = await delete_material_thickness(material, value)
    if status == "material_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    elif status == "deleted":
        await message.answer(
            f"üóë –¢–æ–ª—â–∏–Ω–∞ {format_thickness_value(value)} —É–¥–∞–ª–µ–Ω–∞ —É ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    else:
        await message.answer(
            f"‚ÑπÔ∏è –¢–æ–ª—â–∏–Ω–∞ {format_thickness_value(value)} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ —É ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_THICKNESS_KB,
        )
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == "‚ûñ –£–¥–∞–ª–∏—Ç—å —Ü–≤–µ—Ç")
async def handle_remove_color_button(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    materials = await fetch_materials_with_thicknesses()
    materials_with_colors = [
        item["name"]
        for item in materials
        if item.get("colors") and len(item["colors"]) > 0
    ]
    if not materials_with_colors:
        await message.answer(
            "–ü–æ–∫–∞ –Ω–µ—Ç –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ —Å –¥–æ–±–∞–≤–ª–µ–Ω–Ω—ã–º–∏ —Ü–≤–µ—Ç–∞–º–∏ –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
        await state.clear()
        return
    await state.set_state(
        ManagePlasticMaterialStates.waiting_for_material_name_to_delete_color
    )
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ –º–∞—Ç–µ—Ä–∏–∞–ª, —É –∫–æ—Ç–æ—Ä–æ–≥–æ –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å —Ü–≤–µ—Ç:",
        reply_markup=build_materials_keyboard(materials_with_colors),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_material_name_to_delete_color)
async def process_remove_color_material_selection(
    message: Message, state: FSMContext
) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    name = (message.text or "").strip()
    if not name:
        await message.answer("‚ö†Ô∏è –ù–∞–∑–≤–∞–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.")
        return
    materials = await fetch_materials_with_thicknesses()
    match = next(
        (
            item
            for item in materials
            if item["name"].lower() == name.lower()
            and item.get("colors")
            and len(item["colors"]) > 0
        ),
        None,
    )
    if match is None:
        options = [
            item["name"]
            for item in materials
            if item.get("colors") and len(item["colors"]) > 0
        ]
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω –∏–ª–∏ —É –Ω–µ–≥–æ –Ω–µ—Ç —Ü–≤–µ—Ç–æ–≤. –í—ã–±–µ—Ä–∏—Ç–µ –∏–∑ —Å–ø–∏—Å–∫–∞.",
            reply_markup=build_materials_keyboard(options),
        )
        return
    await state.update_data(selected_material=match["name"])
    await state.set_state(ManagePlasticMaterialStates.waiting_for_color_value_to_delete)
    await message.answer(
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ü–≤–µ—Ç, –∫–æ—Ç–æ—Ä—ã–π –Ω—É–∂–Ω–æ —É–¥–∞–ª–∏—Ç—å:",
        reply_markup=build_colors_keyboard(match["colors"]),
    )


@dp.message(ManagePlasticMaterialStates.waiting_for_color_value_to_delete)
async def process_remove_color_value(message: Message, state: FSMContext) -> None:
    if await _process_cancel_if_requested(message, state):
        return
    data = await state.get_data()
    material = data.get("selected_material")
    if not material:
        await state.clear()
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –Ω–µ –Ω–∞–π–¥–µ–Ω. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –Ω–∞—á–∞—Ç—å –∑–∞–Ω–æ–≤–æ.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
        return
    color = (message.text or "").strip()
    if not color:
        await message.answer(
            "‚ö†Ô∏è –ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–æ–∑–Ω–∞—Ç—å —Ü–≤–µ—Ç. –£–∫–∞–∂–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ —Ü–≤–µ—Ç–∞.",
            reply_markup=build_colors_keyboard(await fetch_material_colors(material)),
        )
        return
    status = await delete_material_color(material, color)
    if status == "material_not_found":
        await message.answer(
            "‚ÑπÔ∏è –ú–∞—Ç–µ—Ä–∏–∞–ª –±–æ–ª—å—à–µ –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ —Å–Ω–æ–≤–∞.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    elif status == "deleted":
        await message.answer(
            f"üóë –¶–≤–µ—Ç ¬´{color}¬ª —É–¥–∞–ª—ë–Ω —É ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    else:
        await message.answer(
            f"‚ÑπÔ∏è –¶–≤–µ—Ç ¬´{color}¬ª –Ω–µ –Ω–∞–π–¥–µ–Ω —É ¬´{material}¬ª.",
            reply_markup=WAREHOUSE_SETTINGS_PLASTIC_COLORS_KB,
        )
    await state.clear()
    await send_plastic_settings_overview(message)


@dp.message(F.text == CANCEL_TEXT)
async def handle_cancel(message: Message, state: FSMContext) -> None:
    if not await ensure_admin_access(message, state):
        return
    current_state = await state.get_state()
    if current_state and current_state.startswith(AddUserStates.__name__):
        await _cancel_add_user_flow(message, state)
        return
    if current_state and current_state.startswith(
        ManageFilmManufacturerStates.__name__
    ):
        await state.clear()
        await send_film_settings_overview(message)
        return
    if current_state and current_state.startswith(
        ManageOrderTypeStates.__name__
    ):
        await state.clear()
        await send_order_type_settings_overview(message)
        return
    if current_state and current_state.startswith(
        ManageFilmSeriesStates.__name__
    ):
        await state.clear()
        await send_film_settings_overview(message)
        return
    if current_state and current_state.startswith(
        ManageFilmStorageStates.__name__
    ):
        await state.clear()
        await send_film_storage_overview(message)
        return
    if current_state and current_state.startswith(
        ManageLedStripManufacturerStates.__name__
    ):
        await state.clear()
        await send_led_strips_settings_overview(message)
        return
    if current_state and current_state.startswith(
        ManageLedModuleManufacturerStates.__name__
    ):
        await state.clear()
        await send_led_modules_settings_overview(message)
        return
    if current_state and current_state.startswith(
        ManageLedModuleSeriesStates.__name__
    ):
        await state.clear()
        await send_led_modules_settings_overview(message)
        return
    if current_state and current_state.startswith(
        AddWarehouseLedModuleStates.__name__
    ):
        await _cancel_add_led_module_flow(message, state)
        return
    if current_state and current_state.startswith(
        GenerateLedModuleStates.__name__
    ):
        await _cancel_generate_led_module_flow(message, state)
        return
    if current_state and current_state.startswith(
        ManageLedModuleVoltageStates.__name__
    ):
        await state.clear()
        await send_led_module_voltage_menu(message)
        return
    if current_state and current_state.startswith(
        ManagePowerSupplyManufacturerStates.__name__
    ):
        await state.clear()
        await send_power_supplies_settings_overview(message)
        return
    await state.clear()
    await send_plastic_settings_overview(message)


# === –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ (–¥–æ–±–∞–≤–ª–µ–Ω–∏–µ/–ø—Ä–æ—Å–º–æ—Ç—Ä) –º–æ–∂–Ω–æ –≤–µ—Ä–Ω—É—Ç—å —Å—é–¥–∞ –ø–æ–∑–∂–µ ===


async def main() -> None:
    """–ó–∞–ø—É—Å–∫–∞–µ—Ç –ø–æ–ª–ª–∏–Ω–≥ Telegram-–±–æ—Ç–∞."""
    bot = Bot(BOT_TOKEN)
    try:
        await dp.start_polling(bot)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
